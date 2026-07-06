import os
from pathlib import Path
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import MasterEntry, CuttingReport, CuttingReportPhoto, FinishingReportPhoto, EmbroideryReport, PrintingReport
from django.db.models import Q
import io
import zipfile

def get_exports_dir():
    exports_dir = Path(settings.BASE_DIR) / 'exports'
    exports_dir.mkdir(exist_ok=True)
    return exports_dir


def export_to_excel(since_date=None):
    """Generate/update the main FabricTrack Excel file with all data."""
    exports_dir = get_exports_dir()
    filepath = exports_dir / 'FabricTrack_Data.xlsx'

    wb = openpyxl.Workbook()

    # ── Sheet 1: Master Entries ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Master Entries'
    _style_sheet(ws1)

    headers1 = ['#', 'Date', 'Job Card Number', 'Created By', 'Created At']
    _write_header_row(ws1, headers1)

    me_qs = MasterEntry.objects.all()
    if since_date:
        me_qs = me_qs.filter(Q(created_at__gt=since_date) | Q(updated_at__gt=since_date))
    for i, entry in enumerate(me_qs.order_by('-date'), start=1):
        ws1.append([
            i,
            entry.date.strftime('%d-%b-%Y'),
            entry.job_card_number,
            entry.created_by.username if entry.created_by else '—',
            entry.created_at.strftime('%d-%b-%Y %H:%M'),
        ])

    _auto_width_and_style(ws1)

    # ── Sheet 2: Consolidated Cutting Reports ───────────────────────────────
    ws2 = wb.create_sheet('Cutting Reports')
    _style_sheet(ws2)

    headers2 = [
        '#', 'Date & Lot No.', 'Report Type', 'Cutting Master Name', 'Cutting Rate', 'Fabric', 'Item Name', 'Job Card No.',
        'S', 'M', 'L', 'XL', '2XL', '3XL', '4XL', 'Grand Total',
        'Submitted By', 'Submitted At',
        'Jobworker', 'Purpose', 'Job Work In Date', 'Job Work Out Date', 'Total Pcs short', 'Total Pcs', 'Any other Problem',
        'Design Jobwork', 'Job Work Rate',
        'Embroidery Worker', 'Embroidery Purpose', 'Embroidery In Date', 'Embroidery Out Date', 'Embroidery Pcs short', 'Embroidery Pcs', 'Embroidery Any other Problem',
        'Design Embroidery', 'Embroidery Rate',
        'Printing Worker', 'Printing Purpose', 'Printing In Date', 'Printing Out Date', 'Printing Pcs short', 'Printing Pcs', 'Printing Any other Problem',
        'Design Printing', 'Printing Rate',
        'Line in Date', 'Line out Date', 'Total Pcs', 'Rate', 'Description', 'Total Rate', 'Option 1',
        'Singleneedle Line in Date', 'Singleneedle Line out Date', 'Singleneedle Total Pcs', 'Singleneedle Rate', 'Singleneedle Description', 'Singleneedle Total Rate', 'Singleneedle Option 1',
        'Sewing Line in Date', 'Sewing Line out Date', 'Sewing Total Pcs', 'Sewing Rate', 'Sewing Description', 'Sewing Total Rate', 'Sewing Option 1',
        'Finishing Date', 'Finishing Total Pcs', 'Finishing Pcs Short', 'Finishing Pcs Packed', 'Finishing Green Tape', 'Finishing Red Tape', 'Finishing Blue Tape', 'Finishing Total Tape', 'Finishing Rate'
    ]
    _write_header_row(ws2, headers2)

    
    all_reports = CuttingReport.objects.select_related('master_entry', 'created_by').prefetch_related(
        'photos', 'color_details', 'embroidery_reports', 'printing_reports', 'stitching_reports', 'finishing_reports',
        'singleneedle_reports', 'sewing_reports'
    ).all()
    if since_date:
        all_reports = all_reports.filter(
            Q(created_at__gt=since_date) | Q(updated_at__gt=since_date) |
            Q(jobwork_reports__created_at__gt=since_date) | Q(jobwork_reports__updated_at__gt=since_date) |
            Q(stitching_reports__created_at__gt=since_date) | Q(stitching_reports__updated_at__gt=since_date) |
            Q(embroidery_reports__created_at__gt=since_date) | Q(embroidery_reports__updated_at__gt=since_date) |
            Q(printing_reports__created_at__gt=since_date) | Q(printing_reports__updated_at__gt=since_date) |
            Q(singleneedle_reports__created_at__gt=since_date) | Q(singleneedle_reports__updated_at__gt=since_date) |
            Q(sewing_reports__created_at__gt=since_date) | Q(sewing_reports__updated_at__gt=since_date) |
            Q(finishing_reports__created_at__gt=since_date) | Q(finishing_reports__updated_at__gt=since_date)
        ).distinct()
        
    all_reports = list(all_reports.order_by('-created_at'))

    for i, report in enumerate(all_reports, start=1):
        job_work = report.jobwork_reports.first()
        embroidery = report.embroidery_reports.first()
        printing = report.printing_reports.first()
        stitching = report.stitching_reports.first()
        singleneedle = report.singleneedle_reports.first()
        sewing = report.sewing_reports.first()
        finishing = report.finishing_reports.first()
            
        if job_work:
            jw_data = [
                job_work.jobworker,
                job_work.purpose,
                job_work.jobwork_in.strftime('%d-%b-%Y') if job_work.jobwork_in else '—',
                job_work.jobwork_out.strftime('%d-%b-%Y') if job_work.jobwork_out else '—',
                job_work.total_pcs_short if job_work.total_pcs_short is not None else '—',
                job_work.total_pcs if job_work.total_pcs is not None else '—',
                job_work.any_other_problem,
                job_work.design_jobwork or '—',
                float(job_work.total_rate) if job_work.total_rate else '—',
            ]
        else:
            jw_data = ['—'] * 9

        if embroidery:
            emb_data = [
                embroidery.embroidery_worker,
                embroidery.purpose,
                embroidery.embroidery_in.strftime('%d-%b-%Y') if embroidery.embroidery_in else '—',
                embroidery.embroidery_out.strftime('%d-%b-%Y') if embroidery.embroidery_out else '—',
                embroidery.total_pcs_short if embroidery.total_pcs_short is not None else '—',
                embroidery.total_pcs if embroidery.total_pcs is not None else '—',
                embroidery.any_other_problem,
                embroidery.design_embroidery or '—',
                float(embroidery.total_rate) if embroidery.total_rate else '—',
            ]
        else:
            emb_data = ['—'] * 9

        if printing:
            print_data = [
                printing.printing_worker,
                printing.purpose,
                printing.printing_in.strftime('%d-%b-%Y') if printing.printing_in else '—',
                printing.printing_out.strftime('%d-%b-%Y') if printing.printing_out else '—',
                printing.total_pcs_short if printing.total_pcs_short is not None else '—',
                printing.total_pcs if printing.total_pcs is not None else '—',
                printing.any_other_problem,
                printing.design_printing or '—',
                float(printing.total_rate) if printing.total_rate else '—',
            ]
        else:
            print_data = ['—'] * 9
            
        if stitching:
            stitch_data = [
                stitching.line_in_date.strftime('%d-%b-%Y') if stitching.line_in_date else '—',
                stitching.line_out_date.strftime('%d-%b-%Y') if stitching.line_out_date else '—',
                stitching.total_pcs if stitching.total_pcs is not None else '—',
                stitching.rate_name or '—',
                stitching.rate_description or '—',
                float(stitching.total_rate) if stitching.total_rate else '—',
                stitching.option_1 or '—',
            ]
        else:
            stitch_data = ['—'] * 7

        if singleneedle:
            singleneedle_data = [
                singleneedle.line_in_date.strftime('%d-%b-%Y') if singleneedle.line_in_date else '—',
                singleneedle.line_out_date.strftime('%d-%b-%Y') if singleneedle.line_out_date else '—',
                singleneedle.total_pcs if singleneedle.total_pcs is not None else '—',
                singleneedle.rate_name or '—',
                singleneedle.rate_description or '—',
                float(singleneedle.total_rate) if singleneedle.total_rate else '—',
                singleneedle.option_1 or '—',
            ]
        else:
            singleneedle_data = ['—'] * 7

        if sewing:
            sewing_data = [
                sewing.line_in_date.strftime('%d-%b-%Y') if sewing.line_in_date else '—',
                sewing.line_out_date.strftime('%d-%b-%Y') if sewing.line_out_date else '—',
                sewing.total_pcs if sewing.total_pcs is not None else '—',
                sewing.rate_name or '—',
                sewing.rate_description or '—',
                float(sewing.total_rate) if sewing.total_rate else '—',
                sewing.option_1 or '—',
            ]
        else:
            sewing_data = ['—'] * 7

        if finishing:
            finish_data = [
                finishing.date.strftime('%d-%b-%Y') if finishing.date else '—',
                finishing.total_pcs if finishing.total_pcs is not None else '—',
                finishing.total_pcs_short if finishing.total_pcs_short is not None else '—',
                finishing.total_pcs_packed if finishing.total_pcs_packed is not None else '—',
                finishing.green_tape if finishing.green_tape is not None else '—',
                finishing.red_tape if finishing.red_tape is not None else '—',
                finishing.blue_tape if finishing.blue_tape is not None else '—',
                finishing.total_tape if finishing.total_tape is not None else '—',
                float(finishing.total_rate) if finishing.total_rate else '—',
            ]
        else:
            finish_data = ['—'] * 9

        ws2.append([
            i,
            report.master_entry.date.strftime('%d-%b-%Y') + " | " + report.master_entry.job_card_number,
            report.get_report_type_display(),
            report.cutting_master_name or '—',
            float(report.cutting_rate) if report.cutting_rate else '—',
            report.fabric_type_quality,
            report.item_name,
            report.job_card_no,
            report.size_s,
            report.size_m,
            report.size_l,
            report.size_xl,
            report.size_2xl,
            report.size_3xl,
            report.size_4xl,
            report.total_pcs,
            report.created_by.username if report.created_by else '—',
            report.created_at.strftime('%d-%b-%Y %H:%M'),
        ] + jw_data + emb_data + print_data + stitch_data + singleneedle_data + sewing_data + finish_data)

    _auto_width_and_style(ws2)

    # ── Sheet 5: Stitching ────────────────────────────────────────────
    ws5 = wb.create_sheet('Stitching')
    _style_sheet(ws5)

    headers5 = [
        '#', 'P1 Date', 'Job Card Number', 'Stitching Master', 'Item Name', 'Line In Date', 'Line Out Date',
        'Total Pcs', 'Rate', 'Description', 'Total Rate', 'Option 1',
        'Submitted By', 'Submitted At'
    ]
    _write_header_row(ws5, headers5)

    from .models import StitchingReport
    p4_qs = StitchingReport.objects.select_related('cutting_report__master_entry', 'created_by').all()
    if since_date:
        p4_qs = p4_qs.filter(created_at__gt=since_date)
    for i, report in enumerate(p4_qs.order_by('-created_at'), start=1):
        ws5.append([
            i,
            report.cutting_report.master_entry.date.strftime('%d-%b-%Y'),
            report.job_card_no,
            report.master_name or '—',
            report.item_name,
            report.line_in_date.strftime('%d-%b-%Y') if report.line_in_date else '—',
            report.line_out_date.strftime('%d-%b-%Y') if report.line_out_date else '—',
            report.total_pcs if report.total_pcs is not None else '—',
            report.rate_name or '—',
            report.rate_description or '—',
            float(report.total_rate) if report.total_rate else '—',
            report.option_1 or '—',
            report.created_by.username if report.created_by else '—',
            report.created_at.strftime('%d-%b-%Y %H:%M'),
        ])

    _auto_width_and_style(ws5)

    # ── Sheet 6: Job Work ────────────────────────────────────────────
    ws6 = wb.create_sheet('Job Work')
    _style_sheet(ws6)

    headers6 = [
        '#', 'P1 Date', 'Job Card Number', 'Jobworker', 'Purpose',
        'In Date', 'Out Date', 'Any other Problem', 'Total Pcs short', 'Total Pcs',
        'Submitted By', 'Submitted At'
    ]
    _write_header_row(ws6, headers6)

    from .models import JobWorkReport
    p5_qs = JobWorkReport.objects.select_related('cutting_report__master_entry', 'created_by').all()
    if since_date:
        p5_qs = p5_qs.filter(created_at__gt=since_date)
    for i, report in enumerate(p5_qs.order_by('-created_at'), start=1):
        ws6.append([
            i,
            report.cutting_report.master_entry.date.strftime('%d-%b-%Y'),
            report.job_card_no,
            report.jobworker,
            report.purpose,
            report.jobwork_in.strftime('%d-%b-%Y') if report.jobwork_in else '—',
            report.jobwork_out.strftime('%d-%b-%Y') if report.jobwork_out else '—',
            report.any_other_problem,
            report.total_pcs_short if report.total_pcs_short is not None else '—',
            report.total_pcs if report.total_pcs is not None else '—',
            report.created_by.username if report.created_by else '—',
            report.created_at.strftime('%d-%b-%Y %H:%M'),
        ])

    _auto_width_and_style(ws6)

    # ── Sheet 7: Finishing Report ────────────────────────────────────────────
    ws7 = wb.create_sheet('Finishing Report')
    _style_sheet(ws7)

    headers7 = [
        '#', 'P1 Date', 'Lot No', 'Finishing Master', 'Date', 'Total Pcs', 'Total Pcs Short',
        'Total Pcs Packed', 'Green Tape', 'Red Tape', 'Blue Tape', 'Total Tape', 'Rate',
        'Submitted By', 'Submitted At'
    ]
    _write_header_row(ws7, headers7)

    from .models import FinishingReport
    p6_qs = FinishingReport.objects.select_related('cutting_report__master_entry', 'created_by').prefetch_related('photos').all()
    if since_date:
        p6_qs = p6_qs.filter(created_at__gt=since_date)
    for i, report in enumerate(p6_qs.order_by('-created_at'), start=1):
        ws7.append([
            i,
            report.cutting_report.master_entry.date.strftime('%d-%b-%Y'),
            report.lot_no,
            report.master_name or '—',
            report.date.strftime('%d-%b-%Y') if report.date else '—',
            report.total_pcs if report.total_pcs is not None else '—',
            report.total_pcs_short if report.total_pcs_short is not None else '—',
            report.total_pcs_packed if report.total_pcs_packed is not None else '—',
            report.green_tape if report.green_tape is not None else '—',
            report.red_tape if report.red_tape is not None else '—',
            report.blue_tape if report.blue_tape is not None else '—',
            report.total_tape if report.total_tape is not None else '—',
            float(report.total_rate) if report.total_rate else '—',
            report.created_by.username if report.created_by else '—',
            report.created_at.strftime('%d-%b-%Y %H:%M'),
        ])

    _auto_width_and_style(ws7)

    # ── Sheet 8: Embroidery ───────────────────────────────────────────
    ws8 = wb.create_sheet('Embroidery')
    _style_sheet(ws8)

    headers8 = [
        '#', 'P1 Date', 'Job Card Number', 'Embroidery Master', 'Embroidery Worker',
        'Purpose', 'In Date', 'Out Date', 'Any other Problem', 'Total Pcs short', 'Total Pcs',
        'Submitted By', 'Submitted At'
    ]
    _write_header_row(ws8, headers8)

    emb_qs = EmbroideryReport.objects.select_related('cutting_report__master_entry', 'created_by').all()
    if since_date:
        emb_qs = emb_qs.filter(created_at__gt=since_date)
    for i, report in enumerate(emb_qs.order_by('-created_at'), start=1):
        ws8.append([
            i,
            report.cutting_report.master_entry.date.strftime('%d-%b-%Y'),
            report.job_card_no,
            report.master_name or '—',
            report.embroidery_worker,
            report.purpose,
            report.embroidery_in.strftime('%d-%b-%Y') if report.embroidery_in else '—',
            report.embroidery_out.strftime('%d-%b-%Y') if report.embroidery_out else '—',
            report.any_other_problem,
            report.total_pcs_short if report.total_pcs_short is not None else '—',
            report.total_pcs if report.total_pcs is not None else '—',
            report.created_by.username if report.created_by else '—',
            report.created_at.strftime('%d-%b-%Y %H:%M'),
        ])

    _auto_width_and_style(ws8)

    # ── Sheet 9: Printing ─────────────────────────────────────────────
    ws9 = wb.create_sheet('Printing')
    _style_sheet(ws9)

    headers9 = [
        '#', 'P1 Date', 'Job Card Number', 'Printing Master', 'Printing Worker',
        'Purpose', 'In Date', 'Out Date', 'Any other Problem', 'Total Pcs short', 'Total Pcs',
        'Submitted By', 'Submitted At'
    ]
    _write_header_row(ws9, headers9)

    prt_qs = PrintingReport.objects.select_related('cutting_report__master_entry', 'created_by').all()
    if since_date:
        prt_qs = prt_qs.filter(created_at__gt=since_date)
    for i, report in enumerate(prt_qs.order_by('-created_at'), start=1):
        ws9.append([
            i,
            report.cutting_report.master_entry.date.strftime('%d-%b-%Y'),
            report.job_card_no,
            report.master_name or '—',
            report.printing_worker,
            report.purpose,
            report.printing_in.strftime('%d-%b-%Y') if report.printing_in else '—',
            report.printing_out.strftime('%d-%b-%Y') if report.printing_out else '—',
            report.any_other_problem,
            report.total_pcs_short if report.total_pcs_short is not None else '—',
            report.total_pcs if report.total_pcs is not None else '—',
            report.created_by.username if report.created_by else '—',
            report.created_at.strftime('%d-%b-%Y %H:%M'),
        ])

    _auto_width_and_style(ws9)

    # ── Sheet 10: Singleneedle ───────────────────────────────────────────
    ws10 = wb.create_sheet('Singleneedle')
    _style_sheet(ws10)

    headers10 = [
        '#', 'P1 Date', 'Job Card Number', 'Singleneedle Master', 'Item Name', 'Line In Date', 'Line Out Date',
        'Total Pcs', 'Rate', 'Description', 'Total Rate', 'Option 1',
        'Submitted By', 'Submitted At'
    ]
    _write_header_row(ws10, headers10)

    from .models import SingleneedleReport
    p9_qs = SingleneedleReport.objects.select_related('cutting_report__master_entry', 'created_by').all()
    if since_date:
        p9_qs = p9_qs.filter(created_at__gt=since_date)
    for i, report in enumerate(p9_qs.order_by('-created_at'), start=1):
        ws10.append([
            i,
            report.cutting_report.master_entry.date.strftime('%d-%b-%Y'),
            report.job_card_no,
            report.master_name or '—',
            report.item_name,
            report.line_in_date.strftime('%d-%b-%Y') if report.line_in_date else '—',
            report.line_out_date.strftime('%d-%b-%Y') if report.line_out_date else '—',
            report.total_pcs if report.total_pcs is not None else '—',
            report.rate_name or '—',
            report.rate_description or '—',
            float(report.total_rate) if report.total_rate else '—',
            report.option_1 or '—',
            report.created_by.username if report.created_by else '—',
            report.created_at.strftime('%d-%b-%Y %H:%M'),
        ])

    _auto_width_and_style(ws10)

    # ── Sheet 11: Sewing ─────────────────────────────────────────────────
    ws11 = wb.create_sheet('Sewing')
    _style_sheet(ws11)

    headers11 = [
        '#', 'P1 Date', 'Job Card Number', 'Sewing Master', 'Item Name', 'Line In Date', 'Line Out Date',
        'Total Pcs', 'Rate', 'Description', 'Total Rate', 'Option 1',
        'Submitted By', 'Submitted At'
    ]
    _write_header_row(ws11, headers11)

    from .models import SewingReport
    p10_qs = SewingReport.objects.select_related('cutting_report__master_entry', 'created_by').all()
    if since_date:
        p10_qs = p10_qs.filter(created_at__gt=since_date)
    for i, report in enumerate(p10_qs.order_by('-created_at'), start=1):
        ws11.append([
            i,
            report.cutting_report.master_entry.date.strftime('%d-%b-%Y'),
            report.job_card_no,
            report.master_name or '—',
            report.item_name,
            report.line_in_date.strftime('%d-%b-%Y') if report.line_in_date else '—',
            report.line_out_date.strftime('%d-%b-%Y') if report.line_out_date else '—',
            report.total_pcs if report.total_pcs is not None else '—',
            report.rate_name or '—',
            report.rate_description or '—',
            float(report.total_rate) if report.total_rate else '—',
            report.option_1 or '—',
            report.created_by.username if report.created_by else '—',
            report.created_at.strftime('%d-%b-%Y %H:%M'),
        ])

    _auto_width_and_style(ws11)

    wb.save(filepath)
    return filepath


# ── Helpers ─────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def _write_header_row(ws, headers):
    ws.append(headers)
    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30


def _style_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'


def _auto_width_and_style(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for i, cell in enumerate(col):
            # Apply borders and alignment to ALL cells
            cell.border = THIN_BORDER
            # We don't overwrite header alignment which is set explicitly
            if i > 0:
                cell.alignment = DATA_ALIGN
            
            try:
                if cell.value:
                    # Calculate max length for auto-width, split by newline if present
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        max_len = max(max_len, len(line))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def generate_backup_zip():
    """Create a ZIP containing the data Excel spreadsheet and all uploaded photos."""
    excel_path = export_to_excel()
    
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Add Excel sheet
        if os.path.exists(excel_path):
            zip_file.write(excel_path, arcname=excel_path.name)
            
        # Add Cutting Report Photos
        for photo in CuttingReportPhoto.objects.select_related('cutting_report').all():
            job_card = photo.cutting_report.job_card_no or "unknown"
            # Sanitize job card number for filename
            job_card_safe = "".join([c for c in job_card if c.isalnum() or c in (' ', '-', '_')]).strip()
            job_card_safe = job_card_safe.replace(' ', '_')
            sanitized_name = "".join([c for c in photo.photo_name if c.isalnum() or c in ('.', '_', '-')])
            filename = f"Photos/Cutting/cutting_photo_{photo.id}_jc_{job_card_safe}_{sanitized_name}"
            zip_file.writestr(filename, bytes(photo.photo_data))
            
        # Add Finishing Report Photos
        for photo in FinishingReportPhoto.objects.select_related('finishing_report').all():
            lot_no = photo.finishing_report.lot_no or "unknown"
            lot_no_safe = "".join([c for c in lot_no if c.isalnum() or c in (' ', '-', '_')]).strip()
            lot_no_safe = lot_no_safe.replace(' ', '_')
            sanitized_name = "".join([c for c in photo.photo_name if c.isalnum() or c in ('.', '_', '-')])
            filename = f"Photos/Finishing/finishing_photo_{photo.id}_lot_{lot_no_safe}_{sanitized_name}"
            zip_file.writestr(filename, bytes(photo.photo_data))
            
        # Add Stitching Report Photos
        from .models import StitchingReportPhoto
        for photo in StitchingReportPhoto.objects.select_related('stitching_report').all():
            job_card = photo.stitching_report.job_card_no or "unknown"
            job_card_safe = "".join([c for c in job_card if c.isalnum() or c in (' ', '-', '_')]).strip()
            job_card_safe = job_card_safe.replace(' ', '_')
            sanitized_name = "".join([c for c in photo.photo_name if c.isalnum() or c in ('.', '_', '-')])
            filename = f"Photos/Stitching/stitching_photo_{photo.id}_jc_{job_card_safe}_{sanitized_name}"
            zip_file.writestr(filename, bytes(photo.photo_data))

        # Add Embroidery Report Photos
        from .models import EmbroideryReportPhoto
        for photo in EmbroideryReportPhoto.objects.select_related('embroidery_report').all():
            job_card = photo.embroidery_report.job_card_no or "unknown"
            job_card_safe = "".join([c for c in job_card if c.isalnum() or c in (' ', '-', '_')]).strip()
            job_card_safe = job_card_safe.replace(' ', '_')
            sanitized_name = "".join([c for c in photo.photo_name if c.isalnum() or c in ('.', '_', '-')])
            filename = f"Photos/Embroidery/embroidery_photo_{photo.id}_jc_{job_card_safe}_{sanitized_name}"
            zip_file.writestr(filename, bytes(photo.photo_data))

        # Add Printing Report Photos
        from .models import PrintingReportPhoto
        for photo in PrintingReportPhoto.objects.select_related('printing_report').all():
            job_card = photo.printing_report.job_card_no or "unknown"
            job_card_safe = "".join([c for c in job_card if c.isalnum() or c in (' ', '-', '_')]).strip()
            job_card_safe = job_card_safe.replace(' ', '_')
            sanitized_name = "".join([c for c in photo.photo_name if c.isalnum() or c in ('.', '_', '-')])
            filename = f"Photos/Printing/printing_photo_{photo.id}_jc_{job_card_safe}_{sanitized_name}"
            zip_file.writestr(filename, bytes(photo.photo_data))

        # Add Singleneedle Report Photos
        from .models import SingleneedleReportPhoto
        for photo in SingleneedleReportPhoto.objects.select_related('singleneedle_report').all():
            job_card = photo.singleneedle_report.job_card_no or "unknown"
            job_card_safe = "".join([c for c in job_card if c.isalnum() or c in (' ', '-', '_')]).strip()
            job_card_safe = job_card_safe.replace(' ', '_')
            sanitized_name = "".join([c for c in photo.photo_name if c.isalnum() or c in ('.', '_', '-')])
            filename = f"Photos/Singleneedle/singleneedle_photo_{photo.id}_jc_{job_card_safe}_{sanitized_name}"
            zip_file.writestr(filename, bytes(photo.photo_data))

        # Add Sewing Report Photos
        from .models import SewingReportPhoto
        for photo in SewingReportPhoto.objects.select_related('sewing_report').all():
            job_card = photo.sewing_report.job_card_no or "unknown"
            job_card_safe = "".join([c for c in job_card if c.isalnum() or c in (' ', '-', '_')]).strip()
            job_card_safe = job_card_safe.replace(' ', '_')
            sanitized_name = "".join([c for c in photo.photo_name if c.isalnum() or c in ('.', '_', '-')])
            filename = f"Photos/Sewing/sewing_photo_{photo.id}_jc_{job_card_safe}_{sanitized_name}"
            zip_file.writestr(filename, bytes(photo.photo_data))
            
    zip_buffer.seek(0)
    return zip_buffer


from django.db.models import Sum, F, ExpressionWrapper, DecimalField
from django.db.models.functions import Coalesce

def calculate_master_earnings(master_name, start_date=None, end_date=None):
    """
    Sums earnings (total_pcs * rate) across all 8 department reports where master_name matches,
    optionally filtered by a date range.
    """
    from .models import (
        CuttingReport, StitchingReport, JobWorkReport, EmbroideryReport,
        PrintingReport, SingleneedleReport, SewingReport, FinishingReport
    )

    earnings = 0.0

    # 1. Cutting (uses cutting_rate, date is in master_entry)
    q_cut = CuttingReport.objects.filter(master_name=master_name)
    if start_date:
        q_cut = q_cut.filter(master_entry__date__gte=start_date)
    if end_date:
        q_cut = q_cut.filter(master_entry__date__lte=end_date)
    
    res = q_cut.annotate(
        val=ExpressionWrapper(F('total_pcs') * Coalesce(F('cutting_rate'), 0.0), output_field=DecimalField())
    ).aggregate(total=Sum('val'))
    earnings += float(res['total'] or 0.0)

    # Helper for standard reports (using total_rate)
    def get_standard_earnings(model_class, date_field):
        q = model_class.objects.filter(master_name=master_name)
        if start_date:
            q = q.filter(**{f"{date_field}__gte": start_date})
        if end_date:
            q = q.filter(**{f"{date_field}__lte": end_date})
            
        r = q.annotate(
            val=ExpressionWrapper(F('total_pcs') * Coalesce(F('total_rate'), 0.0), output_field=DecimalField())
        ).aggregate(total=Sum('val'))
        return float(r['total'] or 0.0)

    earnings += get_standard_earnings(StitchingReport, 'line_in_date')
    earnings += get_standard_earnings(JobWorkReport, 'jobwork_in')
    earnings += get_standard_earnings(EmbroideryReport, 'embroidery_in')
    earnings += get_standard_earnings(PrintingReport, 'printing_in')
    earnings += get_standard_earnings(SingleneedleReport, 'line_in_date')
    earnings += get_standard_earnings(SewingReport, 'line_in_date')
    earnings += get_standard_earnings(FinishingReport, 'date')

    return earnings

