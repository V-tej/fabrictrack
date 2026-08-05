import os
import re
import json
import time
from datetime import datetime, date, timedelta
from django.db.models import Q, Prefetch, Case, When, Value, IntegerField
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.core.paginator import Paginator
from django.utils.text import slugify
from django.utils import timezone

from .models import (
    MasterEntry, CuttingReport, CuttingReportPhoto, CuttingReportColorDetail,
    StitchingReport, StitchingReportPhoto, JobWorkReport, JobWorkReportPhoto,
    FinishingReport, FinishingReportPhoto, UserProfile, SystemSetting,
    JobCardRequirement, EmbroideryReport, PrintingReport, EmbroideryReportPhoto,
    PrintingReportPhoto, SingleneedleReport, SewingReport, SingleneedleReportPhoto,
    SewingReportPhoto, RateDefinition, AccessoriesRecord, AccessoriesItemEntry,
    ACCESSORIES_ITEMS, AccessoryCustomName, AccessoriesPhoto, MiscellaneousReport,
    MiscellaneousReportFile, JobWork1Report, JobWork1ReportPhoto, Sewing1Report, Sewing1ReportPhoto,
    ActivityLog, ChatChannel, ChatMessage, ChatTask, ChatTaskLabel, ChatReaction, ChatBookmark, ChatChannelRead
)
from .forms import (
    MasterEntryForm, CuttingReportForm, StitchingReportForm, JobWorkReportForm,
    FinishingReportForm, EmbroideryReportForm, PrintingReportForm,
    SingleneedleReportForm, SewingReportForm, MiscellaneousReportForm,
    JobWork1ReportForm, Sewing1ReportForm
)
from .utils import export_to_excel, generate_backup_zip


# ── Activity Log Helper ──────────────────────────────────────────────────────

def log_activity(user, action, department, job_card_no='', details=''):
    """Record a Create/Edit/Delete action in the ActivityLog."""
    ActivityLog.objects.create(
        user=user,
        action=action,
        department=department,
        job_card_no=str(job_card_no),
        details=details,
    )

EXCLUDED_FIELDS = {'id', 'created_at', 'updated_at', 'created_by', 'photo_data', 'photo_name', 'photo_content_type'}

def capture_snapshot(obj, fields=None):
    """Dynamically capture a dictionary of ALL field names and values for any model instance."""
    if not obj:
        return {}
    res = {}
    for field in obj._meta.get_fields():
        if field.is_relation and (field.many_to_many or field.one_to_many):
            continue
        field_name = field.name
        if field_name in EXCLUDED_FIELDS:
            continue
        try:
            val = getattr(obj, field_name, None)
        except Exception:
            continue
        
        if field_name in ('signature', 'signature_2'):
            res[field_name] = 'Signed' if val else 'No Signature'
        elif field.is_relation and val:
            res[field_name] = str(val)
        elif isinstance(val, (datetime, date)):
            res[field_name] = val.strftime('%Y-%m-%d %H:%M') if isinstance(val, datetime) else val.strftime('%Y-%m-%d')
        elif val is None or str(val).strip() in ('', 'None'):
            res[field_name] = '—'
        else:
            res[field_name] = str(val).strip()

    if hasattr(obj, 'photos') and hasattr(obj.photos, 'count'):
        try:
            res['photo_count'] = obj.photos.count()
        except Exception:
            res['photo_count'] = 0
    return res

def get_obj_summary(obj):
    """Return a comprehensive summary string of ALL set fields for CREATE / DELETE activity log entries."""
    snap = capture_snapshot(obj)
    parts = []
    for k, v in snap.items():
        if v not in ('—', 'No Signature', 0, '0', '0.00'):
            label = k.replace('_', ' ').title()
            parts.append(f"{label}: {v}")
    return ' | '.join(parts) if parts else 'Entry details recorded'

def build_diff(old_vals: dict, new_vals: dict) -> str:
    """Compare two dicts of field values and return human-readable change summary for ALL modified fields."""
    changes = []
    for key, old_v in old_vals.items():
        new_v = new_vals.get(key)
        if str(old_v) != str(new_v):
            if key == 'photo_count':
                diff_count = (new_v or 0) - (old_v or 0)
                if diff_count > 0:
                    changes.append(f"{diff_count} Photo(s) Added")
                elif diff_count < 0:
                    changes.append(f"{abs(diff_count)} Photo(s) Removed")
            elif key in ('signature', 'signature_2'):
                label = key.replace('_', ' ').title()
                if old_v in ('No Signature', '—') and new_v == 'Signed':
                    changes.append(f"{label} Added")
                elif old_v == 'Signed' and new_v in ('No Signature', '—'):
                    changes.append(f"{label} Removed")
                else:
                    changes.append(f"{label} Updated")
            else:
                label = key.replace('_', ' ').title()
                old_disp = '—' if old_v is None or str(old_v).strip() in ('', 'None') else old_v
                new_disp = '—' if new_v is None or str(new_v).strip() in ('', 'None') else new_v
                changes.append(f"{label}: {old_disp} → {new_disp}")
    return ' | '.join(changes) if changes else 'Updated record'

# ── Auth Views ──────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ── Dashboard ───────────────────────────────────────────────────────────────

@login_required
def dashboard_view(request):
    profile = getattr(request.user, 'profile', None)
    person_type = profile.person_type if profile else 'P1'

    context = {
        'person_type': person_type,
        'master_entries_count': MasterEntry.objects.count(),
        'cutting_reports_count': CuttingReport.objects.count(),
    }

    # Fetch recent reports based on role
    if person_type == 'ADMIN' or request.user.is_superuser:
        context['recent_reports'] = CuttingReport.objects.select_related(
            'master_entry', 'created_by'
        ).prefetch_related(
            Prefetch('photos', queryset=CuttingReportPhoto.objects.defer('photo_data'))
        )[:5]
        context['user_submissions_count'] = CuttingReport.objects.filter(created_by=request.user).count()

    elif person_type in ['P1', 'P2', 'P3']:
        context['recent_reports'] = CuttingReport.objects.filter(created_by=request.user).select_related(
            'master_entry', 'created_by'
        ).prefetch_related(
            Prefetch('photos', queryset=CuttingReportPhoto.objects.defer('photo_data'))
        )[:5]
        context['user_submissions_count'] = CuttingReport.objects.filter(created_by=request.user).count()

    elif person_type == 'P4':
        context['recent_reports'] = StitchingReport.objects.filter(created_by=request.user).select_related(
            'cutting_report__master_entry', 'created_by'
        )[:5]
        context['user_submissions_count'] = StitchingReport.objects.filter(created_by=request.user).count()

    elif person_type == 'P5':
        context['recent_reports'] = JobWorkReport.objects.filter(created_by=request.user).select_related(
            'cutting_report__master_entry', 'created_by'
        )[:5]
        context['user_submissions_count'] = JobWorkReport.objects.filter(created_by=request.user).count()

    elif person_type == 'P6':
        context['recent_reports'] = FinishingReport.objects.filter(created_by=request.user).select_related(
            'cutting_report__master_entry', 'created_by'
        ).prefetch_related(
            Prefetch('photos', queryset=FinishingReportPhoto.objects.defer('photo_data'))
        )[:5]
        context['user_submissions_count'] = FinishingReport.objects.filter(created_by=request.user).count()

    elif person_type == 'P7':
        context['recent_reports'] = EmbroideryReport.objects.filter(created_by=request.user).select_related(
            'cutting_report__master_entry', 'created_by'
        )[:5]
        context['user_submissions_count'] = EmbroideryReport.objects.filter(created_by=request.user).count()

    elif person_type == 'P8':
        context['recent_reports'] = PrintingReport.objects.filter(created_by=request.user).select_related(
            'cutting_report__master_entry', 'created_by'
        )[:5]
        context['user_submissions_count'] = PrintingReport.objects.filter(created_by=request.user).count()

    elif person_type == 'P9':
        context['recent_reports'] = SingleneedleReport.objects.filter(created_by=request.user).select_related(
            'cutting_report__master_entry', 'created_by'
        ).prefetch_related(
            Prefetch('photos', queryset=SingleneedleReportPhoto.objects.defer('photo_data'))
        )[:5]
        context['user_submissions_count'] = SingleneedleReport.objects.filter(created_by=request.user).count()

    elif person_type == 'P10':
        context['recent_reports'] = SewingReport.objects.filter(created_by=request.user).select_related(
            'cutting_report__master_entry', 'created_by'
        ).prefetch_related(
            Prefetch('photos', queryset=SewingReportPhoto.objects.defer('photo_data'))
        )[:5]
        context['user_submissions_count'] = SewingReport.objects.filter(created_by=request.user).count()

    elif person_type == 'P11':
        context['recent_reports'] = JobWork1Report.objects.filter(created_by=request.user).select_related(
            'cutting_report__master_entry', 'created_by'
        )[:5]
        context['user_submissions_count'] = JobWork1Report.objects.filter(created_by=request.user).count()

    elif person_type == 'P12':
        context['recent_reports'] = Sewing1Report.objects.filter(created_by=request.user).select_related(
            'cutting_report__master_entry', 'created_by'
        ).prefetch_related(
            Prefetch('photos', queryset=Sewing1ReportPhoto.objects.defer('photo_data'))
        )[:5]
        context['user_submissions_count'] = Sewing1Report.objects.filter(created_by=request.user).count()

    # Fetch pending tasks — ordered newest date first, then newest id first
    if person_type == 'ADMIN' or request.user.is_superuser:
        context['pending_tasks'] = JobCardRequirement.objects.filter(
            Q(requires_cutting__gt=0, is_cutting_done=False) |
            Q(requires_jobwork__gt=0, is_jobwork_done=False) |
            Q(requires_jobwork1__gt=0, is_jobwork1_done=False) |
            Q(requires_stitching__gt=0, is_stitching_done=False) |
            Q(requires_finishing__gt=0, is_finishing_done=False) |
            Q(requires_embroidery__gt=0, is_embroidery_done=False) |
            Q(requires_printing__gt=0, is_printing_done=False) |
            Q(requires_singleneedle__gt=0, is_singleneedle_done=False) |
            Q(requires_sewing__gt=0, is_sewing_done=False) |
            Q(requires_sewing1__gt=0, is_sewing1_done=False)
        ).order_by('-date', '-id')
    elif person_type in ['P1', 'P2', 'P3']:
        context['pending_tasks'] = JobCardRequirement.objects.filter(
            requires_cutting__gt=0, is_cutting_done=False
        ).order_by('-date', '-id')
    elif person_type == 'P4':
        context['pending_tasks'] = JobCardRequirement.objects.filter(
            requires_stitching__gt=0, is_stitching_done=False
        ).order_by('-date', '-id')
    elif person_type == 'P5':
        context['pending_tasks'] = JobCardRequirement.objects.filter(
            requires_jobwork__gt=0, is_jobwork_done=False
        ).order_by('-date', '-id')
    elif person_type == 'P6':
        context['pending_tasks'] = JobCardRequirement.objects.filter(
            requires_finishing__gt=0, is_finishing_done=False
        ).order_by('-date', '-id')
    elif person_type == 'P7':
        context['pending_tasks'] = JobCardRequirement.objects.filter(
            requires_embroidery__gt=0, is_embroidery_done=False
        ).order_by('-date', '-id')
    elif person_type == 'P8':
        context['pending_tasks'] = JobCardRequirement.objects.filter(
            requires_printing__gt=0, is_printing_done=False
        ).order_by('-date', '-id')
    elif person_type == 'P9':
        context['pending_tasks'] = JobCardRequirement.objects.filter(
            requires_singleneedle__gt=0, is_singleneedle_done=False
        ).order_by('-date', '-id')
    elif person_type == 'P10':
        context['pending_tasks'] = JobCardRequirement.objects.filter(
            requires_sewing__gt=0, is_sewing_done=False
        ).order_by('-date', '-id')
    elif person_type == 'P11':
        context['pending_tasks'] = JobCardRequirement.objects.filter(
            requires_jobwork1__gt=0, is_jobwork1_done=False
        ).order_by('-date', '-id')
    elif person_type == 'P12':
        context['pending_tasks'] = JobCardRequirement.objects.filter(
            requires_sewing1__gt=0, is_sewing1_done=False
        ).order_by('-date', '-id')
    else:
        context['pending_tasks'] = []

    # Completed job cards: all REQUIRED departments are Done
    # A job card is complete when every required dept (requires_X > 0) has is_X_done = True
    if person_type == 'ADMIN' or request.user.is_superuser:
        from django.db.models import Q as _Q
        completed_qs = JobCardRequirement.objects.filter(
            # Exclude any card that still has a pending required step
            ~_Q(requires_cutting__gt=0, is_cutting_done=False),
            ~_Q(requires_jobwork__gt=0, is_jobwork_done=False),
            ~_Q(requires_jobwork1__gt=0, is_jobwork1_done=False),
            ~_Q(requires_stitching__gt=0, is_stitching_done=False),
            ~_Q(requires_finishing__gt=0, is_finishing_done=False),
            ~_Q(requires_embroidery__gt=0, is_embroidery_done=False),
            ~_Q(requires_printing__gt=0, is_printing_done=False),
            ~_Q(requires_singleneedle__gt=0, is_singleneedle_done=False),
            ~_Q(requires_sewing__gt=0, is_sewing_done=False),
            ~_Q(requires_sewing1__gt=0, is_sewing1_done=False),
        ).filter(
            # Must have at least one required department (not a blank entry)
            Q(requires_cutting__gt=0) | Q(requires_jobwork__gt=0) | Q(requires_jobwork1__gt=0) |
            Q(requires_stitching__gt=0) | Q(requires_finishing__gt=0) |
            Q(requires_embroidery__gt=0) | Q(requires_printing__gt=0) |
            Q(requires_singleneedle__gt=0) | Q(requires_sewing__gt=0) | Q(requires_sewing1__gt=0)
        ).order_by('-updated_at')[:50]
        context['completed_tasks'] = completed_qs

    if person_type == 'ADMIN' or request.user.is_superuser:
        context['master_form'] = MasterEntryForm()

    return render(request, 'dashboard.html', context)



# ── Manage Masters (Admin Only) ──────────────────────────────────────────────

@login_required
def manage_masters(request):
    try:
        person_type = request.user.profile.person_type
    except Exception:
        person_type = 'ADMIN'
        
    if person_type != 'ADMIN' and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to view this page.')
        return redirect('dashboard')

    if request.method == 'POST':
        if 'add_master' in request.POST:
            name = request.POST.get('name', '').strip()
            department = request.POST.get('department', '').strip()
            upi_id = request.POST.get('upi_id', '').strip() or None
            article = request.POST.get('article', '').strip() or None
            master_id = request.POST.get('master_id', '').strip()

            # Read photo ONCE right here
            photo_bytes = None
            photo_mime = None
            if department == 'Vendor' and 'vendor_photo' in request.FILES:
                pf = request.FILES['vendor_photo']
                raw = pf.read()
                if raw:  # only if actual bytes were uploaded
                    photo_bytes = raw
                    photo_mime = pf.content_type or 'image/jpeg'

            if name and department:
                from .models import MasterName
                if master_id:
                    # --- EDIT existing record ---
                    master = MasterName.objects.filter(id=master_id).first()
                    if master:
                        master.name = name
                        master.department = department
                        master.upi_id = upi_id
                        master.article = article if department == 'Vendor' else None
                        if photo_bytes is not None:
                            master.photo = photo_bytes
                            master.photo_mime = photo_mime
                        master.save()
                        messages.success(request, f'Successfully updated master {name}!')
                else:
                    # --- CREATE or UPDATE existing ---
                    master_article = article if department == 'Vendor' else None
                    master, created = MasterName.objects.get_or_create(
                        name=name,
                        department=department,
                        article=master_article,
                        defaults={'upi_id': upi_id}
                    )
                    # Update fields whether created or not
                    update_kwargs = {'upi_id': upi_id}
                    if photo_bytes is not None:
                        update_kwargs['photo'] = photo_bytes
                        update_kwargs['photo_mime'] = photo_mime
                    for k, v in update_kwargs.items():
                        setattr(master, k, v)
                    master.save()
                    messages.success(request, f'Successfully added {name} to {department}!')
            return redirect('manage_masters')
            
        elif 'delete_master' in request.POST:
            master_id = request.POST.get('master_id')
            if master_id:
                from .models import MasterName
                master = MasterName.objects.filter(id=master_id).first()
                if master:
                    master.delete()
                    messages.success(request, 'Master name deleted successfully.')
            return redirect('manage_masters')

        elif 'add_rate' in request.POST:
            name = request.POST.get('rate_name')
            description = request.POST.get('rate_description')
            total_rate = request.POST.get('rate_total')
            rate_id = request.POST.get('rate_id')
            if name and description and total_rate:
                from .models import RateDefinition
                if rate_id:
                    rd = RateDefinition.objects.filter(id=rate_id).first()
                    if rd:
                        rd.name = name.strip()
                        rd.description = description.strip()
                        rd.total_rate = total_rate
                        rd.save()
                        messages.success(request, f'Successfully updated rate {name}!')
                else:
                    rd, created = RateDefinition.objects.get_or_create(
                        name=name.strip(),
                        defaults={'description': description.strip(), 'total_rate': total_rate}
                    )
                    if not created:
                        rd.description = description.strip()
                        rd.total_rate = total_rate
                        rd.save()
                        messages.success(request, f'Successfully updated rate {name}!')
                    else:
                        messages.success(request, f'Successfully added rate {name}!')
            return redirect('manage_masters')

        elif 'delete_rate' in request.POST:
            rate_id = request.POST.get('rate_id')
            if rate_id:
                from .models import RateDefinition
                rate = RateDefinition.objects.filter(id=rate_id).first()
                if rate:
                    rate.delete()
                    messages.success(request, 'Rate definition deleted successfully.')
            return redirect('manage_masters')

    from .models import MasterName, RateDefinition
    masters = MasterName.objects.exclude(department='Vendor').order_by('department', 'name')
    vendors = MasterName.objects.filter(department='Vendor').order_by('name', 'article')
    rate_definitions = RateDefinition.objects.all().order_by('name')
    context = {
        'masters': masters,
        'vendors': vendors,
        'rate_definitions': rate_definitions,
        'person_type': person_type
    }
    return render(request, 'manage_masters.html', context)

# ── Manage Users (Admin Only) ────────────────────────────────────────────────

@login_required
def manage_users(request):
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to view this page.')
        return redirect('dashboard')

    from django.contrib.auth.models import User
    from .models import MasterName
    users = User.objects.select_related('profile').order_by('username')
    masters = MasterName.objects.all().order_by('department', 'name')
    context = {
        'users': users,
        'masters': masters,
        'person_choices': UserProfile.PERSON_CHOICES if hasattr(UserProfile, 'PERSON_CHOICES') else [],
    }
    # Load PERSON_CHOICES from models
    from .models import PERSON_CHOICES
    context['person_choices'] = PERSON_CHOICES
    return render(request, 'manage_users.html', context)


@login_required
def add_user(request):
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    if request.method == 'POST':
        from django.contrib.auth.models import User
        from .models import PERSON_CHOICES, MasterName
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        person_type = request.POST.get('person_type', 'P1')
        is_superuser = request.POST.get('is_superuser') == 'on'
        linked_master_ids = request.POST.getlist('linked_masters')
        statement_password = request.POST.get('statement_password', '').strip()

        if not username or not password or not statement_password:
            messages.error(request, 'Username, password and statement password are required.')
            return redirect('manage_users')

        if User.objects.filter(username=username).exists():
            messages.error(request, f'Username "{username}" already exists.')
            return redirect('manage_users')

        user = User.objects.create_user(username=username, password=password)
        user.is_superuser = is_superuser
        user.is_staff = is_superuser
        user.save()

        profile, _ = UserProfile.objects.get_or_create(user=user, defaults={'person_type': person_type})
        profile.statement_password = statement_password if statement_password else None
        profile.plain_password = password
        if linked_master_ids:
            masters = MasterName.objects.filter(pk__in=linked_master_ids)
            profile.linked_masters.set(masters)
        else:
            profile.linked_masters.clear()
        profile.save()

        messages.success(request, f'User "{username}" created successfully!')
    return redirect('manage_users')


@login_required
def delete_user(request, user_id):
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    from django.contrib.auth.models import User
    if request.method == 'POST':
        target_user = get_object_or_404(User, pk=user_id)
        if target_user == request.user:
            messages.error(request, 'You cannot delete your own account.')
        else:
            uname = target_user.username
            target_user.delete()
            messages.success(request, f'User "{uname}" deleted successfully.')
    return redirect('manage_users')


@login_required
def reset_user_password(request, user_id):
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    from django.contrib.auth.models import User
    if request.method == 'POST':
        target_user = get_object_or_404(User, pk=user_id)
        new_password = request.POST.get('new_password', '').strip()
        if not new_password:
            messages.error(request, 'Password cannot be empty.')
        else:
            target_user.set_password(new_password)
            target_user.save()
            
            profile, _ = UserProfile.objects.get_or_create(user=target_user)
            profile.plain_password = new_password
            profile.save()

            messages.success(request, f'Password for "{target_user.username}" reset successfully.')
    return redirect('manage_users')


@login_required
def update_user_role(request, user_id):
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    from django.contrib.auth.models import User
    from .models import MasterName
    if request.method == 'POST':
        target_user = get_object_or_404(User, pk=user_id)
        person_type = request.POST.get('person_type', 'P1')
        is_superuser = request.POST.get('is_superuser') == 'on'
        linked_master_ids = request.POST.getlist('linked_masters')
        statement_password = request.POST.get('statement_password', '').strip()

        if not statement_password:
            messages.error(request, 'Statement password is required.')
            return redirect('manage_users')

        profile, _ = UserProfile.objects.get_or_create(user=target_user)
        profile.person_type = person_type
        profile.statement_password = statement_password if statement_password else None
        if linked_master_ids:
            masters = MasterName.objects.filter(pk__in=linked_master_ids)
            profile.linked_masters.set(masters)
        else:
            profile.linked_masters.clear()
        profile.save()

        target_user.is_superuser = is_superuser
        target_user.is_staff = is_superuser
        target_user.save()
        messages.success(request, f'Role for "{target_user.username}" updated successfully.')
    return redirect('manage_users')


# ── Master Entry (Admin) ─────────────────────────────────────────────────────

@login_required
def create_master_entry(request):
    last_entry = MasterEntry.objects.order_by('-id').first()
    last_job_card = last_entry.job_card_number if last_entry else None

    if request.method == 'POST':
        form = MasterEntryForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.created_by = request.user
            entry.save()

            # Also create/update JobCardRequirement based on sequence fields
            req_data = {
                'date': entry.date,
                'requires_cutting':      form.cleaned_data.get('requires_cutting', 0),
                'requires_jobwork':      form.cleaned_data.get('requires_jobwork', 0),
                'requires_jobwork1':     form.cleaned_data.get('requires_jobwork1', 0),
                'requires_stitching':    form.cleaned_data.get('requires_stitching', 0),
                'requires_finishing':    form.cleaned_data.get('requires_finishing', 0),
                'requires_embroidery':   form.cleaned_data.get('requires_embroidery', 0),
                'requires_printing':     form.cleaned_data.get('requires_printing', 0),
                'requires_singleneedle': form.cleaned_data.get('requires_singleneedle', 0),
                'requires_sewing':       form.cleaned_data.get('requires_sewing', 0),
                'requires_sewing1':      form.cleaned_data.get('requires_sewing1', 0),
            }
            JobCardRequirement.objects.update_or_create(
                job_card_no=entry.job_card_number,
                defaults=req_data,
            )

            # Build detailed create log including sequences
            req_parts = [f"Job Card: {entry.job_card_number}", f"Date: {entry.date}"]
            _seq_labels = {
                'requires_cutting': 'Cutting', 'requires_jobwork': 'Job Work',
                'requires_jobwork1': 'Job Work 1', 'requires_stitching': 'Stitching',
                'requires_finishing': 'Finishing', 'requires_embroidery': 'Embroidery',
                'requires_printing': 'Printing', 'requires_singleneedle': 'Singleneedle',
                'requires_sewing': 'Sewing', 'requires_sewing1': 'Sewing 1',
            }
            for field, label in _seq_labels.items():
                val = req_data.get(field, 0)
                if val:
                    req_parts.append(f"{label} Seq: {val}")
            _create_details = ' | '.join(req_parts)

            messages.success(request, f'Master entry "{entry}" created successfully.')
            log_activity(request.user, 'CREATE', 'Master Entry', entry.job_card_number, _create_details)
            return redirect('dashboard')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = MasterEntryForm()
    return render(request, 'master_entry_form.html', {'form': form, 'last_job_card': last_job_card})

# ── Cutting Report (P1) ───────────────────────────────────────────────────────
@login_required
def cutting_report_view(request):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    is_p1 = hasattr(request.user, 'profile') and request.user.profile.person_type == 'P1'
    is_p2 = hasattr(request.user, 'profile') and request.user.profile.person_type == 'P2'
    is_p3 = hasattr(request.user, 'profile') and request.user.profile.person_type == 'P3'
    
    if not (is_admin or is_p1 or is_p2 or is_p3):
        raise PermissionDenied

    # A Job Card should only ever have ONE cutting report (by any master).
    # Once cut, it should disappear from the "New Report" dropdown for EVERYONE.
    master_entries_qs = MasterEntry.objects.filter(
        cutting_reports__isnull=True
    ).order_by('-date')

    # Enforce sequence flow: only show job cards where cutting step is enabled
    job_card_nos_for_cut = list(master_entries_qs.values_list('job_card_number', flat=True))
    reqs_for_cut = {r.job_card_no: r for r in JobCardRequirement.objects.filter(job_card_no__in=job_card_nos_for_cut)}
    allowed_cut_ids = [me.id for me in master_entries_qs
                       if not reqs_for_cut.get(me.job_card_number) or reqs_for_cut[me.job_card_number].is_cutting_enabled]
    master_entries_qs = master_entries_qs.filter(id__in=allowed_cut_ids)

    # Build a JSON map: { entry_id: job_card_number } for JS auto-fill
    master_entries_json = json.dumps({
        str(e.id): e.job_card_number for e in master_entries_qs
    })

    from .models import RateDefinition
    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    if request.method == 'POST':
        form = CuttingReportForm(request.POST, request.FILES)
        form.fields['master_entry'].queryset = master_entries_qs
        photos = request.FILES.getlist('photos')

        # Validate photo count
        if len(photos) > 5:
            messages.error(request, 'You can upload a maximum of 5 photos.')
            return render(request, 'person1_form.html', {
                'form': form,
                'master_entries': master_entries_qs,
                'master_entries_json': master_entries_json,
                'rate_definitions_json': rate_definitions_json,
                'is_admin': is_admin,
            })

        if len(photos) == 0:
            messages.error(request, 'Please upload at least one Job Card Photo.')
            return render(request, 'person1_form.html', {
                'form': form,
                'master_entries': master_entries_qs,
                'master_entries_json': master_entries_json,
                'rate_definitions_json': rate_definitions_json,
                'is_admin': is_admin,
            })

        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            
            # If the user is NOT an admin, forcefully override report_type based on their profile
            if not is_admin and hasattr(request.user, 'profile'):
                ptype = request.user.profile.person_type
                if ptype in ['P1', 'P2', 'P3']:
                    report.report_type = ptype

            if report.rate_definition:
                report.rate_name = report.rate_definition.name
                report.cutting_rate = report.rate_definition.total_rate
            
            report.save()

            # Save dynamic color size breakdown
            num_colors = report.total_colours
            if num_colors > 0:
                for i in range(1, num_colors + 1):
                    c_name = request.POST.get(f'color_name_{i}', f'C{i}')
                    c_s = int(request.POST.get(f'color_s_{i}') or 0)
                    c_m = int(request.POST.get(f'color_m_{i}') or 0)
                    c_l = int(request.POST.get(f'color_l_{i}') or 0)
                    c_xl = int(request.POST.get(f'color_xl_{i}') or 0)
                    c_2xl = int(request.POST.get(f'color_2xl_{i}') or 0)
                    c_3xl = int(request.POST.get(f'color_3xl_{i}') or 0)
                    c_4xl = int(request.POST.get(f'color_4xl_{i}') or 0)
                    c_weight = request.POST.get(f'color_weight_{i}') or 0.0
                    c_meters = request.POST.get(f'color_meters_{i}') or 0.0
                    CuttingReportColorDetail.objects.create(
                        cutting_report=report, color_name=c_name,
                        size_s=c_s, size_m=c_m, size_l=c_l, size_xl=c_xl,
                        size_2xl=c_2xl, size_3xl=c_3xl, size_4xl=c_4xl,
                        total_weight=c_weight, total_meters=c_meters
                    )

            # Save each photo to database
            for photo_file in photos:
                CuttingReportPhoto.objects.create(
                    cutting_report=report,
                    photo_data=photo_file.read(),
                    photo_name=photo_file.name,
                    photo_content_type=photo_file.content_type
                )

            # Mark pending task as done
            JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(is_cutting_done=True)

            messages.success(request, 'Cutting Report submitted successfully!')
            log_activity(request.user, 'CREATE', 'Cutting Report', report.job_card_no, get_obj_summary(report))
            return redirect('submission_list')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = CuttingReportForm()
        form.fields['master_entry'].queryset = master_entries_qs

    return render(request, 'person1_form.html', {
        'form': form,
        'master_entries': master_entries_qs,
        'master_entries_json': master_entries_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin,
    })


# ── P4: Stitching Report ───────────────────────────────────────────────

@login_required
def stitching_report_view(request):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )

    cutting_reports_qs = CuttingReport.objects.filter(
        stitching_reports__isnull=True,
        job_card_no__in=JobCardRequirement.objects.filter(requires_stitching__gt=0).values('job_card_no')
    ).select_related('master_entry').order_by('-created_at')

    # Enforce sequence flow: only show job cards where stitching step is enabled
    jc_nos_st = [cr.job_card_no for cr in cutting_reports_qs]
    reqs_st = {r.job_card_no: r for r in JobCardRequirement.objects.filter(job_card_no__in=jc_nos_st)}
    cutting_reports_qs = cutting_reports_qs.filter(
        job_card_no__in=[jc for jc in jc_nos_st if reqs_st.get(jc) and reqs_st[jc].is_stitching_enabled]
    )

    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })

    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    if request.method == 'POST':
        form = StitchingReportForm(request.POST, request.FILES)
        form.fields['cutting_report'].queryset = cutting_reports_qs

        if form.is_valid():
            photos = request.FILES.getlist('photos')
            if len(photos) > 5:
                messages.error(request, 'You can upload a maximum of 5 photos.')
                return redirect('stitching_report')

            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                report.rate_name = report.rate_definition.name
                report.rate_description = report.rate_definition.description
                report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos[:5]:
                StitchingReportPhoto.objects.create(
                    stitching_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Mark pending task: in-progress if only Line In, done if Line Out also filled
            if report.line_out_date:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_stitching_done=True, is_stitching_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_stitching_in_progress=True, is_stitching_done=False
                )

            messages.success(request, 'Stitching submitted successfully!')
            log_activity(request.user, 'CREATE', 'Stitching', report.job_card_no, get_obj_summary(report))
            return redirect('submission_list')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = StitchingReportForm()
        form.fields['cutting_report'].queryset = cutting_reports_qs

    return render(request, 'stitching_form.html', {
        'form': form,
        'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin,
    })


# ── P5: Job Work Report ───────────────────────────────────────────────

@login_required
def jobwork_report_view(request):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )

    cutting_reports_qs = CuttingReport.objects.filter(
        jobwork_reports__isnull=True,
        job_card_no__in=JobCardRequirement.objects.filter(requires_jobwork__gt=0).values('job_card_no')
    ).select_related('master_entry').order_by('-created_at')

    # Enforce sequence flow
    jc_nos_jw = [cr.job_card_no for cr in cutting_reports_qs]
    reqs_jw = {r.job_card_no: r for r in JobCardRequirement.objects.filter(job_card_no__in=jc_nos_jw)}
    cutting_reports_qs = cutting_reports_qs.filter(
        job_card_no__in=[jc for jc in jc_nos_jw if reqs_jw.get(jc) and reqs_jw[jc].is_jobwork_enabled]
    )

    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })

    if request.method == 'POST':
        form = JobWorkReportForm(request.POST, request.FILES)
        form.fields['cutting_report'].queryset = cutting_reports_qs

        if form.is_valid():
            photos = request.FILES.getlist('photos')
            if len(photos) == 0:
                messages.error(request, 'Please upload at least one Job Card Photo.')
                return render(request, 'jobwork_form.html', {
                    'form': form,
                    'cutting_reports': cutting_reports_qs,
                    'cutting_reports_json': cutting_reports_json,
                    'is_admin': is_admin,
                })
            if len(photos) > 5:
                messages.error(request, 'You can upload a maximum of 5 photos.')
                return render(request, 'jobwork_form.html', {
                    'form': form,
                    'cutting_reports': cutting_reports_qs,
                    'cutting_reports_json': cutting_reports_json,
                    'is_admin': is_admin,
                })

            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                if not report.total_rate:
                    report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos[:5]:
                JobWorkReportPhoto.objects.create(
                    job_work_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Mark pending task: in-progress if only In date, done if Out date also filled
            job_card_no = report.cutting_report.job_card_no
            if report.jobwork_out:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_jobwork_done=True, is_jobwork_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_jobwork_in_progress=True, is_jobwork_done=False
                )

            messages.success(request, 'Job Work submitted successfully!')
            log_activity(request.user, 'CREATE', 'Job Work', report.cutting_report.job_card_no, get_obj_summary(report))
            return redirect('submission_list')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = JobWorkReportForm()
        form.fields['cutting_report'].queryset = cutting_reports_qs

    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    return render(request, 'jobwork_form.html', {
        'form': form,
        'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin,
    })


@login_required
def jobwork1_report_view(request):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )

    cutting_reports_qs = CuttingReport.objects.filter(
        jobwork1_reports__isnull=True,
        job_card_no__in=JobCardRequirement.objects.filter(requires_jobwork1__gt=0).values('job_card_no')
    ).select_related('master_entry').order_by('-created_at')

    # Enforce sequence flow
    jc_nos_jw = [cr.job_card_no for cr in cutting_reports_qs]
    reqs_jw = {r.job_card_no: r for r in JobCardRequirement.objects.filter(job_card_no__in=jc_nos_jw)}
    cutting_reports_qs = cutting_reports_qs.filter(
        job_card_no__in=[jc for jc in jc_nos_jw if reqs_jw.get(jc) and reqs_jw[jc].is_jobwork1_enabled]
    )

    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })

    if request.method == 'POST':
        form = JobWork1ReportForm(request.POST, request.FILES)
        form.fields['cutting_report'].queryset = cutting_reports_qs

        if form.is_valid():
            photos = request.FILES.getlist('photos')
            if len(photos) == 0:
                messages.error(request, 'Please upload at least one Job Card Photo.')
                return render(request, 'jobwork_form.html', {
                    'form': form,
                    'cutting_reports': cutting_reports_qs,
                    'cutting_reports_json': cutting_reports_json,
                    'is_admin': is_admin,
                })
            if len(photos) > 5:
                messages.error(request, 'You can upload a maximum of 5 photos.')
                return render(request, 'jobwork_form.html', {
                    'form': form,
                    'cutting_reports': cutting_reports_qs,
                    'cutting_reports_json': cutting_reports_json,
                    'is_admin': is_admin,
                })

            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                if not report.total_rate:
                    report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos[:5]:
                JobWork1ReportPhoto.objects.create(
                    job_work1_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Mark pending task: in-progress if only In date, done if Out date also filled
            job_card_no = report.cutting_report.job_card_no
            if report.jobwork_out:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_jobwork1_done=True, is_jobwork1_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_jobwork1_in_progress=True, is_jobwork1_done=False
                )

            messages.success(request, 'Job Work 1 submitted successfully!')
            log_activity(request.user, 'CREATE', 'Job Work 1', report.cutting_report.job_card_no, get_obj_summary(report))
            return redirect('submission_list')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = JobWork1ReportForm()
        form.fields['cutting_report'].queryset = cutting_reports_qs

    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    return render(request, 'jobwork_form.html', {
        'form': form,
        'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin,
    })



# ── P7: Embroidery Report ──────────────────────────────────────────────

@login_required
def embroidery_report_view(request):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )

    cutting_reports_qs = CuttingReport.objects.filter(
        embroidery_reports__isnull=True,
        job_card_no__in=JobCardRequirement.objects.filter(requires_embroidery__gt=0).values('job_card_no')
    ).select_related('master_entry').order_by('-created_at')

    # Enforce sequence flow
    jc_nos_em = [cr.job_card_no for cr in cutting_reports_qs]
    reqs_em = {r.job_card_no: r for r in JobCardRequirement.objects.filter(job_card_no__in=jc_nos_em)}
    cutting_reports_qs = cutting_reports_qs.filter(
        job_card_no__in=[jc for jc in jc_nos_em if reqs_em.get(jc) and reqs_em[jc].is_embroidery_enabled]
    )

    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })

    if request.method == 'POST':
        form = EmbroideryReportForm(request.POST, request.FILES)
        form.fields['cutting_report'].queryset = cutting_reports_qs

        if form.is_valid():
            photos = request.FILES.getlist('photos')
            if len(photos) > 5:
                messages.error(request, 'You can upload a maximum of 5 photos.')
                return redirect('embroidery_report')

            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                if not report.total_rate:
                    report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos[:5]:
                EmbroideryReportPhoto.objects.create(
                    embroidery_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Mark pending task: in-progress if only In date, done if Out date also filled
            job_card_no = report.cutting_report.job_card_no
            if report.embroidery_out:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_embroidery_done=True, is_embroidery_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_embroidery_in_progress=True, is_embroidery_done=False
                )

            messages.success(request, 'Embroidery submitted successfully!')
            log_activity(request.user, 'CREATE', 'Embroidery', report.cutting_report.job_card_no, get_obj_summary(report))
            return redirect('submission_list')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = EmbroideryReportForm()
        form.fields['cutting_report'].queryset = cutting_reports_qs

    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    return render(request, 'embroidery_form.html', {
        'form': form,
        'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin,
    })


# ── P8: Printing Report ────────────────────────────────────────────────

@login_required
def printing_report_view(request):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )

    cutting_reports_qs = CuttingReport.objects.filter(
        printing_reports__isnull=True,
        job_card_no__in=JobCardRequirement.objects.filter(requires_printing__gt=0).values('job_card_no')
    ).select_related('master_entry').order_by('-created_at')

    # Enforce sequence flow
    jc_nos_pr = [cr.job_card_no for cr in cutting_reports_qs]
    reqs_pr = {r.job_card_no: r for r in JobCardRequirement.objects.filter(job_card_no__in=jc_nos_pr)}
    cutting_reports_qs = cutting_reports_qs.filter(
        job_card_no__in=[jc for jc in jc_nos_pr if reqs_pr.get(jc) and reqs_pr[jc].is_printing_enabled]
    )

    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })

    if request.method == 'POST':
        form = PrintingReportForm(request.POST, request.FILES)
        form.fields['cutting_report'].queryset = cutting_reports_qs

        if form.is_valid():
            photos = request.FILES.getlist('photos')
            if len(photos) > 5:
                messages.error(request, 'You can upload a maximum of 5 photos.')
                return redirect('printing_report')

            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                if not report.total_rate:
                    report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos[:5]:
                PrintingReportPhoto.objects.create(
                    printing_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Mark pending task: in-progress if only In date, done if Out date also filled
            job_card_no = report.cutting_report.job_card_no
            if report.printing_out:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_printing_done=True, is_printing_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_printing_in_progress=True, is_printing_done=False
                )

            messages.success(request, 'Printing submitted successfully!')
            log_activity(request.user, 'CREATE', 'Printing', report.cutting_report.job_card_no, get_obj_summary(report))
            return redirect('submission_list')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = PrintingReportForm()
        form.fields['cutting_report'].queryset = cutting_reports_qs

    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    return render(request, 'printing_form.html', {
        'form': form,
        'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin,
    })


# ── P9: Singleneedle Report ───────────────────────────────────────────

@login_required
def singleneedle_report_view(request):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )

    cutting_reports_qs = CuttingReport.objects.filter(
        singleneedle_reports__isnull=True,
        job_card_no__in=JobCardRequirement.objects.filter(requires_singleneedle__gt=0).values('job_card_no')
    ).select_related('master_entry').order_by('-created_at')

    # Enforce sequence flow
    jc_nos_sn = [cr.job_card_no for cr in cutting_reports_qs]
    reqs_sn = {r.job_card_no: r for r in JobCardRequirement.objects.filter(job_card_no__in=jc_nos_sn)}
    cutting_reports_qs = cutting_reports_qs.filter(
        job_card_no__in=[jc for jc in jc_nos_sn if reqs_sn.get(jc) and reqs_sn[jc].is_singleneedle_enabled]
    )

    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })

    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    if request.method == 'POST':
        form = SingleneedleReportForm(request.POST, request.FILES)
        form.fields['cutting_report'].queryset = cutting_reports_qs

        if form.is_valid():
            photos = request.FILES.getlist('photos')
            if len(photos) > 5:
                messages.error(request, 'You can upload a maximum of 5 photos.')
                return redirect('singleneedle_report')

            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                report.rate_name = report.rate_definition.name
                report.rate_description = report.rate_definition.description
                report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos[:5]:
                SingleneedleReportPhoto.objects.create(
                    singleneedle_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Mark pending task: in-progress if only Line In, done if Line Out also filled
            if report.line_out_date:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_singleneedle_done=True, is_singleneedle_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_singleneedle_in_progress=True, is_singleneedle_done=False
                )

            messages.success(request, 'Singleneedle submitted successfully!')
            log_activity(request.user, 'CREATE', 'Singleneedle', report.job_card_no, get_obj_summary(report))
            return redirect('submission_list')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = SingleneedleReportForm()
        form.fields['cutting_report'].queryset = cutting_reports_qs

    return render(request, 'singleneedle_form.html', {
        'form': form,
        'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin,
    })


# ── P10: Sewing Report ───────────────────────────────────────────────────

@login_required
def sewing_report_view(request):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )

    cutting_reports_qs = CuttingReport.objects.filter(
        sewing_reports__isnull=True,
        job_card_no__in=JobCardRequirement.objects.filter(requires_sewing__gt=0).values('job_card_no')
    ).select_related('master_entry').order_by('-created_at')

    # Enforce sequence flow
    jc_nos_sw = [cr.job_card_no for cr in cutting_reports_qs]
    reqs_sw = {r.job_card_no: r for r in JobCardRequirement.objects.filter(job_card_no__in=jc_nos_sw)}
    cutting_reports_qs = cutting_reports_qs.filter(
        job_card_no__in=[jc for jc in jc_nos_sw if reqs_sw.get(jc) and reqs_sw[jc].is_sewing_enabled]
    )

    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })

    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    if request.method == 'POST':
        form = SewingReportForm(request.POST, request.FILES)
        form.fields['cutting_report'].queryset = cutting_reports_qs

        if form.is_valid():
            photos = request.FILES.getlist('photos')
            if len(photos) > 5:
                messages.error(request, 'You can upload a maximum of 5 photos.')
                return redirect('sewing_report')

            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                report.rate_name = report.rate_definition.name
                report.rate_description = report.rate_definition.description
                report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos[:5]:
                SewingReportPhoto.objects.create(
                    sewing_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Mark pending task: in-progress if only Line In, done if Line Out also filled
            if report.line_out_date:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_sewing_done=True, is_sewing_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_sewing_in_progress=True, is_sewing_done=False
                )

            messages.success(request, 'Sewing submitted successfully!')
            log_activity(request.user, 'CREATE', 'Sewing', report.job_card_no, get_obj_summary(report))
            return redirect('submission_list')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = SewingReportForm()
        form.fields['cutting_report'].queryset = cutting_reports_qs

    return render(request, 'sewing_form.html', {
        'form': form,
        'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin,
    })


@login_required
def sewing1_report_view(request):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )

    cutting_reports_qs = CuttingReport.objects.filter(
        sewing1_reports__isnull=True,
        job_card_no__in=JobCardRequirement.objects.filter(requires_sewing1__gt=0).values('job_card_no')
    ).select_related('master_entry').order_by('-created_at')

    # Enforce sequence flow
    jc_nos_sw = [cr.job_card_no for cr in cutting_reports_qs]
    reqs_sw = {r.job_card_no: r for r in JobCardRequirement.objects.filter(job_card_no__in=jc_nos_sw)}
    cutting_reports_qs = cutting_reports_qs.filter(
        job_card_no__in=[jc for jc in jc_nos_sw if reqs_sw.get(jc) and reqs_sw[jc].is_sewing1_enabled]
    )

    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })

    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    if request.method == 'POST':
        form = Sewing1ReportForm(request.POST, request.FILES)
        form.fields['cutting_report'].queryset = cutting_reports_qs

        if form.is_valid():
            photos = request.FILES.getlist('photos')
            if len(photos) > 5:
                messages.error(request, 'You can upload a maximum of 5 photos.')
                return redirect('sewing1_report')

            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                report.rate_name = report.rate_definition.name
                report.rate_description = report.rate_definition.description
                report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos[:5]:
                Sewing1ReportPhoto.objects.create(
                    sewing1_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Mark pending task: in-progress if only Line In, done if Line Out also filled
            if report.line_out_date:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_sewing1_done=True, is_sewing1_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_sewing1_in_progress=True, is_sewing1_done=False
                )

            messages.success(request, 'Sewing 1 submitted successfully!')
            log_activity(request.user, 'CREATE', 'Sewing 1', report.job_card_no, get_obj_summary(report))
            return redirect('submission_list')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = Sewing1ReportForm()
        form.fields['cutting_report'].queryset = cutting_reports_qs

    return render(request, 'sewing_form.html', {
        'form': form,
        'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin,
    })



# ── P6: Finishing Report ───────────────────────────────────────────────

@login_required
def finishing_report_view(request):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )

    cutting_reports_qs = CuttingReport.objects.filter(
        finishing_reports__isnull=True,
        job_card_no__in=JobCardRequirement.objects.filter(requires_finishing__gt=0).values('job_card_no')
    ).select_related('master_entry').order_by('-created_at')

    # Enforce sequence flow
    jc_nos_fi = [cr.job_card_no for cr in cutting_reports_qs]
    reqs_fi = {r.job_card_no: r for r in JobCardRequirement.objects.filter(job_card_no__in=jc_nos_fi)}
    cutting_reports_qs = cutting_reports_qs.filter(
        job_card_no__in=[jc for jc in jc_nos_fi if reqs_fi.get(jc) and reqs_fi[jc].is_finishing_enabled]
    )

    # Build JSON map for auto-fill based on Cutting Report selection
    # We want to fill Date (from master_entry) and Lot No (from master_entry.job_card_number or cutting_report.job_card_no)
    # Since the user requested it comes from the cutting report, we'll map both:
    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'date': cr.master_entry.date.strftime('%Y-%m-%d'),
            'lot_no': cr.job_card_no,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })

    if request.method == 'POST':
        form = FinishingReportForm(request.POST, request.FILES)
        form.fields['cutting_report'].queryset = cutting_reports_qs
        photos = request.FILES.getlist('photos')

        if len(photos) > 5:
            messages.error(request, 'You can upload a maximum of 5 photos.')
            return render(request, 'finishing_form.html', {
                'form': form,
                'cutting_reports': cutting_reports_qs,
                'cutting_reports_json': cutting_reports_json,
                'is_admin': is_admin,
            })

        if len(photos) == 0:
            messages.error(request, 'Please upload at least one Job Card Photo.')
            return render(request, 'finishing_form.html', {
                'form': form,
                'cutting_reports': cutting_reports_qs,
                'cutting_reports_json': cutting_reports_json,
                'is_admin': is_admin,
            })

        rate_definitions = RateDefinition.objects.all()
        rate_definitions_json = json.dumps({
            str(r.id): {
                'name': r.name,
                'description': r.description,
                'total_rate': str(r.total_rate)
            } for r in rate_definitions
        })

        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                if not report.total_rate:
                    report.total_rate = report.rate_definition.total_rate
            report.save()

            for photo_file in photos:
                FinishingReportPhoto.objects.create(
                    finishing_report=report,
                    photo_data=photo_file.read(),
                    photo_name=photo_file.name,
                    photo_content_type=photo_file.content_type
                )



            # Mark pending task as done
            JobCardRequirement.objects.filter(job_card_no=report.cutting_report.job_card_no).update(is_finishing_done=True)

            messages.success(request, 'Finishing Report submitted successfully!')
            log_activity(request.user, 'CREATE', 'Finishing', report.cutting_report.job_card_no, get_obj_summary(report))
            return redirect('submission_list')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = FinishingReportForm()
        form.fields['cutting_report'].queryset = cutting_reports_qs

    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    return render(request, 'finishing_form.html', {
        'form': form,
        'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin,
    })


# ── Submissions List ──────────────────────────────────────────────────────────

@login_required
def submission_list_view(request):
    profile = getattr(request.user, 'profile', None)
    person_type = profile.person_type if profile else 'P1'
    filter_param = request.GET.get('filter')
    page_number = request.GET.get('page', '1')

    date_from_cutting = request.GET.get('date_from_cutting', '')
    date_to_cutting = request.GET.get('date_to_cutting', '')
    date_from_stitching = request.GET.get('date_from_stitching', '')
    date_to_stitching = request.GET.get('date_to_stitching', '')
    date_from_job_work = request.GET.get('date_from_job_work', '')
    date_to_job_work = request.GET.get('date_to_job_work', '')
    date_from_job_work1 = request.GET.get('date_from_job_work1', '')
    date_to_job_work1 = request.GET.get('date_to_job_work1', '')
    date_from_finishing = request.GET.get('date_from_finishing', '')
    date_to_finishing = request.GET.get('date_to_finishing', '')
    date_from_embroidery = request.GET.get('date_from_embroidery', '')
    date_to_embroidery = request.GET.get('date_to_embroidery', '')
    date_from_printing = request.GET.get('date_from_printing', '')
    date_to_printing = request.GET.get('date_to_printing', '')
    date_from_singleneedle = request.GET.get('date_from_singleneedle', '')
    date_to_singleneedle = request.GET.get('date_to_singleneedle', '')
    date_from_sewing = request.GET.get('date_from_sewing', '')
    date_to_sewing = request.GET.get('date_to_sewing', '')
    date_from_sewing1 = request.GET.get('date_from_sewing1', '')
    date_to_sewing1 = request.GET.get('date_to_sewing1', '')

    # Job card search — filters Cutting Reports by job_card_no
    jc_search = request.GET.get('jc', '').strip()

    # Base querysets with optimized prefetching to avoid N+1 queries
    # All users see all querysets, ordered by current user's first, then by date descending
    reports_qs = CuttingReport.objects.select_related(
        'master_entry', 'created_by'
    ).prefetch_related(
        Prefetch('photos', queryset=CuttingReportPhoto.objects.defer('photo_data'))
    ).annotate(
        is_mine=Case(
            When(created_by=request.user, then=Value(1)),
            default=Value(0),
            output_field=IntegerField()
        )
    ).order_by('-created_at')
    p4_qs = StitchingReport.objects.select_related(
        'cutting_report__master_entry', 'created_by'
    ).prefetch_related(
        Prefetch('photos', queryset=StitchingReportPhoto.objects.defer('photo_data'))
    ).annotate(
        is_mine=Case(
            When(created_by=request.user, then=Value(1)),
            default=Value(0),
            output_field=IntegerField()
        )
    ).order_by('-created_at')
    p5_qs = JobWorkReport.objects.select_related(
        'cutting_report__master_entry', 'created_by'
    ).annotate(
        is_mine=Case(
            When(created_by=request.user, then=Value(1)),
            default=Value(0),
            output_field=IntegerField()
        )
    ).order_by('-created_at')
    p6_qs = FinishingReport.objects.select_related(
        'cutting_report__master_entry', 'created_by'
    ).prefetch_related(
        Prefetch('photos', queryset=FinishingReportPhoto.objects.defer('photo_data'))
    ).annotate(
        is_mine=Case(
            When(created_by=request.user, then=Value(1)),
            default=Value(0),
            output_field=IntegerField()
        )
    ).order_by('-created_at')
    p7_qs = EmbroideryReport.objects.select_related(
        'cutting_report__master_entry', 'created_by'
    ).annotate(
        is_mine=Case(
            When(created_by=request.user, then=Value(1)),
            default=Value(0),
            output_field=IntegerField()
        )
    ).order_by('-created_at')
    p8_qs = PrintingReport.objects.select_related(
        'cutting_report__master_entry', 'created_by'
    ).annotate(
        is_mine=Case(
            When(created_by=request.user, then=Value(1)),
            default=Value(0),
            output_field=IntegerField()
        )
    ).order_by('-created_at')
    p9_qs = SingleneedleReport.objects.select_related(
        'cutting_report__master_entry', 'created_by'
    ).prefetch_related(
        Prefetch('photos', queryset=SingleneedleReportPhoto.objects.defer('photo_data'))
    ).annotate(
        is_mine=Case(
            When(created_by=request.user, then=Value(1)),
            default=Value(0),
            output_field=IntegerField()
        )
    ).order_by('-created_at')
    p10_qs = SewingReport.objects.select_related(
        'cutting_report__master_entry', 'created_by'
    ).prefetch_related(
        Prefetch('photos', queryset=SewingReportPhoto.objects.defer('photo_data'))
    ).annotate(
        is_mine=Case(
            When(created_by=request.user, then=Value(1)),
            default=Value(0),
            output_field=IntegerField()
        )
    ).order_by('-created_at')

    p11_qs = JobWork1Report.objects.select_related(
        'cutting_report__master_entry', 'created_by'
    ).prefetch_related(
        Prefetch('photos', queryset=JobWork1ReportPhoto.objects.defer('photo_data'))
    ).annotate(
        is_mine=Case(
            When(created_by=request.user, then=Value(1)),
            default=Value(0),
            output_field=IntegerField()
        )
    ).order_by('-created_at')

    p12_qs = Sewing1Report.objects.select_related(
        'cutting_report__master_entry', 'created_by'
    ).prefetch_related(
        Prefetch('photos', queryset=Sewing1ReportPhoto.objects.defer('photo_data'))
    ).annotate(
        is_mine=Case(
            When(created_by=request.user, then=Value(1)),
            default=Value(0),
            output_field=IntegerField()
        )
    ).order_by('-created_at')

    # Apply Date Filters independently if present
    if date_from_cutting:
        try:
            reports_qs = reports_qs.filter(created_at__date__gte=datetime.strptime(date_from_cutting, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_cutting:
        try:
            reports_qs = reports_qs.filter(created_at__date__lte=datetime.strptime(date_to_cutting, '%Y-%m-%d').date())
        except ValueError:
            pass

    # Apply job card number search filter (server-side)
    if jc_search:
        reports_qs = reports_qs.filter(job_card_no__icontains=jc_search)

    if date_from_stitching:
        try:
            p4_qs = p4_qs.filter(created_at__date__gte=datetime.strptime(date_from_stitching, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_stitching:
        try:
            p4_qs = p4_qs.filter(created_at__date__lte=datetime.strptime(date_to_stitching, '%Y-%m-%d').date())
        except ValueError:
            pass

    if date_from_job_work:
        try:
            p5_qs = p5_qs.filter(created_at__date__gte=datetime.strptime(date_from_job_work, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_job_work:
        try:
            p5_qs = p5_qs.filter(created_at__date__lte=datetime.strptime(date_to_job_work, '%Y-%m-%d').date())
        except ValueError:
            pass

    if date_from_embroidery:
        try:
            p7_qs = p7_qs.filter(created_at__date__gte=datetime.strptime(date_from_embroidery, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_embroidery:
        try:
            p7_qs = p7_qs.filter(created_at__date__lte=datetime.strptime(date_to_embroidery, '%Y-%m-%d').date())
        except ValueError:
            pass

    if date_from_printing:
        try:
            p8_qs = p8_qs.filter(created_at__date__gte=datetime.strptime(date_from_printing, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_printing:
        try:
            p8_qs = p8_qs.filter(created_at__date__lte=datetime.strptime(date_to_printing, '%Y-%m-%d').date())
        except ValueError:
            pass

    if date_from_finishing:
        try:
            p6_qs = p6_qs.filter(created_at__date__gte=datetime.strptime(date_from_finishing, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_finishing:
        try:
            p6_qs = p6_qs.filter(created_at__date__lte=datetime.strptime(date_to_finishing, '%Y-%m-%d').date())
        except ValueError:
            pass

    if date_from_singleneedle:
        try:
            p9_qs = p9_qs.filter(created_at__date__gte=datetime.strptime(date_from_singleneedle, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_singleneedle:
        try:
            p9_qs = p9_qs.filter(created_at__date__lte=datetime.strptime(date_to_singleneedle, '%Y-%m-%d').date())
        except ValueError:
            pass

    if date_from_sewing:
        try:
            p10_qs = p10_qs.filter(created_at__date__gte=datetime.strptime(date_from_sewing, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_sewing:
        try:
            p10_qs = p10_qs.filter(created_at__date__lte=datetime.strptime(date_to_sewing, '%Y-%m-%d').date())
        except ValueError:
            pass

    if date_from_job_work1:
        try:
            p11_qs = p11_qs.filter(created_at__date__gte=datetime.strptime(date_from_job_work1, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_job_work1:
        try:
            p11_qs = p11_qs.filter(created_at__date__lte=datetime.strptime(date_to_job_work1, '%Y-%m-%d').date())
        except ValueError:
            pass

    if date_from_sewing1:
        try:
            p12_qs = p12_qs.filter(created_at__date__gte=datetime.strptime(date_from_sewing1, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_sewing1:
        try:
            p12_qs = p12_qs.filter(created_at__date__lte=datetime.strptime(date_to_sewing1, '%Y-%m-%d').date())
        except ValueError:
            pass

    # Apply job card number search to ALL department querysets (server-side, searches full DB)
    if jc_search:
        p4_qs  = p4_qs.filter(job_card_no__icontains=jc_search)
        p5_qs  = p5_qs.filter(job_card_no__icontains=jc_search)
        # FinishingReport has no job_card_no — search via lot_no and cutting_report FK
        p6_qs  = p6_qs.filter(
            Q(lot_no__icontains=jc_search) |
            Q(cutting_report__job_card_no__icontains=jc_search)
        )
        p7_qs  = p7_qs.filter(job_card_no__icontains=jc_search)
        p8_qs  = p8_qs.filter(job_card_no__icontains=jc_search)
        p9_qs  = p9_qs.filter(job_card_no__icontains=jc_search)
        p10_qs = p10_qs.filter(job_card_no__icontains=jc_search)
        p11_qs = p11_qs.filter(job_card_no__icontains=jc_search)
        p12_qs = p12_qs.filter(job_card_no__icontains=jc_search)


    # Apply Department-Specific Master/Worker Filters if present
    master_name_cutting = request.GET.get('master_name_cutting')
    master_name_stitching = request.GET.get('master_name_stitching')
    master_name_job_work = request.GET.get('master_name_job_work')
    master_name_job_work1 = request.GET.get('master_name_job_work1')
    master_name_finishing = request.GET.get('master_name_finishing')
    master_name_embroidery = request.GET.get('master_name_embroidery')
    master_name_printing = request.GET.get('master_name_printing')
    master_name_singleneedle = request.GET.get('master_name_singleneedle')
    master_name_sewing = request.GET.get('master_name_sewing')
    master_name_sewing1 = request.GET.get('master_name_sewing1')

    if master_name_cutting:
        reports_qs = reports_qs.filter(Q(master_name=master_name_cutting) | Q(cutting_master_name=master_name_cutting))
    if master_name_stitching:
        p4_qs = p4_qs.filter(Q(master_name=master_name_stitching) | Q(stitching_master_name=master_name_stitching))
    if master_name_job_work:
        p5_qs = p5_qs.filter(Q(master_name=master_name_job_work) | Q(jobworker=master_name_job_work))
    if master_name_job_work1:
        p11_qs = p11_qs.filter(Q(master_name=master_name_job_work1) | Q(jobworker=master_name_job_work1))
    if master_name_finishing:
        p6_qs = p6_qs.filter(Q(master_name=master_name_finishing) | Q(finishing_master_name=master_name_finishing))
    if master_name_embroidery:
        p7_qs = p7_qs.filter(Q(master_name=master_name_embroidery) | Q(embroidery_worker=master_name_embroidery))
    if master_name_printing:
        p8_qs = p8_qs.filter(Q(master_name=master_name_printing) | Q(printing_worker=master_name_printing))
    if master_name_singleneedle:
        p9_qs = p9_qs.filter(Q(master_name=master_name_singleneedle) | Q(singleneedle_master_name=master_name_singleneedle))
    if master_name_sewing:
        p10_qs = p10_qs.filter(Q(master_name=master_name_sewing) | Q(sewing_master_name=master_name_sewing))
    if master_name_sewing1:
        p12_qs = p12_qs.filter(Q(master_name=master_name_sewing1) | Q(sewing_master_name=master_name_sewing1))

    from core.models import MasterName
    master_names = MasterName.objects.all().order_by('department', 'name')

    reports = []
    p4_in_progress = []
    p4_completed = []
    p5_in_progress = []
    p5_completed = []
    p11_in_progress = []
    p11_completed = []
    p6_reports = []
    p7_in_progress = []
    p7_completed = []
    p8_in_progress = []
    p8_completed = []
    p9_in_progress = []
    p9_completed = []
    p10_in_progress = []
    p10_completed = []
    p12_in_progress = []
    p12_completed = []
    is_paginated = False
    page_obj = None

    ITEMS_PER_PAGE = 20
    ITEMS_OVERVIEW = 10

    if filter_param:
        is_paginated = True
        if filter_param in ['p1', 'p2', 'p3']:
            paginator = Paginator(reports_qs, ITEMS_PER_PAGE)
            page_obj = paginator.get_page(page_number)
            reports = page_obj
        elif filter_param == 'p4':
            paginator = Paginator(p4_qs, ITEMS_PER_PAGE)
            page_obj = paginator.get_page(page_number)
            p4_in_progress = [r for r in page_obj if not r.line_out_date]
            p4_completed = [r for r in page_obj if r.line_out_date]
        elif filter_param == 'p5':
            paginator = Paginator(p5_qs, ITEMS_PER_PAGE)
            page_obj = paginator.get_page(page_number)
            p5_in_progress = [r for r in page_obj if not r.jobwork_out]
            p5_completed = [r for r in page_obj if r.jobwork_out]
        elif filter_param == 'p11':
            paginator = Paginator(p11_qs, ITEMS_PER_PAGE)
            page_obj = paginator.get_page(page_number)
            p11_in_progress = [r for r in page_obj if not r.jobwork_out]
            p11_completed = [r for r in page_obj if r.jobwork_out]
        elif filter_param == 'p6':
            paginator = Paginator(p6_qs, ITEMS_PER_PAGE)
            page_obj = paginator.get_page(page_number)
            p6_reports = page_obj
        elif filter_param == 'p7':
            paginator = Paginator(p7_qs, ITEMS_PER_PAGE)
            page_obj = paginator.get_page(page_number)
            p7_in_progress = [r for r in page_obj if not r.embroidery_out]
            p7_completed = [r for r in page_obj if r.embroidery_out]
        elif filter_param == 'p8':
            paginator = Paginator(p8_qs, ITEMS_PER_PAGE)
            page_obj = paginator.get_page(page_number)
            p8_in_progress = [r for r in page_obj if not r.printing_out]
            p8_completed = [r for r in page_obj if r.printing_out]
        elif filter_param == 'p9':
            paginator = Paginator(p9_qs, ITEMS_PER_PAGE)
            page_obj = paginator.get_page(page_number)
            p9_in_progress = [r for r in page_obj if not r.line_out_date]
            p9_completed = [r for r in page_obj if r.line_out_date]
        elif filter_param == 'p10':
            paginator = Paginator(p10_qs, ITEMS_PER_PAGE)
            page_obj = paginator.get_page(page_number)
            p10_in_progress = [r for r in page_obj if not r.line_out_date]
            p10_completed = [r for r in page_obj if r.line_out_date]
        elif filter_param == 'p12':
            paginator = Paginator(p12_qs, ITEMS_PER_PAGE)
            page_obj = paginator.get_page(page_number)
            p12_in_progress = [r for r in page_obj if not r.line_out_date]
            p12_completed = [r for r in page_obj if r.line_out_date]
    else:
        if jc_search:
            # When searching by job card, show ALL matching records across every department
            reports = reports_qs
            p4_in_progress = list(p4_qs.filter(line_out_date__isnull=True))
            p4_completed   = list(p4_qs.filter(line_out_date__isnull=False))
            p5_in_progress = list(p5_qs.filter(jobwork_out__isnull=True))
            p5_completed   = list(p5_qs.filter(jobwork_out__isnull=False))
            p11_in_progress = list(p11_qs.filter(jobwork_out__isnull=True))
            p11_completed   = list(p11_qs.filter(jobwork_out__isnull=False))
            p6_reports = p6_qs
            p7_in_progress = list(p7_qs.filter(embroidery_out__isnull=True))
            p7_completed   = list(p7_qs.filter(embroidery_out__isnull=False))
            p8_in_progress = list(p8_qs.filter(printing_out__isnull=True))
            p8_completed   = list(p8_qs.filter(printing_out__isnull=False))
            p9_in_progress = list(p9_qs.filter(line_out_date__isnull=True))
            p9_completed   = list(p9_qs.filter(line_out_date__isnull=False))
            p10_in_progress = list(p10_qs.filter(line_out_date__isnull=True))
            p10_completed   = list(p10_qs.filter(line_out_date__isnull=False))
            p12_in_progress = list(p12_qs.filter(line_out_date__isnull=True))
            p12_completed   = list(p12_qs.filter(line_out_date__isnull=False))
        else:
            # Overview mode: show latest 10 items for each list to ensure fast rendering
            reports = reports_qs[:ITEMS_OVERVIEW]
            p4_in_progress = p4_qs.filter(line_out_date__isnull=True)[:ITEMS_OVERVIEW]
            p4_completed = p4_qs.filter(line_out_date__isnull=False)[:ITEMS_OVERVIEW]
            p5_in_progress = p5_qs.filter(jobwork_out__isnull=True)[:ITEMS_OVERVIEW]
            p5_completed = p5_qs.filter(jobwork_out__isnull=False)[:ITEMS_OVERVIEW]
            p11_in_progress = p11_qs.filter(jobwork_out__isnull=True)[:ITEMS_OVERVIEW]
            p11_completed = p11_qs.filter(jobwork_out__isnull=False)[:ITEMS_OVERVIEW]
        
        p6_reports = p6_qs[:ITEMS_OVERVIEW]
        
        p7_in_progress = p7_qs.filter(embroidery_out__isnull=True)[:ITEMS_OVERVIEW]
        p7_completed = p7_qs.filter(embroidery_out__isnull=False)[:ITEMS_OVERVIEW]
        
        p8_in_progress = p8_qs.filter(printing_out__isnull=True)[:ITEMS_OVERVIEW]
        p8_completed = p8_qs.filter(printing_out__isnull=False)[:ITEMS_OVERVIEW]
        
        p9_in_progress = p9_qs.filter(line_out_date__isnull=True)[:ITEMS_OVERVIEW]
        p9_completed = p9_qs.filter(line_out_date__isnull=False)[:ITEMS_OVERVIEW]

        p10_in_progress = p10_qs.filter(line_out_date__isnull=True)[:ITEMS_OVERVIEW]
        p10_completed = p10_qs.filter(line_out_date__isnull=False)[:ITEMS_OVERVIEW]

        p12_in_progress = p12_qs.filter(line_out_date__isnull=True)[:ITEMS_OVERVIEW]
        p12_completed = p12_qs.filter(line_out_date__isnull=False)[:ITEMS_OVERVIEW]

    return render(request, 'submission_list.html', {
        'reports': reports,
        'p4_in_progress': p4_in_progress,
        'p4_completed': p4_completed,
        'p5_in_progress': p5_in_progress,
        'p5_completed': p5_completed,
        'p11_in_progress': p11_in_progress,
        'p11_completed': p11_completed,
        'p6_reports': p6_reports,
        'p7_in_progress': p7_in_progress,
        'p7_completed': p7_completed,
        'p8_in_progress': p8_in_progress,
        'p8_completed': p8_completed,
        'p9_in_progress': p9_in_progress,
        'p9_completed': p9_completed,
        'p10_in_progress': p10_in_progress,
        'p10_completed': p10_completed,
        'p12_in_progress': p12_in_progress,
        'p12_completed': p12_completed,
        'person_type': person_type,
        'filter_param': filter_param,
        'is_paginated': is_paginated,
        'page_obj': page_obj,
        'show_p1': not filter_param or filter_param in ['p1', 'p2', 'p3'],
        'show_p2': not filter_param or filter_param in ['p1', 'p2', 'p3'],
        'show_p3': not filter_param or filter_param in ['p1', 'p2', 'p3'],
        'show_p4': not filter_param or filter_param == 'p4',
        'show_p5': not filter_param or filter_param == 'p5',
        'show_p11': not filter_param or filter_param == 'p11',
        'show_p6': not filter_param or filter_param == 'p6',
        'show_p7': not filter_param or filter_param == 'p7',
        'show_p8': not filter_param or filter_param == 'p8',
        'show_p9': not filter_param or filter_param == 'p9',
        'show_p10': not filter_param or filter_param == 'p10',
        'show_p12': not filter_param or filter_param == 'p12',
        'master_names': master_names,
        'master_name_cutting': master_name_cutting,
        'master_name_stitching': master_name_stitching,
        'master_name_job_work': master_name_job_work,
        'master_name_job_work1': master_name_job_work1,
        'master_name_finishing': master_name_finishing,
        'master_name_embroidery': master_name_embroidery,
        'master_name_printing': master_name_printing,
        'master_name_singleneedle': master_name_singleneedle,
        'master_name_sewing': master_name_sewing,
        'master_name_sewing1': master_name_sewing1,
        'date_from_cutting': date_from_cutting,
        'date_to_cutting': date_to_cutting,
        'date_from_stitching': date_from_stitching,
        'date_to_stitching': date_to_stitching,
        'date_from_job_work': date_from_job_work,
        'date_to_job_work': date_to_job_work,
        'date_from_job_work1': date_from_job_work1,
        'date_to_job_work1': date_to_job_work1,
        'date_from_finishing': date_from_finishing,
        'date_to_finishing': date_to_finishing,
        'date_from_embroidery': date_from_embroidery,
        'date_to_embroidery': date_to_embroidery,
        'date_from_printing': date_from_printing,
        'date_to_printing': date_to_printing,
        'date_from_singleneedle': date_from_singleneedle,
        'date_to_singleneedle': date_to_singleneedle,
        'date_from_sewing': date_from_sewing,
        'date_to_sewing': date_to_sewing,
        'date_from_sewing1': date_from_sewing1,
        'date_to_sewing1': date_to_sewing1,
        'jc_search': jc_search,
    })


from django.core.exceptions import PermissionDenied

@login_required
def users_reports_view(request):
    if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN')):
        raise PermissionDenied

    query = request.GET.get('q', '').strip()

    p4_qs = StitchingReport.objects.select_related('cutting_report__master_entry', 'created_by').prefetch_related(
        Prefetch('photos', queryset=StitchingReportPhoto.objects.defer('photo_data'))
    ).order_by('-created_at')
    p5_qs = JobWorkReport.objects.filter(created_by__profile__person_type='P5').select_related('cutting_report__master_entry', 'created_by').order_by('-created_at')
    p11_qs = JobWork1Report.objects.select_related('cutting_report__master_entry', 'created_by').prefetch_related(
        Prefetch('photos', queryset=JobWork1ReportPhoto.objects.defer('photo_data'))
    ).order_by('-created_at')
    p6_qs = FinishingReport.objects.select_related('cutting_report__master_entry', 'created_by').prefetch_related(
        Prefetch('photos', queryset=FinishingReportPhoto.objects.defer('photo_data'))
    ).order_by('-created_at')
    p7_qs = EmbroideryReport.objects.select_related('cutting_report__master_entry', 'created_by').order_by('-created_at')
    p8_qs = PrintingReport.objects.select_related('cutting_report__master_entry', 'created_by').order_by('-created_at')
    p9_qs = SingleneedleReport.objects.select_related('cutting_report__master_entry', 'created_by').prefetch_related(
        Prefetch('photos', queryset=SingleneedleReportPhoto.objects.defer('photo_data'))
    ).order_by('-created_at')
    p10_qs = SewingReport.objects.filter(created_by__profile__person_type='P10').select_related('cutting_report__master_entry', 'created_by').prefetch_related(
        Prefetch('photos', queryset=SewingReportPhoto.objects.defer('photo_data'))
    ).order_by('-created_at')
    p12_qs = Sewing1Report.objects.select_related('cutting_report__master_entry', 'created_by').prefetch_related(
        Prefetch('photos', queryset=Sewing1ReportPhoto.objects.defer('photo_data'))
    ).order_by('-created_at')
    misc_qs = MiscellaneousReport.objects.select_related('created_by').prefetch_related('files').order_by('-created_at')

    if query:
        p4_qs = p4_qs.filter(job_card_no__icontains=query)
        p5_qs = p5_qs.filter(job_card_no__icontains=query)
        p11_qs = p11_qs.filter(job_card_no__icontains=query)
        p6_qs = p6_qs.filter(cutting_report__job_card_no__icontains=query)
        p7_qs = p7_qs.filter(job_card_no__icontains=query)
        p8_qs = p8_qs.filter(job_card_no__icontains=query)
        p9_qs = p9_qs.filter(job_card_no__icontains=query)
        p10_qs = p10_qs.filter(job_card_no__icontains=query)
        p12_qs = p12_qs.filter(job_card_no__icontains=query)
        misc_qs = misc_qs.filter(job_card_no__icontains=query)
        limit = 50
    else:
        limit = 5

    # Split P4 reports
    p4_in_progress = p4_qs.filter(line_out_date__isnull=True)[:limit]
    p4_completed = p4_qs.filter(line_out_date__isnull=False)[:limit]
    # Split P9 reports
    p9_in_progress = p9_qs.filter(line_out_date__isnull=True)[:limit]
    p9_completed = p9_qs.filter(line_out_date__isnull=False)[:limit]
    # Split P10 reports
    p10_in_progress = p10_qs.filter(line_out_date__isnull=True)[:limit]
    p10_completed = p10_qs.filter(line_out_date__isnull=False)[:limit]
    # Split P12 reports
    p12_in_progress = p12_qs.filter(line_out_date__isnull=True)[:limit]
    p12_completed = p12_qs.filter(line_out_date__isnull=False)[:limit]

    context = {
        'search_query': query,
        'person1_reports_count': CuttingReport.objects.filter(report_type='P1').count(),
        'person2_reports_count': CuttingReport.objects.filter(report_type='P2').count(),
        'person3_reports_count': CuttingReport.objects.filter(report_type='P3').count(),
        'stitching_reports_count': StitchingReport.objects.count(),
        'jobwork_reports_count': JobWorkReport.objects.filter(created_by__profile__person_type='P5').count(),
        'jobwork1_reports_count': JobWork1Report.objects.count(),
        'finishing_reports_count': FinishingReport.objects.count(),
        'embroidery_reports_count': EmbroideryReport.objects.count(),
        'printing_reports_count': PrintingReport.objects.count(),
        'singleneedle_reports_count': SingleneedleReport.objects.count(),
        'sewing_reports_count': SewingReport.objects.filter(created_by__profile__person_type='P10').count(),
        'sewing1_reports_count': Sewing1Report.objects.count(),
        'miscellaneous_reports_count': MiscellaneousReport.objects.count(),
        'recent_p4_in_progress': p4_in_progress,
        'recent_p4_completed': p4_completed,
        'recent_jobwork_reports': p5_qs[:limit],
        'recent_jobwork1_reports': p11_qs[:limit],
        'recent_finishing_reports': p6_qs[:limit],
        'recent_embroidery_reports': p7_qs[:limit],
        'recent_printing_reports': p8_qs[:limit],
        'recent_p9_in_progress': p9_in_progress,
        'recent_p9_completed': p9_completed,
        'recent_p10_in_progress': p10_in_progress,
        'recent_p10_completed': p10_completed,
        'recent_p12_in_progress': p12_in_progress,
        'recent_p12_completed': p12_completed,
        'recent_misc_reports': misc_qs[:limit],
    }
    return render(request, 'users_reports.html', context)

# ── QR Code & Detail Views ──────────────────────────────────────────────────

def job_card_detail_view(request, pk):
    """Publicly accessible view for scanned QR codes"""
    master_entry = get_object_or_404(MasterEntry, pk=pk)
    report_type = request.GET.get('type', None)  # e.g. P1, P2, P3, P4, P5, P6

    # Get all cutting reports grouped by type
    all_cutting = master_entry.cutting_reports.prefetch_related('color_details', 'photos').all()

    # Resolve which reports to show based on ?type= param
    show_all = report_type is None or report_type == 'ADMIN'

    p1_report = all_cutting.filter(report_type='P1').first() if (show_all or report_type == 'P1') else None
    p2_report = all_cutting.filter(report_type='P2').first() if (show_all or report_type == 'P2') else None
    p3_report = all_cutting.filter(report_type='P3').first() if (show_all or report_type == 'P3') else None

    # Link subsequent reports (P4, P5, P6, P7, P8, P9, P10) to whatever cutting report was submitted
    base_cutting = all_cutting.first()
    stitching_report = base_cutting.stitching_reports.first() if (base_cutting and (show_all or report_type == 'P4')) else None
    jobwork_report = base_cutting.jobwork_reports.first() if (base_cutting and (show_all or report_type == 'P5')) else None
    finishing_report = base_cutting.finishing_reports.first() if (base_cutting and (show_all or report_type == 'P6')) else None
    embroidery_report = base_cutting.embroidery_reports.first() if (base_cutting and (show_all or report_type == 'P7')) else None
    printing_report = base_cutting.printing_reports.first() if (base_cutting and (show_all or report_type == 'P8')) else None
    singleneedle_report = base_cutting.singleneedle_reports.first() if (base_cutting and (show_all or report_type == 'P9')) else None
    sewing_report = base_cutting.sewing_reports.first() if (base_cutting and (show_all or report_type == 'P10')) else None

    return render(request, 'job_card_detail.html', {
        'master_entry': master_entry,
        'cutting_report': p1_report,
        'person2_report': p2_report,
        'person3_report': p3_report,
        'stitching_report': stitching_report,
        'jobwork_report': jobwork_report,
        'finishing_report': finishing_report,
        'embroidery_report': embroidery_report,
        'printing_report': printing_report,
        'singleneedle_report': singleneedle_report,
        'sewing_report': sewing_report,
        'report_type': report_type,
    })

@login_required
def job_card_print_view(request, pk):
    """View to generate and print the QR code — embeds user's person_type in URL"""
    master_entry = get_object_or_404(MasterEntry, pk=pk)

    # Determine the person type of the current user
    try:
        person_type = request.user.profile.person_type
    except Exception:
        person_type = None

    # Build URL with type filter so QR only shows that user's report
    base_url = request.build_absolute_uri(f'/job-card/{pk}/')
    if person_type and person_type != 'ADMIN':
        qr_url = f"{base_url}?type={person_type}"
    else:
        qr_url = base_url  # Admins see all

    return render(request, 'job_card_print.html', {
        'master_entry': master_entry,
        'qr_url': qr_url,
    })

# ── Excel Download ────────────────────────────────────────────────────────────

from django.utils import timezone

@login_required
def export_options_view(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    
    settings = SystemSetting.get_settings()
    
    return render(request, 'export_options.html', {
        'last_download_at': settings.last_excel_download_at
    })

@login_required
def export_excel_view(request):
    if not request.user.is_superuser:
        raise PermissionDenied
        
    export_type = request.GET.get('type', 'all')
    settings = SystemSetting.get_settings()
    
    since_date = None
    if export_type == 'new' and settings.last_excel_download_at:
        since_date = settings.last_excel_download_at
        
    try:
        filepath = export_to_excel(since_date=since_date)
        if os.path.exists(filepath):
            # Update last downloaded time
            settings.last_excel_download_at = timezone.now()
            settings.save()
            
            response = FileResponse(
                open(filepath, 'rb'),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="FabricTrack_Data.xlsx"'
            return response
        raise Http404("Export file not found.")
    except Exception as e:
        messages.error(request, f'Export failed: {e}')
        return redirect('dashboard')

@login_required
def download_sample_excel_view(request):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Job Cards Import'

    headers = [
        'Date', 'Jobcard number', 'Cutting', 'Jobworker', 'Jobwork 1',
        'Stitching', 'Embroidery', 'Printing', 'Singleneedle', 'Sewing', 'Sewing 1', 'Finishing'
    ]
    ws.append(headers)

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    sample_data = [
        ['2026-07-23', 'JC-9001', 1, 2, 0, 3, 0, 0, 0, 0, 0, 4],
        ['2026-07-23', 'JC-9002', 1, 0, 0, 2, 3, 4, 0, 0, 0, 5],
        ['2026-07-24', 'JC-9003', 1, 0, 0, 2, 0, 0, 0, 0, 0, 3],
        ['2026-07-24', 'JC-9004', 1, 2, 3, 4, 0, 0, 0, 0, 0, 5],
        ['2026-07-25', 'JC-9005', 1, 0, 0, 2, 0, 0, 3, 4, 0, 5],
    ]

    for row_data in sample_data:
        ws.append(row_data)

    for row in ws.iter_rows(min_row=2, max_row=len(sample_data)+1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = align_center
            cell.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample_job_cards_import.xlsx"'
    return response

import openpyxl
from datetime import datetime

@login_required
def import_job_cards_view(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            # Assuming headers in row 1: Date, Jobcard number, Cutting, Jobworker, Stiching, Finishing
            # We will find the column indices dynamically to be safe
            headers = [cell.value for cell in sheet[1]]
            header_map = {str(h).strip().lower(): idx for idx, h in enumerate(headers) if h}
            
            # Check required columns
            required_cols = ['jobcard number']
            if not all(col in header_map for col in required_cols):
                messages.error(request, f"Missing required columns. Found headers: {', '.join([str(h) for h in headers if h])}")
                return redirect('create_master_entry')
            
            created_count = 0
            updated_count = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                job_card_no = str(row[header_map.get('jobcard number', -1)]).strip()
                if not job_card_no or job_card_no == 'None': continue
                
                def get_val(key1, key2=None):
                    idx = header_map.get(key1)
                    if idx is None and key2:
                        idx = header_map.get(key2)
                    if idx is not None and idx < len(row):
                        return row[idx]
                    return None
                
                def parse_sequence_val(val):
                    """Parse Excel cell: numeric → int, yes/true → 1, no/false/blank → 0"""
                    if val is None:
                        return 0
                    val_str = str(val).strip().lower()
                    if val_str in ('no', 'false', '0', 'none', ''):
                        return 0
                    if val_str in ('yes', 'true'):
                        return 1
                    try:
                        return int(float(val_str))
                    except (ValueError, TypeError):
                        return 0

                cutting_req     = parse_sequence_val(get_val('cutting'))
                jobwork_req     = parse_sequence_val(get_val('jobwork', 'jobworker'))
                jobwork1_req    = parse_sequence_val(get_val('jobwork1', 'jobwork 1'))
                stitching_req   = parse_sequence_val(get_val('stitching', 'stiching'))
                finishing_req   = parse_sequence_val(get_val('finishing'))
                embroidery_req  = parse_sequence_val(get_val('embroidery'))
                printing_req    = parse_sequence_val(get_val('printing'))
                singleneedle_req= parse_sequence_val(get_val('singleneedle'))
                sewing_req      = parse_sequence_val(get_val('sewing'))
                sewing1_req     = parse_sequence_val(get_val('sewing1', 'sewing 1'))
                
                # Parse date if possible
                date_val = row[header_map.get('date', -1)]
                if isinstance(date_val, datetime):
                    date_obj = date_val.date()
                elif date_val:
                    try:
                        date_obj = datetime.strptime(str(date_val).strip(), '%m/%d/%Y').date()
                    except ValueError:
                        date_obj = timezone.now().date()
                else:
                    date_obj = timezone.now().date()

                obj, created = JobCardRequirement.objects.update_or_create(
                    job_card_no=job_card_no,
                    defaults={
                        'date': date_obj,
                        'requires_cutting': cutting_req,
                        'requires_jobwork': jobwork_req,
                        'requires_jobwork1': jobwork1_req,
                        'requires_stitching': stitching_req,
                        'requires_finishing': finishing_req,
                        'requires_embroidery': embroidery_req,
                        'requires_printing': printing_req,
                        'requires_singleneedle': singleneedle_req,
                        'requires_sewing': sewing_req,
                        'requires_sewing1': sewing1_req,
                    }
                )

                # Create MasterEntry if it doesn't exist
                MasterEntry.objects.get_or_create(
                    job_card_number=job_card_no,
                    defaults={
                        'date': date_obj,
                        'created_by': request.user
                    }
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1
                    
            messages.success(request, f"Successfully imported! Created {created_count} new tasks, updated {updated_count} existing tasks.")
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            
        return redirect('create_master_entry')
        
    return redirect('create_master_entry')

# ── Edit and Delete Views ───────────────────────────────────────────────────

@login_required
@require_POST
def delete_pending_task(request, pk):
    if not request.user.is_superuser and getattr(request.user, 'profile', None) and request.user.profile.person_type != 'ADMIN':
        raise PermissionDenied
    task = get_object_or_404(JobCardRequirement, pk=pk)
    task.delete()
    messages.success(request, f"Pending task for {task.job_card_no} deleted.")
    return redirect('dashboard')


from django.core.exceptions import PermissionDenied

def check_edit_permission(request, obj):
    if request.user.is_superuser: return True
    if hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN': return True
    if getattr(obj, 'created_by', None) == request.user: return True
    return False

@login_required
def edit_master_entry(request, pk):
    entry = get_object_or_404(MasterEntry, pk=pk)
    if not check_edit_permission(request, entry): raise PermissionDenied

    # Sequence field mapping: form field -> human label
    _SEQ_LABELS = {
        'requires_cutting': 'Cutting Seq', 'requires_jobwork': 'Job Work Seq',
        'requires_jobwork1': 'Job Work 1 Seq', 'requires_stitching': 'Stitching Seq',
        'requires_finishing': 'Finishing Seq', 'requires_embroidery': 'Embroidery Seq',
        'requires_printing': 'Printing Seq', 'requires_singleneedle': 'Singleneedle Seq',
        'requires_sewing': 'Sewing Seq', 'requires_sewing1': 'Sewing 1 Seq',
    }

    if request.method == 'POST':
        # Snapshot BEFORE: MasterEntry fields + existing JobCardRequirement sequences
        _old = {
            'Date': str(entry.date),
            'Job Card Number': str(entry.job_card_number),
        }
        _old_req = JobCardRequirement.objects.filter(job_card_no=entry.job_card_number).first()
        for field, label in _SEQ_LABELS.items():
            _old[label] = str(getattr(_old_req, field, 0) or 0) if _old_req else '0'

        form = MasterEntryForm(request.POST, instance=entry)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.created_by = request.user
            entry.save()

            # Update the JobCardRequirement with sequence values
            new_req_data = {
                'date': entry.date,
                'requires_cutting':      form.cleaned_data.get('requires_cutting', 0),
                'requires_jobwork':      form.cleaned_data.get('requires_jobwork', 0),
                'requires_jobwork1':     form.cleaned_data.get('requires_jobwork1', 0),
                'requires_stitching':    form.cleaned_data.get('requires_stitching', 0),
                'requires_finishing':    form.cleaned_data.get('requires_finishing', 0),
                'requires_embroidery':   form.cleaned_data.get('requires_embroidery', 0),
                'requires_printing':     form.cleaned_data.get('requires_printing', 0),
                'requires_singleneedle': form.cleaned_data.get('requires_singleneedle', 0),
                'requires_sewing':       form.cleaned_data.get('requires_sewing', 0),
                'requires_sewing1':      form.cleaned_data.get('requires_sewing1', 0),
            }
            JobCardRequirement.objects.update_or_create(
                job_card_no=entry.job_card_number,
                defaults=new_req_data,
            )

            # Snapshot AFTER: updated MasterEntry + new requirement values
            _new = {
                'Date': str(entry.date),
                'Job Card Number': str(entry.job_card_number),
            }
            for field, label in _SEQ_LABELS.items():
                _new[label] = str(new_req_data.get(field, 0))

            _diff = build_diff(_old, _new)
            messages.success(request, 'Master entry updated.')
            log_activity(request.user, 'EDIT', 'Master Entry', entry.job_card_number, _diff)
            return redirect('dashboard')
    else:
        # Pre-fill form with existing JobCardRequirement values
        req = JobCardRequirement.objects.filter(job_card_no=entry.job_card_number).first()
        initial_data = {}
        if req:
            initial_data = {
                'requires_cutting':      req.requires_cutting,
                'requires_jobwork':      req.requires_jobwork,
                'requires_jobwork1':     req.requires_jobwork1,
                'requires_stitching':    req.requires_stitching,
                'requires_finishing':    req.requires_finishing,
                'requires_embroidery':   req.requires_embroidery,
                'requires_printing':     req.requires_printing,
                'requires_singleneedle': req.requires_singleneedle,
                'requires_sewing':       req.requires_sewing,
                'requires_sewing1':      req.requires_sewing1,
            }
        form = MasterEntryForm(instance=entry, initial=initial_data)
    return render(request, 'master_entry_form.html', {'form': form, 'is_edit': True})

@login_required
def delete_master_entry(request, pk):
    entry = get_object_or_404(MasterEntry, pk=pk)
    if not check_edit_permission(request, entry): raise PermissionDenied
    if request.method == 'POST':
        jc_no = entry.job_card_number
        # Build rich delete summary including requirement sequences
        _summary_parts = [f"Job Card: {entry.job_card_number}", f"Date: {entry.date}"]
        _del_req = JobCardRequirement.objects.filter(job_card_no=jc_no).first()
        if _del_req:
            _del_seq_map = {
                'requires_cutting': 'Cutting', 'requires_jobwork': 'Job Work',
                'requires_jobwork1': 'Job Work 1', 'requires_stitching': 'Stitching',
                'requires_finishing': 'Finishing', 'requires_embroidery': 'Embroidery',
                'requires_printing': 'Printing', 'requires_singleneedle': 'Singleneedle',
                'requires_sewing': 'Sewing', 'requires_sewing1': 'Sewing 1',
            }
            for field, label in _del_seq_map.items():
                val = getattr(_del_req, field, 0)
                if val:
                    _summary_parts.append(f"{label} Seq: {val}")
        _summary = ' | '.join(_summary_parts)
        entry.delete()
        messages.success(request, 'Master entry deleted.')
        log_activity(request.user, 'DELETE', 'Master Entry', jc_no, _summary)
        return redirect('dashboard')
    return render(request, 'confirm_delete.html', {'object': entry, 'cancel_url': 'dashboard'})

@login_required
def edit_cutting_report(request, pk):
    report = get_object_or_404(CuttingReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    
    # Handle photo deletion if requested via URL param
    delete_photo_id = request.GET.get('delete_photo')
    if delete_photo_id:
        photo_to_delete = get_object_or_404(CuttingReportPhoto, pk=delete_photo_id, cutting_report=report)
        photo_to_delete.delete()
        messages.success(request, 'Photo deleted successfully.')
        return redirect(request.path)

    is_admin = request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN')
    master_entries_qs = MasterEntry.objects.filter(Q(cutting_reports__isnull=True) | Q(id=report.master_entry_id)).distinct().order_by('-date')
    master_entries_json = json.dumps({str(e.id): e.job_card_number for e in master_entries_qs})
    colors_qs = report.color_details.order_by('id')
    colors_json = json.dumps([
        {
            'color_name': c.color_name,
            'size_s': c.size_s, 'size_m': c.size_m, 'size_l': c.size_l, 'size_xl': c.size_xl,
            'size_2xl': c.size_2xl, 'size_3xl': c.size_3xl, 'size_4xl': c.size_4xl,
            'total_weight': float(c.total_weight) if c.total_weight is not None else 0.0,
            'total_meters': float(c.total_meters) if c.total_meters is not None else 0.0,
        } for c in colors_qs
    ])
    
    from .models import RateDefinition
    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })
    
    if request.method == 'POST':
        _fields = ['item_name', 'total_pcs', 'cutting_rate']
        _old = capture_snapshot(report, _fields)
        form = CuttingReportForm(request.POST, request.FILES, instance=report)
        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                report.rate_name = report.rate_definition.name
                report.cutting_rate = report.rate_definition.total_rate
            report.save()
            
            # Re-save dynamic color size breakdown
            report.color_details.all().delete()
            num_colors = report.total_colours
            if num_colors > 0:
                for i in range(1, num_colors + 1):
                    c_name = request.POST.get(f'color_name_{i}', f'C{i}')
                    c_s = int(request.POST.get(f'color_s_{i}') or 0)
                    c_m = int(request.POST.get(f'color_m_{i}') or 0)
                    c_l = int(request.POST.get(f'color_l_{i}') or 0)
                    c_xl = int(request.POST.get(f'color_xl_{i}') or 0)
                    c_2xl = int(request.POST.get(f'color_2xl_{i}') or 0)
                    c_3xl = int(request.POST.get(f'color_3xl_{i}') or 0)
                    c_4xl = int(request.POST.get(f'color_4xl_{i}') or 0)
                    c_weight = request.POST.get(f'color_weight_{i}') or 0.0
                    c_meters = request.POST.get(f'color_meters_{i}') or 0.0
                    CuttingReportColorDetail.objects.create(
                        cutting_report=report, color_name=c_name,
                        size_s=c_s, size_m=c_m, size_l=c_l, size_xl=c_xl,
                        size_2xl=c_2xl, size_3xl=c_3xl, size_4xl=c_4xl,
                        total_weight=c_weight, total_meters=c_meters
                    )
            photos = request.FILES.getlist('photos')
            for photo_file in photos:
                if report.photos.count() < 5:
                    CuttingReportPhoto.objects.create(
                        cutting_report=report,
                        photo_data=photo_file.read(),
                        photo_name=photo_file.name,
                        photo_content_type=photo_file.content_type
                    )
            messages.success(request, 'Cutting Report updated.')
            _diff = build_diff(_old, capture_snapshot(report, _fields))
            log_activity(request.user, 'EDIT', 'Cutting Report', report.job_card_no, _diff)
            return redirect('submission_list')
    else:
        form = CuttingReportForm(instance=report)
    return render(request, 'person1_form.html', {
        'form': form, 'master_entries': master_entries_qs,
        'master_entries_json': master_entries_json, 'is_admin': is_admin, 'is_edit': True, 'report': report,
        'colors_json': colors_json,
        'rate_definitions_json': rate_definitions_json
    })

@login_required
def delete_cutting_report(request, pk):
    report = get_object_or_404(CuttingReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    if request.method == 'POST':
        jc_no = report.job_card_no
        _summary = get_obj_summary(report)
        JobCardRequirement.objects.filter(job_card_no=jc_no).update(is_cutting_done=False)
        report.delete()
        messages.success(request, 'Cutting Report deleted.')
        log_activity(request.user, 'DELETE', 'Cutting Report', jc_no, _summary)
        return redirect('submission_list')
    return render(request, 'confirm_delete.html', {'object': report, 'cancel_url': 'submission_list'})

@login_required
def edit_stitching_report(request, pk):
    report = get_object_or_404(StitchingReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    is_admin = request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN')
    cutting_reports_qs = CuttingReport.objects.filter(Q(stitching_reports__isnull=True) | Q(id=report.cutting_report_id)).distinct().select_related('master_entry').order_by('-created_at')
    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })

    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    delete_photo_id = request.GET.get('delete_photo')
    if delete_photo_id:
        photo_to_delete = get_object_or_404(StitchingReportPhoto, pk=delete_photo_id, stitching_report=report)
        photo_to_delete.delete()
        messages.success(request, 'Photo deleted successfully.')
        return redirect('edit_stitching_report', pk=report.id)

    if request.method == 'POST':
        _fields = ['item_name', 'total_pcs', 'total_rate']
        _old = capture_snapshot(report, _fields)
        form = StitchingReportForm(request.POST, request.FILES, instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
        photos = request.FILES.getlist('photos')

        if len(photos) + report.photos.count() > 5:
            messages.error(request, 'You can upload a maximum of 5 photos total.')
            return redirect('edit_stitching_report', pk=report.id)

        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                report.rate_name = report.rate_definition.name
                report.rate_description = report.rate_definition.description
                report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos:
                StitchingReportPhoto.objects.create(
                    stitching_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Update pending task status based on Line In/Out dates
            if report.line_out_date:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_stitching_done=True, is_stitching_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_stitching_in_progress=True, is_stitching_done=False
                )

            messages.success(request, 'Stitching updated.')
            _diff = build_diff(_old, capture_snapshot(report, _fields))
            log_activity(request.user, 'EDIT', 'Stitching', report.job_card_no, _diff)
            return redirect('submission_list')
    else:
        form = StitchingReportForm(instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
    return render(request, 'stitching_form.html', {
        'form': form, 'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin, 'is_edit': True, 'report': report
    })

@login_required
def delete_stitching_report(request, pk):
    report = get_object_or_404(StitchingReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    if request.method == 'POST':
        jc_no = report.job_card_no
        _summary = get_obj_summary(report)
        JobCardRequirement.objects.filter(job_card_no=jc_no).update(
            is_stitching_done=False, is_stitching_in_progress=False
        )
        report.delete()
        messages.success(request, 'Stitching deleted.')
        log_activity(request.user, 'DELETE', 'Stitching', jc_no, _summary)
        return redirect('submission_list')
    return render(request, 'confirm_delete.html', {'object': report, 'cancel_url': 'submission_list'})

@login_required
def edit_singleneedle_report(request, pk):
    report = get_object_or_404(SingleneedleReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    is_admin = request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN')
    cutting_reports_qs = CuttingReport.objects.filter(Q(singleneedle_reports__isnull=True) | Q(id=report.cutting_report_id)).distinct().select_related('master_entry').order_by('-created_at')
    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })

    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    delete_photo_id = request.GET.get('delete_photo')
    if delete_photo_id:
        photo_to_delete = get_object_or_404(SingleneedleReportPhoto, pk=delete_photo_id, singleneedle_report=report)
        photo_to_delete.delete()
        messages.success(request, 'Photo deleted successfully.')
        return redirect('edit_singleneedle_report', pk=report.id)

    if request.method == 'POST':
        _fields = ['item_name', 'total_pcs', 'total_rate']
        _old = capture_snapshot(report, _fields)
        form = SingleneedleReportForm(request.POST, request.FILES, instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
        photos = request.FILES.getlist('photos')

        if len(photos) + report.photos.count() > 5:
            messages.error(request, 'You can upload a maximum of 5 photos total.')
            return redirect('edit_singleneedle_report', pk=report.id)

        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                report.rate_name = report.rate_definition.name
                report.rate_description = report.rate_definition.description
                report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos:
                SingleneedleReportPhoto.objects.create(
                    singleneedle_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Update pending task status based on Line In/Out dates
            if report.line_out_date:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_singleneedle_done=True, is_singleneedle_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_singleneedle_in_progress=True, is_singleneedle_done=False
                )

            messages.success(request, 'Singleneedle updated.')
            _diff = build_diff(_old, capture_snapshot(report, _fields))
            log_activity(request.user, 'EDIT', 'Singleneedle', report.job_card_no, _diff)
            return redirect('submission_list')
    else:
        form = SingleneedleReportForm(instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
    return render(request, 'singleneedle_form.html', {
        'form': form, 'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin, 'is_edit': True, 'report': report
    })

@login_required
def delete_singleneedle_report(request, pk):
    report = get_object_or_404(SingleneedleReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    if request.method == 'POST':
        jc_no = report.job_card_no
        _summary = get_obj_summary(report)
        JobCardRequirement.objects.filter(job_card_no=jc_no).update(
            is_singleneedle_done=False, is_singleneedle_in_progress=False
        )
        report.delete()
        messages.success(request, 'Singleneedle deleted.')
        log_activity(request.user, 'DELETE', 'Singleneedle', jc_no, _summary)
        return redirect('submission_list')
    return render(request, 'confirm_delete.html', {'object': report, 'cancel_url': 'submission_list'})


@login_required
def edit_sewing_report(request, pk):
    report = get_object_or_404(SewingReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    is_admin = request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN')
    cutting_reports_qs = CuttingReport.objects.filter(Q(sewing_reports__isnull=True) | Q(id=report.cutting_report_id)).distinct().select_related('master_entry').order_by('-created_at')
    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })

    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    delete_photo_id = request.GET.get('delete_photo')
    if delete_photo_id:
        photo_to_delete = get_object_or_404(SewingReportPhoto, pk=delete_photo_id, sewing_report=report)
        photo_to_delete.delete()
        messages.success(request, 'Photo deleted successfully.')
        return redirect('edit_sewing_report', pk=report.id)

    if request.method == 'POST':
        _fields = ['item_name', 'total_pcs', 'total_rate']
        _old = capture_snapshot(report, _fields)
        form = SewingReportForm(request.POST, request.FILES, instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
        photos = request.FILES.getlist('photos')

        if len(photos) + report.photos.count() > 5:
            messages.error(request, 'You can upload a maximum of 5 photos total.')
            return redirect('edit_sewing_report', pk=report.id)

        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                report.rate_name = report.rate_definition.name
                report.rate_description = report.rate_definition.description
                report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos:
                SewingReportPhoto.objects.create(
                    sewing_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Update pending task status based on Line In/Out dates
            if report.line_out_date:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_sewing_done=True, is_sewing_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_sewing_in_progress=True, is_sewing_done=False
                )

            messages.success(request, 'Sewing updated.')
            _diff = build_diff(_old, capture_snapshot(report, _fields))
            log_activity(request.user, 'EDIT', 'Sewing', report.job_card_no, _diff)
            return redirect('submission_list')
    else:
        form = SewingReportForm(instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
    return render(request, 'sewing_form.html', {
        'form': form, 'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin, 'is_edit': True, 'report': report
    })

@login_required
def delete_sewing_report(request, pk):
    report = get_object_or_404(SewingReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    if request.method == 'POST':
        jc_no = report.job_card_no
        _summary = get_obj_summary(report)
        JobCardRequirement.objects.filter(job_card_no=jc_no).update(
            is_sewing_done=False, is_sewing_in_progress=False
        )
        report.delete()
        messages.success(request, 'Sewing deleted.')
        log_activity(request.user, 'DELETE', 'Sewing', jc_no, _summary)
        return redirect('submission_list')
    return render(request, 'confirm_delete.html', {'object': report, 'cancel_url': 'submission_list'})


@login_required
def edit_sewing1_report(request, pk):
    report = get_object_or_404(Sewing1Report, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    is_admin = request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN')
    cutting_reports_qs = CuttingReport.objects.filter(Q(sewing1_reports__isnull=True) | Q(id=report.cutting_report_id)).distinct().select_related('master_entry').order_by('-created_at')
    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })

    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    delete_photo_id = request.GET.get('delete_photo')
    if delete_photo_id:
        photo_to_delete = get_object_or_404(Sewing1ReportPhoto, pk=delete_photo_id, sewing1_report=report)
        photo_to_delete.delete()
        messages.success(request, 'Photo deleted successfully.')
        return redirect('edit_sewing1_report', pk=report.id)

    if request.method == 'POST':
        _fields = ['item_name', 'total_pcs', 'total_rate']
        _old = capture_snapshot(report, _fields)
        form = Sewing1ReportForm(request.POST, request.FILES, instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
        photos = request.FILES.getlist('photos')

        if len(photos) + report.photos.count() > 5:
            messages.error(request, 'You can upload a maximum of 5 photos total.')
            return redirect('edit_sewing1_report', pk=report.id)

        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                report.rate_name = report.rate_definition.name
                report.rate_description = report.rate_definition.description
                report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos:
                Sewing1ReportPhoto.objects.create(
                    sewing1_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Update pending task status based on Line In/Out dates
            if report.line_out_date:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_sewing1_done=True, is_sewing1_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=report.job_card_no).update(
                    is_sewing1_in_progress=True, is_sewing1_done=False
                )

            messages.success(request, 'Sewing 1 updated.')
            _diff = build_diff(_old, capture_snapshot(report, _fields))
            log_activity(request.user, 'EDIT', 'Sewing 1', report.job_card_no, _diff)
            return redirect('submission_list')
    else:
        form = Sewing1ReportForm(instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
    return render(request, 'sewing_form.html', {
        'form': form, 'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin, 'is_edit': True, 'report': report
    })

@login_required
def delete_sewing1_report(request, pk):
    report = get_object_or_404(Sewing1Report, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    if request.method == 'POST':
        jc_no = report.job_card_no
        _summary = get_obj_summary(report)
        JobCardRequirement.objects.filter(job_card_no=jc_no).update(
            is_sewing1_done=False, is_sewing1_in_progress=False
        )
        report.delete()
        messages.success(request, 'Sewing 1 deleted.')
        log_activity(request.user, 'DELETE', 'Sewing 1', jc_no, _summary)
        return redirect('submission_list')
    return render(request, 'confirm_delete.html', {'object': report, 'cancel_url': 'submission_list'})



@login_required
def edit_jobwork_report(request, pk):
    report = get_object_or_404(JobWorkReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    is_admin = request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN')
    cutting_reports_qs = CuttingReport.objects.filter(Q(jobwork_reports__isnull=True) | Q(id=report.cutting_report_id)).distinct().select_related('master_entry').order_by('-created_at')
    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })
    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })
    delete_photo_id = request.GET.get('delete_photo')
    if delete_photo_id:
        if report.photos.count() <= 1:
            messages.error(request, 'Cannot delete. At least one Job Card Photo is required.')
        else:
            photo_to_delete = get_object_or_404(JobWorkReportPhoto, pk=delete_photo_id, job_work_report=report)
            photo_to_delete.delete()
            messages.success(request, 'Photo deleted successfully.')
        return redirect('edit_jobwork_report', pk=report.id)

    if request.method == 'POST':
        _fields = ['jobworker', 'purpose', 'total_pcs', 'total_rate']
        _old = capture_snapshot(report, _fields)
        form = JobWorkReportForm(request.POST, request.FILES, instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
        photos = request.FILES.getlist('photos')

        if len(photos) + report.photos.count() == 0:
            messages.error(request, 'At least one Job Card Photo is required.')
            return render(request, 'jobwork_form.html', {
                'form': form, 'cutting_reports': cutting_reports_qs,
                'cutting_reports_json': cutting_reports_json, 'is_admin': is_admin, 'is_edit': True, 'report': report
            })

        if len(photos) + report.photos.count() > 5:
            messages.error(request, 'You can upload a maximum of 5 photos total.')
            return render(request, 'jobwork_form.html', {
                'form': form, 'cutting_reports': cutting_reports_qs,
                'cutting_reports_json': cutting_reports_json, 'is_admin': is_admin, 'is_edit': True, 'report': report
            })

        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                if not report.total_rate:
                    report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos:
                JobWorkReportPhoto.objects.create(
                    job_work_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Update pending task status based on In/Out dates
            job_card_no = report.cutting_report.job_card_no
            if report.jobwork_out:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_jobwork_done=True, is_jobwork_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_jobwork_in_progress=True, is_jobwork_done=False
                )

            messages.success(request, 'Job Work updated.')
            _diff = build_diff(_old, capture_snapshot(report, _fields))
            log_activity(request.user, 'EDIT', 'Job Work', report.cutting_report.job_card_no, _diff)
            return redirect('submission_list')
    else:
        form = JobWorkReportForm(instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
    return render(request, 'jobwork_form.html', {
        'form': form, 'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin, 'is_edit': True, 'report': report
    })

@login_required
def delete_jobwork_report(request, pk):
    report = get_object_or_404(JobWorkReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    if request.method == 'POST':
        jc_no = report.cutting_report.job_card_no
        _summary = get_obj_summary(report)
        JobCardRequirement.objects.filter(job_card_no=jc_no).update(
            is_jobwork_done=False, is_jobwork_in_progress=False
        )
        report.delete()
        messages.success(request, 'Job Work deleted.')
        log_activity(request.user, 'DELETE', 'Job Work', jc_no, _summary)
        return redirect('submission_list')
    return render(request, 'confirm_delete.html', {'object': report, 'cancel_url': 'submission_list'})


@login_required
def edit_jobwork1_report(request, pk):
    report = get_object_or_404(JobWork1Report, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    is_admin = request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN')
    cutting_reports_qs = CuttingReport.objects.filter(Q(jobwork1_reports__isnull=True) | Q(id=report.cutting_report_id)).distinct().select_related('master_entry').order_by('-created_at')
    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })
    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })
    delete_photo_id = request.GET.get('delete_photo')
    if delete_photo_id:
        if report.photos.count() <= 1:
            messages.error(request, 'Cannot delete. At least one Job Card Photo is required.')
        else:
            photo_to_delete = get_object_or_404(JobWork1ReportPhoto, pk=delete_photo_id, job_work1_report=report)
            photo_to_delete.delete()
            messages.success(request, 'Photo deleted successfully.')
        return redirect('edit_jobwork1_report', pk=report.id)

    if request.method == 'POST':
        _fields = ['jobworker', 'purpose', 'total_pcs', 'total_rate']
        _old = capture_snapshot(report, _fields)
        form = JobWork1ReportForm(request.POST, request.FILES, instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
        photos = request.FILES.getlist('photos')

        if len(photos) + report.photos.count() == 0:
            messages.error(request, 'At least one Job Card Photo is required.')
            return render(request, 'jobwork_form.html', {
                'form': form, 'cutting_reports': cutting_reports_qs,
                'cutting_reports_json': cutting_reports_json, 'is_admin': is_admin, 'is_edit': True, 'report': report
            })

        if len(photos) + report.photos.count() > 5:
            messages.error(request, 'You can upload a maximum of 5 photos total.')
            return render(request, 'jobwork_form.html', {
                'form': form, 'cutting_reports': cutting_reports_qs,
                'cutting_reports_json': cutting_reports_json, 'is_admin': is_admin, 'is_edit': True, 'report': report
            })

        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                if not report.total_rate:
                    report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos:
                JobWork1ReportPhoto.objects.create(
                    job_work1_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Update pending task status based on In/Out dates
            job_card_no = report.cutting_report.job_card_no
            if report.jobwork_out:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_jobwork1_done=True, is_jobwork1_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_jobwork1_in_progress=True, is_jobwork1_done=False
                )

            messages.success(request, 'Job Work 1 updated.')
            _diff = build_diff(_old, capture_snapshot(report, _fields))
            log_activity(request.user, 'EDIT', 'Job Work 1', report.cutting_report.job_card_no, _diff)
            return redirect('submission_list')
    else:
        form = JobWork1ReportForm(instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
    return render(request, 'jobwork_form.html', {
        'form': form, 'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin, 'is_edit': True, 'report': report
    })

@login_required
def delete_jobwork1_report(request, pk):
    report = get_object_or_404(JobWork1Report, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    if request.method == 'POST':
        jc_no = report.cutting_report.job_card_no
        _summary = get_obj_summary(report)
        JobCardRequirement.objects.filter(job_card_no=jc_no).update(
            is_jobwork1_done=False, is_jobwork1_in_progress=False
        )
        report.delete()
        messages.success(request, 'Job Work 1 deleted.')
        log_activity(request.user, 'DELETE', 'Job Work 1', jc_no, _summary)
        return redirect('submission_list')
    return render(request, 'confirm_delete.html', {'object': report, 'cancel_url': 'submission_list'})


@login_required
def edit_embroidery_report(request, pk):
    report = get_object_or_404(EmbroideryReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    is_admin = request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN')
    cutting_reports_qs = CuttingReport.objects.filter(Q(embroidery_reports__isnull=True) | Q(id=report.cutting_report_id)).distinct().select_related('master_entry').order_by('-created_at')
    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })
    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })
    delete_photo_id = request.GET.get('delete_photo')
    if delete_photo_id:
        photo_to_delete = get_object_or_404(EmbroideryReportPhoto, pk=delete_photo_id, embroidery_report=report)
        photo_to_delete.delete()
        messages.success(request, 'Photo deleted successfully.')
        return redirect('edit_embroidery_report', pk=report.id)

    if request.method == 'POST':
        _fields = ['embroidery_worker', 'purpose', 'total_pcs', 'total_rate']
        _old = capture_snapshot(report, _fields)
        form = EmbroideryReportForm(request.POST, request.FILES, instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
        photos = request.FILES.getlist('photos')

        if len(photos) + report.photos.count() > 5:
            messages.error(request, 'You can upload a maximum of 5 photos total.')
            return redirect('edit_embroidery_report', pk=report.id)

        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                if not report.total_rate:
                    report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos:
                EmbroideryReportPhoto.objects.create(
                    embroidery_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Update pending task status based on In/Out dates
            job_card_no = report.cutting_report.job_card_no
            if report.embroidery_out:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_embroidery_done=True, is_embroidery_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_embroidery_in_progress=True, is_embroidery_done=False
                )

            messages.success(request, 'Embroidery updated.')
            _diff = build_diff(_old, capture_snapshot(report, _fields))
            log_activity(request.user, 'EDIT', 'Embroidery', report.cutting_report.job_card_no, _diff)
            return redirect('submission_list')
    else:
        form = EmbroideryReportForm(instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
    return render(request, 'embroidery_form.html', {
        'form': form, 'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin, 'is_edit': True, 'report': report
    })

@login_required
def delete_embroidery_report(request, pk):
    report = get_object_or_404(EmbroideryReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    if request.method == 'POST':
        jc_no = report.cutting_report.job_card_no
        _summary = get_obj_summary(report)
        JobCardRequirement.objects.filter(job_card_no=jc_no).update(
            is_embroidery_done=False, is_embroidery_in_progress=False
        )
        report.delete()
        messages.success(request, 'Embroidery deleted.')
        log_activity(request.user, 'DELETE', 'Embroidery', jc_no, _summary)
        return redirect('submission_list')
    return render(request, 'confirm_delete.html', {'object': report, 'cancel_url': 'submission_list'})

@login_required
def edit_printing_report(request, pk):
    report = get_object_or_404(PrintingReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    is_admin = request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN')
    cutting_reports_qs = CuttingReport.objects.filter(Q(printing_reports__isnull=True) | Q(id=report.cutting_report_id)).distinct().select_related('master_entry').order_by('-created_at')
    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'job_card_no': cr.job_card_no,
            'item_name': cr.item_name,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })
    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })
    delete_photo_id = request.GET.get('delete_photo')
    if delete_photo_id:
        photo_to_delete = get_object_or_404(PrintingReportPhoto, pk=delete_photo_id, printing_report=report)
        photo_to_delete.delete()
        messages.success(request, 'Photo deleted successfully.')
        return redirect('edit_printing_report', pk=report.id)

    if request.method == 'POST':
        _fields = ['printing_worker', 'purpose', 'total_pcs', 'total_rate']
        _old = capture_snapshot(report, _fields)
        form = PrintingReportForm(request.POST, request.FILES, instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
        photos = request.FILES.getlist('photos')

        if len(photos) + report.photos.count() > 5:
            messages.error(request, 'You can upload a maximum of 5 photos total.')
            return redirect('edit_printing_report', pk=report.id)

        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                if not report.total_rate:
                    report.total_rate = report.rate_definition.total_rate
            report.save()

            for p in photos:
                PrintingReportPhoto.objects.create(
                    printing_report=report,
                    photo_data=p.read(),
                    photo_name=p.name,
                    photo_content_type=p.content_type
                )

            # Update pending task status based on In/Out dates
            job_card_no = report.cutting_report.job_card_no
            if report.printing_out:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_printing_done=True, is_printing_in_progress=False
                )
            else:
                JobCardRequirement.objects.filter(job_card_no=job_card_no).update(
                    is_printing_in_progress=True, is_printing_done=False
                )

            messages.success(request, 'Printing updated.')
            _diff = build_diff(_old, capture_snapshot(report, _fields))
            log_activity(request.user, 'EDIT', 'Printing', report.cutting_report.job_card_no, _diff)
            return redirect('submission_list')
    else:
        form = PrintingReportForm(instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
    return render(request, 'printing_form.html', {
        'form': form, 'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin, 'is_edit': True, 'report': report
    })

@login_required
def delete_printing_report(request, pk):
    report = get_object_or_404(PrintingReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    if request.method == 'POST':
        jc_no = report.cutting_report.job_card_no
        _summary = get_obj_summary(report)
        JobCardRequirement.objects.filter(job_card_no=jc_no).update(
            is_printing_done=False, is_printing_in_progress=False
        )
        report.delete()
        messages.success(request, 'Printing deleted.')
        log_activity(request.user, 'DELETE', 'Printing', jc_no, _summary)
        return redirect('submission_list')
    return render(request, 'confirm_delete.html', {'object': report, 'cancel_url': 'submission_list'})

@login_required
def edit_finishing_report(request, pk):
    report = get_object_or_404(FinishingReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    
    # Handle photo deletion if requested via URL param
    delete_photo_id = request.GET.get('delete_photo')
    if delete_photo_id:
        photo_to_delete = get_object_or_404(FinishingReportPhoto, pk=delete_photo_id, finishing_report=report)
        photo_to_delete.delete()
        messages.success(request, 'Photo deleted successfully.')
        return redirect(request.path)

    is_admin = request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN')
    cutting_reports_qs = CuttingReport.objects.filter(Q(finishing_reports__isnull=True) | Q(id=report.cutting_report_id)).distinct().select_related('master_entry').order_by('-created_at')
    cutting_reports_json = json.dumps({
        str(cr.id): {
            'master_entry_id': cr.master_entry_id,
            'date': cr.master_entry.date.strftime('%Y-%m-%d'),
            'lot_no': cr.job_card_no,
            'total_pcs': cr.total_pcs
        } for cr in cutting_reports_qs
    })

    rate_definitions = RateDefinition.objects.all()
    rate_definitions_json = json.dumps({
        str(r.id): {
            'name': r.name,
            'description': r.description,
            'total_rate': str(r.total_rate)
        } for r in rate_definitions
    })

    if request.method == 'POST':
        _fields = ['lot_no', 'total_pcs', 'total_pcs_packed', 'total_rate']
        _old = capture_snapshot(report, _fields)
        form = FinishingReportForm(request.POST, request.FILES, instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            if report.rate_definition:
                if not report.total_rate:
                    report.total_rate = report.rate_definition.total_rate
            report.save()
            photos = request.FILES.getlist('photos')
            for photo_file in photos:
                if report.photos.count() < 5:
                    FinishingReportPhoto.objects.create(
                        finishing_report=report,
                        photo_data=photo_file.read(),
                        photo_name=photo_file.name,
                        photo_content_type=photo_file.content_type
                    )
            messages.success(request, 'Finishing Report updated.')
            _diff = build_diff(_old, capture_snapshot(report, _fields))
            log_activity(request.user, 'EDIT', 'Finishing', report.cutting_report.job_card_no, _diff)
            return redirect('submission_list')
    else:
        form = FinishingReportForm(instance=report)
        form.fields['cutting_report'].queryset = cutting_reports_qs
    return render(request, 'finishing_form.html', {
        'form': form, 'cutting_reports': cutting_reports_qs,
        'cutting_reports_json': cutting_reports_json,
        'rate_definitions_json': rate_definitions_json,
        'is_admin': is_admin, 'is_edit': True, 'report': report
    })

@login_required
def delete_finishing_report(request, pk):
    report = get_object_or_404(FinishingReport, pk=pk)
    if not check_edit_permission(request, report): raise PermissionDenied
    if request.method == 'POST':
        jc_no = report.cutting_report.job_card_no
        _summary = get_obj_summary(report)
        JobCardRequirement.objects.filter(job_card_no=jc_no).update(is_finishing_done=False)
        report.delete()
        messages.success(request, 'Finishing Report deleted.')
        log_activity(request.user, 'DELETE', 'Finishing', jc_no, _summary)
        return redirect('submission_list')
    return render(request, 'confirm_delete.html', {'object': report, 'cancel_url': 'submission_list'})

from django.utils import timezone

# ── Activity Log View ─────────────────────────────────────────────────────────

@login_required
def activity_log_view(request):
    """Admin-only: shows a paginated, filterable history of all actions."""
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        raise PermissionDenied

    from django.contrib.auth.models import User as AuthUser

    logs = ActivityLog.objects.select_related('user').all()

    # ── Filters ─────────────────────────────────────────────────────────────
    filter_user     = request.GET.get('user', '').strip()
    filter_action   = request.GET.get('action', '').strip()
    filter_dept     = request.GET.get('department', '').strip()
    filter_jc       = request.GET.get('job_card_no', '').strip()
    filter_date     = request.GET.get('date', '').strip()

    if filter_user:
        logs = logs.filter(user__username__icontains=filter_user)
    if filter_action:
        logs = logs.filter(action=filter_action)
    if filter_dept:
        logs = logs.filter(department__icontains=filter_dept)
    if filter_jc:
        logs = logs.filter(job_card_no__icontains=filter_jc)
    if filter_date:
        try:
            from datetime import datetime
            date_obj = datetime.strptime(filter_date, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date=date_obj)
        except ValueError:
            pass

    paginator   = Paginator(logs, 15)
    page_number = request.GET.get('page', 1)
    page_obj    = paginator.get_page(page_number)

    all_users = AuthUser.objects.filter(activitylog__isnull=False).distinct().order_by('username')
    departments = ActivityLog.objects.values_list('department', flat=True).distinct().order_by('department')

    return render(request, 'activity_log.html', {
        'page_obj':      page_obj,
        'all_users':     all_users,
        'departments':   departments,
        'filter_user':   filter_user,
        'filter_action': filter_action,
        'filter_dept':   filter_dept,
        'filter_jc':     filter_jc,
        'filter_date':   filter_date,
        'total_count':   logs.count(),
    })


@login_required
def download_database(request):
    if not request.user.is_superuser:
        raise PermissionDenied("Only superusers can download the database backup.")
    
    try:
        zip_buffer = generate_backup_zip()
        
        # Update last downloaded time
        system_settings = SystemSetting.get_settings()
        system_settings.last_excel_download_at = timezone.now()
        system_settings.save()
        
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="FabricTrack_Backup.zip"'
        return response
    except Exception as e:
        messages.error(request, f'Backup failed: {e}')
        return redirect('dashboard')


def serve_db_image(request, model_name, photo_id):
    if model_name == 'cutting':
        photo = get_object_or_404(CuttingReportPhoto, pk=photo_id)
    elif model_name == 'finishing':
        photo = get_object_or_404(FinishingReportPhoto, pk=photo_id)
    elif model_name == 'p4':
        photo = get_object_or_404(StitchingReportPhoto, pk=photo_id)
    elif model_name == 'jobwork':
        photo = get_object_or_404(JobWorkReportPhoto, pk=photo_id)
    elif model_name == 'jobwork1':
        photo = get_object_or_404(JobWork1ReportPhoto, pk=photo_id)
    elif model_name == 'embroidery':
        photo = get_object_or_404(EmbroideryReportPhoto, pk=photo_id)
    elif model_name == 'printing':
        photo = get_object_or_404(PrintingReportPhoto, pk=photo_id)
    elif model_name == 'singleneedle':
        photo = get_object_or_404(SingleneedleReportPhoto, pk=photo_id)
    elif model_name == 'sewing':
        photo = get_object_or_404(SewingReportPhoto, pk=photo_id)
    elif model_name == 'sewing1':
        photo = get_object_or_404(Sewing1ReportPhoto, pk=photo_id)
    elif model_name == 'accessories':
        photo = get_object_or_404(AccessoriesPhoto, pk=photo_id)
    else:
        raise Http404("Invalid photo model")
    
    response = HttpResponse(photo.photo_data, content_type=photo.photo_content_type)
    response['Content-Disposition'] = f'inline; filename="{photo.photo_name}"'
    response['Cache-Control'] = 'public, max-age=31536000'  # Cache for 1 year
    return response


def serve_accessories_cell_photo(request, entry_id, col):
    from .models import AccessoriesItemEntry
    entry = get_object_or_404(AccessoriesItemEntry, pk=entry_id)
    c = col.lower()
    photo_data = getattr(entry, f'photo_data_{c}', None)
    photo_name = getattr(entry, f'photo_name_{c}', None) or 'photo.jpg'
    content_type = getattr(entry, f'photo_content_type_{c}', None) or 'image/jpeg'
    
    if not photo_data:
        raise Http404("Photo not found")
        
    response = HttpResponse(photo_data, content_type=content_type)
    response['Content-Disposition'] = f'inline; filename="{photo_name}"'
    response['Cache-Control'] = 'public, max-age=31536000'
    return response



@login_required
def serve_vendor_photo(request, vendor_id):
    """Serve vendor photo stored as binary in MasterName model."""
    from .models import MasterName
    vendor = get_object_or_404(MasterName, pk=vendor_id, department='Vendor')
    if not vendor.photo:
        raise Http404("Photo not found")
    content_type = vendor.photo_mime or 'image/jpeg'
    response = HttpResponse(bytes(vendor.photo), content_type=content_type)
    response['Content-Disposition'] = f'inline; filename="vendor_{vendor_id}.jpg"'
    # No caching — always serve the latest photo from DB
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    response['Pragma'] = 'no-cache'
    return response


@login_required
def reset_database_view(request):
    if not request.user.is_superuser:
        raise PermissionDenied("Only superusers can reset the database.")
    
    if request.method == 'POST':
        # Delete MasterEntry (cascades to all other data including reports, colors, photos)
        MasterEntry.objects.all().delete()
        
        # Reset last downloaded time
        system_settings = SystemSetting.get_settings()
        system_settings.last_excel_download_at = None
        system_settings.save()
        

            
        messages.success(request, 'Database successfully reset! All master entries, reports, and photos have been permanently deleted.')
        return redirect('dashboard')
        
    return render(request, 'confirm_reset.html')


from django.db.models import Sum
from .forms import MasterPaymentForm
from .models import MasterPayment, MasterName
from .utils import calculate_master_earnings

# ── Ledger Password Lock ──────────────────────────────────────────────────────

_LEDGER_SESSION_KEY = 'ledger_unlocked'

def _ledger_is_unlocked(request):
    return request.session.get(_LEDGER_SESSION_KEY) is True

@login_required
def ledger_unlock_view(request):
    """Password gate for the Master Ledger section."""
    from django.conf import settings as dj_settings
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        raise PermissionDenied

    if _ledger_is_unlocked(request):
        return redirect('master_ledger_list')

    error = None
    if request.method == 'POST':
        entered = request.POST.get('password', '')
        if entered == dj_settings.LEDGER_PASSWORD:
            request.session[_LEDGER_SESSION_KEY] = True
            next_url = request.POST.get('next') or request.GET.get('next') or ''
            return redirect(next_url if next_url else 'master_ledger_list')
        else:
            error = 'Incorrect password. Please try again.'

    return render(request, 'ledger_unlock.html', {
        'error': error,
        'next': request.GET.get('next', ''),
    })

@login_required
def ledger_lock_view(request):
    """Clears the ledger session key — locks the section again."""
    request.session.pop(_LEDGER_SESSION_KEY, None)
    messages.success(request, 'Master Ledger has been locked.')
    return redirect('ledger_unlock')

# ── Ledger Views ──────────────────────────────────────────────────────────────

@login_required
def master_ledger_list_view(request):
    if not _ledger_is_unlocked(request):
        return redirect(f'/ledger/unlock/?next=/ledger/')
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        raise PermissionDenied("Only administrators can view the ledger.")

    from django.utils.dateparse import parse_date
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str   = request.GET.get('end_date', '').strip()
    item_name_str  = request.GET.get('item_name', '').strip()

    start_date = parse_date(start_date_str) if start_date_str else None
    end_date   = parse_date(end_date_str)   if end_date_str   else None

    masters = MasterName.objects.all().order_by('department', 'name')

    # Filter by item_name (article field on MasterName)
    if item_name_str:
        masters = masters.filter(
            Q(article__icontains=item_name_str) |
            Q(name__icontains=item_name_str)
        )

    ledger_data = []

    for master in masters:
        total_earnings = calculate_master_earnings(master.name, start_date, end_date)
        
        p_qs = master.payments.all()
        if start_date:
            p_qs = p_qs.filter(date__gte=start_date)
        if end_date:
            p_qs = p_qs.filter(date__lte=end_date)
            
        total_paid = float(p_qs.aggregate(total=Sum('amount'))['total'] or 0.0)
        balance = total_earnings - total_paid

        # Hide inactive masters in the filtered view
        if (start_date or end_date) and total_earnings == 0.0 and total_paid == 0.0 and balance == 0.0:
            continue

        ledger_data.append({
            'master': master,
            'total_earnings': total_earnings,
            'total_paid': total_paid,
            'balance': balance,
        })

    return render(request, 'master_ledger_list.html', {
        'ledger_data': ledger_data,
        'is_admin': is_admin,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'item_name': item_name_str,
    })


@login_required
def master_ledger_detail_view(request, pk):
    if not _ledger_is_unlocked(request):
        return redirect(f'/ledger/unlock/?next=/ledger/{pk}/')
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        raise PermissionDenied("Only administrators can view the ledger.")

    master = get_object_or_404(MasterName, pk=pk)

    from django.utils.dateparse import parse_date
    start_date_str  = request.GET.get('start_date', '').strip()
    end_date_str    = request.GET.get('end_date', '').strip()
    item_name_filter = request.GET.get('item_name', '').strip()

    start_date_obj = parse_date(start_date_str) if start_date_str else None
    end_date_obj   = parse_date(end_date_str)   if end_date_str   else None

    from .models import (
        CuttingReport, StitchingReport, JobWorkReport, EmbroideryReport,
        PrintingReport, SingleneedleReport, SewingReport, FinishingReport
    )

    events = []

    # 1. Cutting Report
    for r in CuttingReport.objects.filter(master_name=master.name).select_related('master_entry'):
        rate = float(r.cutting_rate or 0.0)
        pcs = int(r.total_pcs or 0)
        amount = rate * pcs
        if amount > 0:
            item_label = f" [{r.item_name}]" if getattr(r, 'item_name', '') else ''
            events.append({
                'date': r.master_entry.date,
                'created_at': r.created_at,
                'type': 'earning',
                'item_name': getattr(r, 'item_name', '') or '',
                'description': f"Cutting{item_label}: {pcs} Pcs @ ₹{rate:.2f} (Job Card: {r.job_card_no})",
                'amount': amount
            })

    # Helper for standard reports
    def add_reports(qs, label, date_field, rate_field='total_rate', jc_field='job_card_no'):
        for r in qs:
            rate = float(getattr(r, rate_field) or 0.0)
            pcs = int(r.total_pcs or 0)
            amount = rate * pcs
            if amount > 0:
                d = getattr(r, date_field) or r.created_at.date()
                jc = getattr(r, jc_field, '')
                item_label = f" [{r.item_name}]" if getattr(r, 'item_name', '') else ''
                events.append({
                    'date': d,
                    'created_at': r.created_at,
                    'type': 'earning',
                    'item_name': getattr(r, 'item_name', '') or '',
                    'description': f"{label}{item_label}: {pcs} Pcs @ ₹{rate:.2f} (Job Card/Lot: {jc})",
                    'amount': amount
                })

    add_reports(StitchingReport.objects.filter(master_name=master.name), "Stitching", "line_in_date")
    add_reports(JobWorkReport.objects.filter(master_name=master.name), "Job Work", "jobwork_in")
    add_reports(EmbroideryReport.objects.filter(master_name=master.name), "Embroidery", "embroidery_in")
    add_reports(PrintingReport.objects.filter(master_name=master.name), "Printing", "printing_in")
    add_reports(SingleneedleReport.objects.filter(master_name=master.name), "Singleneedle", "line_in_date")
    add_reports(SewingReport.objects.filter(master_name=master.name), "Sewing", "line_in_date")
    add_reports(FinishingReport.objects.filter(master_name=master.name), "Finishing", "date", jc_field='lot_no')

    # Payments
    for p in MasterPayment.objects.filter(master=master):
        period_str = f" (Period: {p.start_date.strftime('%d %b %Y')} to {p.end_date.strftime('%d %b %Y')})" if p.start_date and p.end_date else ""
        events.append({
            'date': p.date,
            'created_at': p.created_at,
            'type': 'payment',
            'item_name': '',
            'description': f"Paid via {p.get_payment_mode_display()}{period_str}" + (f" (Ref: {p.reference_no})" if p.reference_no else "") + (f" - {p.remarks}" if p.remarks else ""),
            'amount': float(p.amount),
            'payment_id': p.id
        })

    # Sort chronologically
    events = sorted(events, key=lambda x: (x['date'], x['created_at']))

    # Split into pre-range and active range
    pre_events = []
    active_events = []
    
    for e in events:
        if start_date_obj and e['date'] < start_date_obj:
            pre_events.append(e)
        elif end_date_obj and e['date'] > end_date_obj:
            pass
        else:
            active_events.append(e)

    # Compute opening balance
    opening_balance = sum(e['amount'] for e in pre_events if e['type'] == 'earning') - sum(e['amount'] for e in pre_events if e['type'] == 'payment')

    running_balance = opening_balance
    for e in active_events:
        if e['type'] == 'earning':
            running_balance += e['amount']
        else:
            running_balance -= e['amount']
        e['balance'] = running_balance

    # Overall totals
    total_earnings_all = calculate_master_earnings(master.name)
    total_paid_all = float(master.payments.aggregate(total=Sum('amount'))['total'] or 0.0)
    current_balance_all = total_earnings_all - total_paid_all

    # Filtered range totals
    range_earnings = sum(e['amount'] for e in active_events if e['type'] == 'earning')
    range_paid = sum(e['amount'] for e in active_events if e['type'] == 'payment')
    range_balance_change = range_earnings - range_paid

    # Apply item_name filter to active events (earnings only, payments always shown)
    if item_name_filter:
        active_events = [
            e for e in active_events
            if e['type'] == 'payment' or item_name_filter.lower() in e['item_name'].lower()
        ]

    return render(request, 'master_ledger_detail.html', {
        'master': master,
        'events': reversed(active_events),  # Show newest first in table
        'total_earnings': range_earnings if (start_date_obj or end_date_obj) else total_earnings_all,
        'total_paid': range_paid if (start_date_obj or end_date_obj) else total_paid_all,
        'current_balance': (opening_balance + range_balance_change) if (start_date_obj or end_date_obj) else current_balance_all,

        'total_earnings_all': total_earnings_all,
        'total_paid_all': total_paid_all,
        'current_balance_all': current_balance_all,
        'opening_balance': opening_balance,

        'start_date': start_date_str,
        'end_date': end_date_str,
        'item_name_filter': item_name_filter,
        'is_admin': is_admin,
    })


@login_required
def my_statement_unlock_view(request):
    """Password prompt for viewing a user's statement."""
    profile = getattr(request.user, 'profile', None)
    statement_password = getattr(profile, 'statement_password', '') if profile else ''

    if not statement_password:
        return redirect('my_statement')

    session_key = f"mystatement_unlocked_{request.user.id}"
    if request.session.get(session_key) is True:
        return redirect('my_statement')

    error = None
    if request.method == 'POST':
        entered = request.POST.get('password', '')
        if entered == statement_password:
            request.session[session_key] = True
            next_url = request.POST.get('next') or request.GET.get('next') or ''
            return redirect(next_url if next_url else 'my_statement')
        else:
            error = 'Incorrect statement password. Please try again.'

    return render(request, 'my_statement_unlock.html', {
        'error': error,
        'next': request.GET.get('next', ''),
    })


@login_required
def my_statement_lock_view(request):
    """Locks the statement page for this user."""
    session_key = f"mystatement_unlocked_{request.user.id}"
    request.session.pop(session_key, None)
    messages.success(request, 'Your statement view has been locked.')
    return redirect('my_statement')


@login_required
def my_statement_view(request):
    """
    Allows any logged-in user to view their own ledger statement.
    No password required if admin hasn't set one. Otherwise, prompts for password.
    """
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if is_admin:
        return redirect('master_ledger_list')

    profile = getattr(request.user, 'profile', None)
    statement_password = getattr(profile, 'statement_password', '') if profile else ''

    if statement_password:
        session_key = f"mystatement_unlocked_{request.user.id}"
        if request.session.get(session_key) is not True:
            return redirect(f'/my-statement/unlock/?next={request.path}')

    linked_masters = list(profile.linked_masters.all()) if profile else []

    if not linked_masters:
        return render(request, 'my_statement_no_link.html', {})

    # Select which master to show
    selected_master_id = request.GET.get('master_id', '').strip()
    master = None
    if selected_master_id:
        for lm in linked_masters:
            if str(lm.id) == selected_master_id:
                master = lm
                break

    if not master:
        master = linked_masters[0]

    from django.utils.dateparse import parse_date
    from .models import (
        CuttingReport, StitchingReport, JobWorkReport, EmbroideryReport,
        PrintingReport, SingleneedleReport, SewingReport, FinishingReport,
        MasterPayment,
    )

    start_date_str   = request.GET.get('start_date', '').strip()
    end_date_str     = request.GET.get('end_date', '').strip()
    item_name_filter = request.GET.get('item_name', '').strip()
    start_date_obj   = parse_date(start_date_str) if start_date_str else None
    end_date_obj     = parse_date(end_date_str)   if end_date_str   else None

    events = []

    # 1. Cutting Report
    for r in CuttingReport.objects.filter(master_name=master.name).select_related('master_entry'):
        rate = float(r.cutting_rate or 0.0)
        pcs = int(r.total_pcs or 0)
        amount = rate * pcs
        if amount > 0:
            item_label = f" [{r.item_name}]" if getattr(r, 'item_name', '') else ''
            events.append({
                'date': r.master_entry.date,
                'created_at': r.created_at,
                'type': 'earning',
                'item_name': getattr(r, 'item_name', '') or '',
                'description': f"Cutting{item_label}: {pcs} Pcs @ ₹{rate:.2f} (Job Card: {r.job_card_no})",
                'amount': amount
            })

    # Helper for standard reports
    def add_reports(qs, label, date_field, rate_field='total_rate', jc_field='job_card_no'):
        for r in qs:
            rate = float(getattr(r, rate_field) or 0.0)
            pcs = int(r.total_pcs or 0)
            amount = rate * pcs
            if amount > 0:
                d = getattr(r, date_field) or r.created_at.date()
                jc = getattr(r, jc_field, '')
                item_label = f" [{r.item_name}]" if getattr(r, 'item_name', '') else ''
                events.append({
                    'date': d,
                    'created_at': r.created_at,
                    'type': 'earning',
                    'item_name': getattr(r, 'item_name', '') or '',
                    'description': f"{label}{item_label}: {pcs} Pcs @ ₹{rate:.2f} (Job Card/Lot: {jc})",
                    'amount': amount
                })

    add_reports(StitchingReport.objects.filter(master_name=master.name), "Stitching", "line_in_date")
    add_reports(JobWorkReport.objects.filter(master_name=master.name), "Job Work", "jobwork_in")
    add_reports(EmbroideryReport.objects.filter(master_name=master.name), "Embroidery", "embroidery_in")
    add_reports(PrintingReport.objects.filter(master_name=master.name), "Printing", "printing_in")
    add_reports(SingleneedleReport.objects.filter(master_name=master.name), "Singleneedle", "line_in_date")
    add_reports(SewingReport.objects.filter(master_name=master.name), "Sewing", "line_in_date")
    add_reports(FinishingReport.objects.filter(master_name=master.name), "Finishing", "date", jc_field='lot_no')

    # Payments
    for p in MasterPayment.objects.filter(master=master):
        period_str = f" (Period: {p.start_date.strftime('%d %b %Y')} to {p.end_date.strftime('%d %b %Y')})" if p.start_date and p.end_date else ""
        events.append({
            'date': p.date,
            'created_at': p.created_at,
            'type': 'payment',
            'item_name': '',
            'description': f"Paid via {p.get_payment_mode_display()}{period_str}" + (f" (Ref: {p.reference_no})" if p.reference_no else "") + (f" - {p.remarks}" if p.remarks else ""),
            'amount': float(p.amount),
            'payment_id': p.id
        })

    # Sort chronologically
    events = sorted(events, key=lambda x: (x['date'], x['created_at']))

    # Split into pre-range and active range
    pre_events = []
    active_events = []
    
    for e in events:
        if start_date_obj and e['date'] < start_date_obj:
            pre_events.append(e)
        elif end_date_obj and e['date'] > end_date_obj:
            pass
        else:
            active_events.append(e)

    # Compute opening balance
    opening_balance = 0.0
    for e in pre_events:
        if e['type'] == 'earning':
            opening_balance += e['amount']
        else:
            opening_balance -= e['amount']

    # Calculate running balance starting from the very beginning (all-time)
    running = 0.0
    for e in events:
        if e['type'] == 'earning':
            running += e['amount']
        else:
            running -= e['amount']
        e['balance'] = running

    # Overall totals
    from .views import calculate_master_earnings
    from django.db.models import Sum
    total_earnings_all = calculate_master_earnings(master.name)
    total_paid_all = float(master.payments.aggregate(total=Sum('amount'))['total'] or 0.0)
    current_balance_all = total_earnings_all - total_paid_all

    # Filtered range totals
    range_earnings = sum(e['amount'] for e in active_events if e['type'] == 'earning')
    range_paid = sum(e['amount'] for e in active_events if e['type'] == 'payment')
    range_balance_change = range_earnings - range_paid

    # Apply item_name filter to active events (earnings only, payments always shown)
    if item_name_filter:
        active_events = [
            e for e in active_events
            if e['type'] == 'payment' or item_name_filter.lower() in e['item_name'].lower()
        ]

    return render(request, 'my_statement.html', {
        'master': master,
        'linked_masters': linked_masters,
        'events': reversed(active_events),  # Show newest first in table
        'total_earnings': range_earnings if (start_date_obj or end_date_obj) else total_earnings_all,
        'total_paid': range_paid if (start_date_obj or end_date_obj) else total_paid_all,
        'current_balance': (opening_balance + range_balance_change) if (start_date_obj or end_date_obj) else current_balance_all,

        'total_earnings_all': total_earnings_all,
        'total_paid_all': total_paid_all,
        'current_balance_all': current_balance_all,
        'opening_balance': opening_balance,

        'start_date': start_date_str,
        'end_date': end_date_str,
        'item_name_filter': item_name_filter,
        'is_admin': False,
    })




@login_required
def record_payment_view(request, pk=None):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        raise PermissionDenied("Only administrators can record payments.")

    from django.utils.dateparse import parse_date
    from django.urls import reverse
    
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str = request.GET.get('end_date', '').strip()

    start_date_obj = parse_date(start_date_str) if start_date_str else None
    end_date_obj = parse_date(end_date_str) if end_date_str else None

    initial_data = {
        'date': timezone.now().date(),
    }
    if start_date_obj:
        initial_data['start_date'] = start_date_obj
    if end_date_obj:
        initial_data['end_date'] = end_date_obj

    master = None
    if pk:
        master = get_object_or_404(MasterName, pk=pk)
        initial_data['master'] = master
        total_earnings = calculate_master_earnings(master.name, start_date_obj, end_date_obj)
        p_qs = master.payments.all()
        if start_date_obj:
            p_qs = p_qs.filter(date__gte=start_date_obj)
        if end_date_obj:
            p_qs = p_qs.filter(date__lte=end_date_obj)
        total_paid = float(p_qs.aggregate(total=Sum('amount'))['total'] or 0.0)
        initial_data['amount'] = max(0.0, total_earnings - total_paid)

    if request.method == 'POST':
        form = MasterPaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.created_by = request.user
            payment.save()
            messages.success(request, f"Successfully recorded payment of ₹{payment.amount} to {payment.master.name}!")
            
            detail_url = reverse('master_ledger_detail', kwargs={'pk': payment.master.pk})
            if start_date_str or end_date_str:
                query_params = []
                if start_date_str: query_params.append(f"start_date={start_date_str}")
                if end_date_str: query_params.append(f"end_date={end_date_str}")
                detail_url += "?" + "&".join(query_params)
            return redirect(detail_url)
    else:
        form = MasterPaymentForm(initial=initial_data)

    masters_json = json.dumps({
        str(m.id): {
            'name': m.name,
            'upi_id': m.upi_id or '',
            'outstanding': max(0.0, calculate_master_earnings(m.name, start_date_obj, end_date_obj) - float(
                m.payments.filter(
                    **({'date__gte': start_date_obj} if start_date_obj else {}),
                    **({'date__lte': end_date_obj} if end_date_obj else {})
                ).aggregate(total=Sum('amount'))['total'] or 0.0
            ))
        } for m in MasterName.objects.all()
    })

    return render(request, 'record_payment.html', {
        'form': form,
        'master': master,
        'masters_json': masters_json,
        'is_admin': is_admin,
        'start_date': start_date_str,
        'end_date': end_date_str,
    })


@login_required
def delete_payment_view(request, pk):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        raise PermissionDenied("Only administrators can delete payments.")

    payment = get_object_or_404(MasterPayment, pk=pk)
    master_pk = payment.master.pk
    if request.method == 'POST':
        payment.delete()
        messages.success(request, "Payment deleted successfully.")
        return redirect('master_ledger_detail', pk=master_pk)
    return render(request, 'confirm_delete.html', {'object': payment, 'cancel_url': 'master_ledger_detail'})


@login_required
def get_master_outstanding_api(request):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        from django.http import JsonResponse
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    from django.http import JsonResponse

    from django.utils.dateparse import parse_date
    master_id = request.GET.get('master_id')
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str = request.GET.get('end_date', '').strip()

    if not master_id:
        return JsonResponse({'error': 'master_id is required'}, status=400)

    start_date = parse_date(start_date_str) if start_date_str else None
    end_date = parse_date(end_date_str) if end_date_str else None

    master = get_object_or_404(MasterName, id=master_id)

    total_earnings = calculate_master_earnings(master.name, start_date, end_date)

    p_qs = master.payments.all()
    if start_date:
        p_qs = p_qs.filter(date__gte=start_date)
    if end_date:
        p_qs = p_qs.filter(date__lte=end_date)

    total_paid = float(p_qs.aggregate(total=Sum('amount'))['total'] or 0.0)
    outstanding = max(0.0, total_earnings - total_paid)

    return JsonResponse({
        'master_id': master.id,
        'name': master.name,
        'upi_id': master.upi_id or '',
        'outstanding': outstanding
    })



# ── Accessories Views ──────────────────────────────────────────────────────

@login_required
def accessories_view(request):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    from django.db.models import Sum, Max
    jc_data = (
        CuttingReport.objects
        .values('job_card_no')
        .annotate(total=Sum('total_pcs'), item_name=Max('item_name'))
        .order_by('job_card_no')
    )
    existing = {r.job_card_no: r for r in AccessoriesRecord.objects.prefetch_related('entries', 'photos').all()}
    cutting_reports = {cr.job_card_no: cr for cr in CuttingReport.objects.prefetch_related('photos').all()}

    job_cards = []
    yellow_jc_count = 0
    red_jc_count = 0
    green_jc_count = 0

    for jc in jc_data:
        rec = existing.get(jc['job_card_no'])
        cr = cutting_reports.get(jc['job_card_no'])
        if rec and rec.is_started:
            status = 'complete' if rec.is_complete else 'pending'
        else:
            status = 'new'

        if status == 'complete':
            green_jc_count += 1
        elif status == 'pending':
            yellow_jc_count += 1
        else:
            red_jc_count += 1
            
        # Collect photos from both cutting report and accessories record
        job_card_photos = []
        if cr:
            for p in cr.photos.all():
                job_card_photos.append({'id': p.id, 'type': 'cutting'})
        if rec:
            for p in rec.photos.all():
                job_card_photos.append({'id': p.id, 'type': 'accessories'})

        job_cards.append({
            'job_card_no': jc['job_card_no'],
            'total_pcs':   jc['total'],
            'item_name':   jc['item_name'] or '—',
            'record':      rec,
            'status':      status,
            'photos':      job_card_photos,
        })

    # Find all unique item names in database entries, subtract standard ones, and sort the rest
    db_item_names = set(AccessoriesItemEntry.objects.values_list('item_name', flat=True))
    custom_names = sorted(list(db_item_names - set(ACCESSORIES_ITEMS)))
    all_accessories_list = list(ACCESSORIES_ITEMS) + custom_names

    # Count cell-level statuses across all A, B, C, D columns of all entries
    cell_yellow_count = 0
    cell_red_count = 0
    cell_green_count = 0
    total_cells_count = 0

    for entry in AccessoriesItemEntry.objects.all():
        for col in ['a', 'b', 'c', 'd']:
            st = getattr(entry, f'status_{col}', '')
            qty = getattr(entry, f'qty_{col}', None)
            if st == 'yellow':
                cell_yellow_count += 1
                total_cells_count += 1
            elif st == 'green':
                cell_green_count += 1
                total_cells_count += 1
            elif st == 'red':
                cell_red_count += 1
                total_cells_count += 1
            elif qty is not None and qty > 0:
                cell_red_count += 1
                total_cells_count += 1

    kpi_summary = {
        'total_jc': len(job_cards),
        'yellow_jc': yellow_jc_count,
        'red_jc': red_jc_count,
        'green_jc': green_jc_count,
        'cell_yellow': cell_yellow_count,
        'cell_red': cell_red_count,
        'cell_green': cell_green_count,
        'total_cells': total_cells_count,
    }

    return render(request, 'accessories.html', {
        'job_cards': job_cards,
        'all_accessories_list': all_accessories_list,
        'kpi_summary': kpi_summary,
        'is_admin': is_admin
    })


@login_required
def accessories_detail_view(request, job_card_no):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    from django.db.models import Sum, Max
    total_pcs = CuttingReport.objects.filter(job_card_no=job_card_no).aggregate(t=Sum('total_pcs'))['t'] or 0

    record, created = AccessoriesRecord.objects.get_or_create(
        job_card_no=job_card_no,
        defaults={'total_pcs': total_pcs, 'created_by': request.user}
    )
    if not created:
        record.total_pcs = total_pcs
        record.save(update_fields=['total_pcs', 'updated_at'])

    # Handle photo deletion if requested via URL param
    delete_photo_id = request.GET.get('delete_photo')
    if delete_photo_id:
        photo_to_delete = get_object_or_404(AccessoriesPhoto, pk=delete_photo_id, accessories_record=record)
        photo_to_delete.delete()
        messages.success(request, 'Photo deleted successfully.')
        return redirect('accessories_detail', job_card_no=job_card_no)

    # Ensure all standard accessory entries exist for this record
    existing_entries = {e.item_name: e for e in record.entries.all()}
    for name in ACCESSORIES_ITEMS:
        if name not in existing_entries:
            max_sr = record.entries.aggregate(m=Max('sr_no'))['m'] or 0
            AccessoriesItemEntry.objects.create(record=record, sr_no=max_sr + 1, item_name=name)

    # Re-fetch and normalize sr_no order
    entries = list(record.entries.order_by('sr_no'))
    for i, entry in enumerate(entries):
        if entry.sr_no != i + 1:
            entry.sr_no = i + 1
            entry.save(update_fields=['sr_no'])

    if request.method == 'POST':
        if 'add_accessory_item' in request.POST:
            new_name = request.POST.get('new_item_name', '').strip().upper()
            if new_name:
                if not record.entries.filter(item_name=new_name).exists():
                    max_sr = record.entries.aggregate(m=Max('sr_no'))['m'] or 0
                    AccessoriesItemEntry.objects.create(record=record, sr_no=max_sr + 1, item_name=new_name)
                    messages.success(request, f'Accessory item "{new_name}" added.')
                else:
                    messages.warning(request, f'Accessory item "{new_name}" already exists on this job card.')
            return redirect('accessories_detail', job_card_no=job_card_no)

        elif 'delete_accessory_item' in request.POST:
            item_id = request.POST.get('item_id')
            if item_id:
                entry = record.entries.filter(id=item_id).first()
                if entry:
                    if entry.item_name not in ACCESSORIES_ITEMS:
                        entry.delete()
                        messages.success(request, 'Accessory item deleted.')
                    else:
                        messages.error(request, 'Standard accessory items cannot be deleted.')
            return redirect('accessories_detail', job_card_no=job_card_no)

        # Standard save accessories form
        for entry in entries:
            prefix = f'item_{entry.sr_no}'

            def get_dec(key, p=prefix):
                val = request.POST.get(f'{p}_{key}', '').strip()
                try:
                    return float(val) if val else None
                except ValueError:
                    return None

            entry.qty_a  = get_dec('a')
            entry.qty_b  = get_dec('b')
            entry.qty_c  = get_dec('c')
            entry.qty_d  = get_dec('d')
            entry.status_a = request.POST.get(f'{prefix}_status_a', '')
            entry.status_b = request.POST.get(f'{prefix}_status_b', '')
            entry.status_c = request.POST.get(f'{prefix}_status_c', '')
            entry.status_d = request.POST.get(f'{prefix}_status_d', '')
            
            # Save vendor and article values (always save or only if status is yellow)
            entry.vendor_a  = request.POST.get(f'{prefix}_vendor_a', '').strip() or None
            entry.vendor_b  = request.POST.get(f'{prefix}_vendor_b', '').strip() or None
            entry.vendor_c  = request.POST.get(f'{prefix}_vendor_c', '').strip() or None
            entry.vendor_d  = request.POST.get(f'{prefix}_vendor_d', '').strip() or None
            entry.article_a = request.POST.get(f'{prefix}_article_a', '').strip() or None
            entry.article_b = request.POST.get(f'{prefix}_article_b', '').strip() or None
            entry.article_c = request.POST.get(f'{prefix}_article_c', '').strip() or None
            entry.article_d = request.POST.get(f'{prefix}_article_d', '').strip() or None

            # Save per-cell yellow-status photos if uploaded
            for col in ['a', 'b', 'c', 'd']:
                file_key = f'{prefix}_photo_{col}'
                photo_file = request.FILES.get(file_key)
                if photo_file:
                    setattr(entry, f'photo_data_{col}', photo_file.read())
                    setattr(entry, f'photo_name_{col}', photo_file.name)
                    setattr(entry, f'photo_content_type_{col}', photo_file.content_type)

            entry.save()


        record.notes = request.POST.get('notes', '').strip()
        record.save(update_fields=['notes', 'updated_at'])

        # Save uploaded photos up to 5
        photos = request.FILES.getlist('photos')
        uploaded_count = 0
        skipped_count = 0
        for photo_file in photos:
            if record.photos.count() < 5:
                AccessoriesPhoto.objects.create(
                    accessories_record=record,
                    photo_data=photo_file.read(),
                    photo_name=photo_file.name,
                    photo_content_type=photo_file.content_type
                )
                uploaded_count += 1
            else:
                skipped_count += 1
        
        if uploaded_count > 0:
            messages.success(request, f'{uploaded_count} photo(s) uploaded successfully.')
        if skipped_count > 0:
            messages.warning(request, f'{skipped_count} photo(s) skipped because the limit is max 5 photos.')

        messages.success(request, f'Accessories for {job_card_no} saved successfully.')
        return redirect('accessories_detail', job_card_no=job_card_no)

    custom_names_obj = record.entries.exclude(item_name__in=ACCESSORIES_ITEMS).order_by('sr_no')
    
    # Query masters in Vendor department and group articles by vendor
    from collections import defaultdict
    from .models import MasterName
    vendor_groups = defaultdict(list)
    # Build vendor photo map: {vendor_name: {article: photo_url}}
    vendor_photo_map = defaultdict(dict)
    for m in MasterName.objects.filter(department='Vendor').order_by('name', 'article'):
        if m.article and m.article not in vendor_groups[m.name]:
            vendor_groups[m.name].append(m.article)
        if m.article and m.photo:
            from django.urls import reverse as url_reverse
            import time as _time
            photo_url = url_reverse('serve_vendor_photo', args=[m.id])
            vendor_photo_map[m.name][m.article] = f'{photo_url}?t={int(_time.time())}'
            
    import json
    vendor_groups_json = json.dumps(dict(vendor_groups))
    vendor_photo_map_json = json.dumps({k: dict(v) for k, v in vendor_photo_map.items()})

    return render(request, 'accessories_detail.html', {
        'record': record, 'entries': entries, 'is_admin': is_admin,
        'custom_names_obj': custom_names_obj,
        'vendor_groups_json': vendor_groups_json,
        'vendor_photo_map_json': vendor_photo_map_json,
    })


@login_required
def accessories_print_view(request, job_card_no):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    record  = get_object_or_404(AccessoriesRecord, job_card_no=job_card_no)
    entries = list(record.entries.order_by('sr_no'))
    return render(request, 'accessories_print.html', {'record': record, 'entries': entries})


@login_required
@require_POST
def accessories_add_item_view(request):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    item_name = request.POST.get('new_item_name', '').strip().upper()
    if item_name:
        if AccessoryCustomName.objects.filter(name=item_name).exists() or item_name in ACCESSORIES_ITEMS:
            messages.error(request, f'Item "{item_name}" already exists.')
        else:
            AccessoryCustomName.objects.create(name=item_name)
            messages.success(request, f'Global Accessory Item "{item_name}" added successfully.')
    else:
        messages.error(request, 'Item name cannot be empty.')
    return redirect('accessories')


@login_required
@require_POST
def accessories_delete_item_view(request, pk):
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    custom_name = get_object_or_404(AccessoryCustomName, pk=pk)
    name = custom_name.name
    custom_name.delete()
    messages.success(request, f'Global Accessory Item "{name}" deleted successfully.')
    return redirect('accessories')


# ── Miscellaneous Report ────────────────────────────────────────────────

@login_required
def miscellaneous_report_view(request):
    """Only admin users can submit a miscellaneous report."""
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    if request.method == 'POST':
        form = MiscellaneousReportForm(request.POST, request.FILES)
        if form.is_valid():
            files = request.FILES.getlist('misc_files')
            if len(files) > 5:
                messages.error(request, 'You can upload a maximum of 5 files.')
                return redirect('miscellaneous_report')

            report = form.save(commit=False)
            report.created_by = request.user
            report.save()

            for f in files[:5]:
                MiscellaneousReportFile.objects.create(
                    report=report,
                    file_data=f.read(),
                    file_name=f.name,
                    content_type=f.content_type or 'application/octet-stream'
                )

            messages.success(request, 'Miscellaneous report submitted successfully!')
            return redirect('miscellaneous_report_list')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = MiscellaneousReportForm()

    return render(request, 'miscellaneous_form.html', {'form': form})


@login_required
def miscellaneous_report_list_view(request):
    """List all miscellaneous reports — admin sees all, others see their own."""
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if is_admin:
        reports = MiscellaneousReport.objects.prefetch_related('files').all()
    else:
        reports = MiscellaneousReport.objects.prefetch_related('files').filter(created_by=request.user)

    return render(request, 'miscellaneous_list.html', {
        'reports': reports,
        'is_admin': is_admin,
    })


@login_required
def serve_misc_file(request, file_id):
    """Serve a MiscellaneousReportFile. Images are served inline; other files are attachments."""
    f = get_object_or_404(MiscellaneousReportFile, pk=file_id)
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin and f.report.created_by != request.user:
        raise Http404
    response = HttpResponse(bytes(f.file_data), content_type=f.content_type)
    
    is_image = f.content_type.startswith('image/') or f.file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp', '.heic'))
    if is_image:
        response['Content-Disposition'] = f'inline; filename="{f.file_name}"'
    else:
        response['Content-Disposition'] = f'attachment; filename="{f.file_name}"'
    return response


@login_required
def delete_miscellaneous_report(request, pk):
    """Delete a miscellaneous report — admin or owner only."""
    report = get_object_or_404(MiscellaneousReport, pk=pk)
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin and report.created_by != request.user:
        messages.error(request, 'You do not have permission to delete this report.')
        return redirect('miscellaneous_report_list')
    report.delete()
    messages.success(request, 'Miscellaneous report deleted successfully.')
    return redirect('miscellaneous_report_list')


@login_required
def edit_miscellaneous_report(request, pk):
    """Only admin users can edit a miscellaneous report."""
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    report = get_object_or_404(MiscellaneousReport, pk=pk)

    if request.method == 'POST':
        form = MiscellaneousReportForm(request.POST, request.FILES, instance=report)
        if form.is_valid():
            delete_file_id = request.GET.get('delete_file')
            if delete_file_id:
                file_obj = get_object_or_404(MiscellaneousReportFile, pk=delete_file_id, report=report)
                file_obj.delete()
                messages.success(request, 'Attached file deleted successfully.')
                return redirect('edit_miscellaneous_report', pk=report.id)

            files = request.FILES.getlist('misc_files')
            current_files_count = report.files.count()
            if current_files_count + len(files) > 5:
                messages.error(request, f'You can upload a maximum of 5 files. You currently have {current_files_count} files.')
                return redirect('edit_miscellaneous_report', pk=report.id)

            updated_report = form.save()

            for f in files:
                MiscellaneousReportFile.objects.create(
                    report=updated_report,
                    file_data=f.read(),
                    file_name=f.name,
                    content_type=f.content_type or 'application/octet-stream'
                )

            messages.success(request, 'Miscellaneous report updated successfully!')
            return redirect('miscellaneous_report_list')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        # If GET request has a delete_file parameter, handle it
        delete_file_id = request.GET.get('delete_file')
        if delete_file_id:
            file_obj = get_object_or_404(MiscellaneousReportFile, pk=delete_file_id, report=report)
            file_obj.delete()
            messages.success(request, 'Attached file deleted successfully.')
            return redirect('edit_miscellaneous_report', pk=report.id)

        form = MiscellaneousReportForm(instance=report)

    return render(request, 'miscellaneous_form.html', {
        'form': form,
        'report': report,
        'is_edit': True,
    })


from decimal import Decimal

@login_required
def print_report_view(request, report_type, pk):
    # Mapping report types to Django models and their image serve strings
    model_mapping = {
        'cutting': (CuttingReport, 'cutting'),
        'stitching': (StitchingReport, 'p4'),
        'jobwork': (JobWorkReport, 'jobwork'),
        'jobwork1': (JobWork1Report, 'jobwork1'),
        'embroidery': (EmbroideryReport, 'embroidery'),
        'printing': (PrintingReport, 'printing'),
        'singleneedle': (SingleneedleReport, 'singleneedle'),
        'sewing': (SewingReport, 'sewing'),
        'sewing1': (Sewing1Report, 'sewing1'),
        'finishing': (FinishingReport, 'finishing'),
        'miscellaneous': (MiscellaneousReport, 'misc'),
    }

    if report_type not in model_mapping:
        raise Http404("Report type not found")

    model_class, image_model_name = model_mapping[report_type]
    report = get_object_or_404(model_class, pk=pk)

    # Get all model fields in a key-value format (excluding technical/relation fields)
    excluded_fields = {
        'id', 'signature', 'signature_2', 'created_at', 'updated_at',
        'created_by', 'cutting_report', 'master_entry', 'rate_definition'
    }
    
    fields_list = []
    for field in report._meta.fields:
        if field.name in excluded_fields:
            continue
        
        val = getattr(report, field.name)
        if val is None:
            val_str = '—'
        elif isinstance(val, bool):
            val_str = 'Yes' if val else 'No'
        elif hasattr(val, 'strftime'):
            val_str = val.strftime('%d-%b-%Y')
        elif isinstance(val, (int, float, Decimal)):
            val_str = str(val)
        else:
            val_str = str(val)

        fields_list.append({
            'label': field.verbose_name.title() if hasattr(field, 'verbose_name') else field.name.replace('_', ' ').title(),
            'value': val_str
        })

    color_details = None
    if report_type == 'cutting' and hasattr(report, 'color_details'):
        color_details = report.color_details.all()

    photos = []
    files = []
    if report_type == 'miscellaneous':
        files = report.files.all() if hasattr(report, 'files') else []
    else:
        photos = report.photos.all() if hasattr(report, 'photos') else []

    context = {
        'report': report,
        'report_type': report_type,
        'report_type_title': report_type.replace('_', ' ').title(),
        'fields_list': fields_list,
        'color_details': color_details,
        'photos': photos,
        'files': files,
        'image_model_name': image_model_name,
    }

    return render(request, 'report_print_detail.html', context)


@login_required
def vendor_report_view(request):
    from collections import defaultdict
    from .models import MasterName, AccessoriesItemEntry

    selected_vendor = request.GET.get('vendor', '').strip()
    selected_article = request.GET.get('article', '').strip()
    search_q = request.GET.get('q', '').strip()

    # Query registered Vendor masters
    vendor_masters = MasterName.objects.filter(department='Vendor').order_by('name', 'article')
    
    registered_vendor_articles = defaultdict(list)
    all_vendors_set = set()
    all_articles_set = set()

    for vm in vendor_masters:
        all_vendors_set.add(vm.name)
        if vm.article:
            if vm.article not in registered_vendor_articles[vm.name]:
                registered_vendor_articles[vm.name].append(vm.article)
            all_articles_set.add(vm.article)

    # Build a lookup for vendor master photos: {(vendor_name, article_name): photo_url}
    from django.urls import reverse as url_reverse
    vendor_master_photos = {}
    for vm in vendor_masters:
        if vm.article and vm.photo:
            key = (vm.name, vm.article)
            if key not in vendor_master_photos:
                vendor_master_photos[key] = url_reverse('serve_vendor_photo', args=[vm.id])

    # Collect entries from AccessoriesItemEntry across columns A, B, C, D
    cols = ['a', 'b', 'c', 'd']
    entries = AccessoriesItemEntry.objects.select_related('record').all()

    vendor_tree = defaultdict(lambda: defaultdict(list))
    vendor_job_cards = defaultdict(set)
    vendor_totals = defaultdict(lambda: {'pcs': 0, 'articles': set()})

    for entry in entries:
        jc_no = entry.record.job_card_no
        item_name = entry.item_name
        total_pcs = entry.record.total_pcs

        for col in cols:
            v_name = getattr(entry, f'vendor_{col}', None)
            a_name = getattr(entry, f'article_{col}', None)
            qty = getattr(entry, f'qty_{col}', None)
            tot = getattr(entry, f'total_{col}', None)
            status = getattr(entry, f'status_{col}', '')
            has_photo = bool(getattr(entry, f'photo_data_{col}', None))

            if v_name:
                v_name = v_name.strip()
                art_key = a_name.strip() if a_name else 'Unspecified Article'

                all_vendors_set.add(v_name)
                if a_name:
                    all_articles_set.add(a_name.strip())

                # Apply Filters
                if selected_vendor and selected_vendor.lower() != v_name.lower():
                    continue
                if selected_article and selected_article.lower() != art_key.lower():
                    continue
                if search_q:
                    q_lower = search_q.lower()
                    if q_lower not in jc_no.lower() and q_lower not in art_key.lower() and q_lower not in item_name.lower():
                        continue

                item_dict = {
                    'entry_id': entry.id,
                    'job_card_no': jc_no,
                    'item_name': item_name,
                    'total_pcs': total_pcs,
                    'col': col.upper(),
                    'qty': qty,
                    'total': tot,
                    'status': status,
                    'has_photo': has_photo,
                    'updated_at': entry.record.updated_at,
                    'vendor_photo_url': vendor_master_photos.get((v_name, art_key), ''),
                }

                vendor_tree[v_name][art_key].append(item_dict)

    # Build report list structure
    display_vendors = sorted(list(all_vendors_set))
    if selected_vendor:
        display_vendors = [v for v in display_vendors if v.lower() == selected_vendor.lower()]

    vendor_report_list = []
    grand_total_articles = 0
    grand_total_jobs = 0
    grand_total_pcs = 0

    for v_name in display_vendors:
        articles_dict = vendor_tree[v_name]
        
        # Add registered articles if not already present
        registered_arts = registered_vendor_articles.get(v_name, [])
        for r_art in registered_arts:
            if r_art not in articles_dict and not selected_article and not search_q:
                articles_dict[r_art] = []

        articles_list = []
        vendor_job_set = set()
        vendor_pcs_sum = 0

        for art_name, jobs in articles_dict.items():
            if selected_article and selected_article.lower() != art_name.lower():
                continue
            
            art_job_set = set(j['job_card_no'] for j in jobs)
            vendor_job_set.update(art_job_set)
            art_pcs = sum((j['total'] or 0) for j in jobs)
            vendor_pcs_sum += art_pcs

            articles_list.append({
                'name': art_name,
                'jobs': jobs,
                'job_count': len(art_job_set),
                'total_entries': len(jobs),
                'total_pcs': art_pcs,
                'job_card_numbers': sorted(list(art_job_set))
            })

        articles_list.sort(key=lambda x: x['name'])
        
        if articles_list or not (selected_vendor or selected_article or search_q):
            grand_total_articles += len(articles_list)
            grand_total_jobs += len(vendor_job_set)
            grand_total_pcs += vendor_pcs_sum

            vendor_report_list.append({
                'name': v_name,
                'articles': articles_list,
                'article_count': len(articles_list),
                'job_card_count': len(vendor_job_set),
                'unique_job_cards': sorted(list(vendor_job_set)),
                'total_pcs': vendor_pcs_sum,
            })

    all_vendors_sorted = sorted(list(all_vendors_set))
    all_articles_sorted = sorted(list(all_articles_set))

    return render(request, 'vendor_report.html', {
        'vendor_report_list': vendor_report_list,
        'all_vendors': all_vendors_sorted,
        'all_articles': all_articles_sorted,
        'selected_vendor': selected_vendor,
        'selected_article': selected_article,
        'search_q': search_q,
        'total_vendors_count': len(vendor_report_list),
        'grand_total_articles': grand_total_articles,
        'grand_total_jobs': grand_total_jobs,
        'grand_total_pcs': grand_total_pcs,
    })




@login_required
def progress_tracker_view(request):
    "Pending task progress grouped by user level for admin."
    is_admin = request.user.is_superuser or (
        hasattr(request.user, 'profile') and request.user.profile.person_type == 'ADMIN'
    )
    if not is_admin:
        raise PermissionDenied('Only administrators can view the progress tracker.')

    base_qs = JobCardRequirement.objects.all().order_by('-date', '-id')
    departments = [
        {'id': 'cutting',     'label': 'Cutting',     'levels': 'P1/P2/P3', 'color': '#3b82f6', 'tasks': base_qs.filter(requires_cutting__gt=0,     is_cutting_done=False),     'field': 'cutting'},
        {'id': 'stitching',   'label': 'Stitching',   'levels': 'P4',       'color': '#ec4899', 'tasks': base_qs.filter(requires_stitching__gt=0,   is_stitching_done=False),   'field': 'stitching'},
        {'id': 'jobwork',     'label': 'Job Work',    'levels': 'P5',       'color': '#f59e0b', 'tasks': base_qs.filter(requires_jobwork__gt=0,      is_jobwork_done=False),     'field': 'jobwork'},
        {'id': 'finishing',   'label': 'Finishing',   'levels': 'P6',       'color': '#8b5cf6', 'tasks': base_qs.filter(requires_finishing__gt=0,   is_finishing_done=False),   'field': 'finishing'},
        {'id': 'embroidery',  'label': 'Embroidery',  'levels': 'P7',       'color': '#06b6d4', 'tasks': base_qs.filter(requires_embroidery__gt=0,  is_embroidery_done=False),  'field': 'embroidery'},
        {'id': 'printing',    'label': 'Printing',    'levels': 'P8',       'color': '#10b981', 'tasks': base_qs.filter(requires_printing__gt=0,    is_printing_done=False),    'field': 'printing'},
        {'id': 'singleneedle','label': 'Singleneedle','levels': 'P9',       'color': '#f97316', 'tasks': base_qs.filter(requires_singleneedle__gt=0,is_singleneedle_done=False),'field': 'singleneedle'},
        {'id': 'sewing',      'label': 'Sewing',      'levels': 'P10',      'color': '#a78bfa', 'tasks': base_qs.filter(requires_sewing__gt=0,      is_sewing_done=False),      'field': 'sewing'},
        {'id': 'jobwork1',    'label': 'Job Work 1',  'levels': 'P11',      'color': '#fbbf24', 'tasks': base_qs.filter(requires_jobwork1__gt=0,    is_jobwork1_done=False),    'field': 'jobwork1'},
        {'id': 'sewing1',     'label': 'Sewing 1',    'levels': 'P12',      'color': '#34d399', 'tasks': base_qs.filter(requires_sewing1__gt=0,     is_sewing1_done=False),     'field': 'sewing1'},
    ]
    for dept in departments:
        dept['count'] = dept['tasks'].count()
    total_pending = sum(d['count'] for d in departments)
    return render(request, 'progress_tracker.html', {
        'departments': departments,
        'total_pending': total_pending,
    })


# ── Team Chat (SlackTask-style) ──────────────────────────────────────────────

DEFAULT_CHAT_CHANNELS = [
    ('production', 'Production updates & job cards', False),
    ('cutting', 'Cutting department', False),
    ('stitching', 'Stitching department', False),
    ('finishing', 'Finishing department', False),
    ('general', 'General team discussion', False),
    ('admin-sync', 'Private admin sync', True),
]


def ensure_default_chat_channels():
    for name, desc, private in DEFAULT_CHAT_CHANNELS:
        ChatChannel.objects.get_or_create(
            slug=slugify(name),
            defaults={
                'name': name,
                'description': desc,
                'is_default': name == 'general',
                'is_private': private,
                'is_dm': False,
            },
        )


def user_can_access_channel(user, channel):
    if channel.is_dm:
        return channel.members.filter(pk=user.pk).exists()
    if channel.is_private:
        return user.is_superuser or (
            hasattr(user, 'profile') and user.profile.person_type == 'ADMIN'
        ) or channel.members.filter(pk=user.pk).exists()
    return True


def parse_task_natural_language(text):
    """Simple NLP: strip today/tomorrow/next week, !p1 priority, @username assignee."""
    original = text.strip()
    priority = 'p3'
    due = None
    assignee_name = None
    today = timezone.localdate()

    pr = re.search(r'!(p[1-4])\b', original, re.I)
    if pr:
        priority = pr.group(1).lower()
        original = re.sub(r'!(p[1-4])\b', '', original, flags=re.I)

    am = re.search(r'@([A-Za-z0-9_.-]+)', original)
    if am:
        assignee_name = am.group(1)
        original = re.sub(r'@([A-Za-z0-9_.-]+)', '', original, count=1)

    lower = original.lower()
    if re.search(r'\btoday\b', lower):
        due = today
        original = re.sub(r'\btoday\b', '', original, flags=re.I)
    elif re.search(r'\btomorrow\b', lower):
        due = today + timedelta(days=1)
        original = re.sub(r'\btomorrow\b', '', original, flags=re.I)
    elif re.search(r'\bnext\s+week\b', lower):
        due = today + timedelta(days=7)
        original = re.sub(r'\bnext\s+week\b', '', original, flags=re.I)
    elif re.search(r'\bfriday\b', lower):
        days = (4 - today.weekday()) % 7
        due = today + timedelta(days=days or 7)
        original = re.sub(r'\bfriday\b', '', original, flags=re.I)

    title = re.sub(r'\s+', ' ', original).strip(' -–|')
    return title or text.strip(), due, priority, assignee_name


def resolve_task_assignee(assignee_name, fallback_user):
    if assignee_name:
        u = User.objects.filter(username__iexact=assignee_name, is_active=True).first()
        if u:
            return u
    return fallback_user


def serialize_chat_task(task):
    return {
        'id': task.id,
        'task_key': task.task_key or f'tsk_{task.id}',
        'title': task.title,
        'due_date': task.due_date.isoformat() if task.due_date else None,
        'due_label': task.due_date.strftime('%d %b') if task.due_date else None,
        'priority': task.priority,
        'completed': task.completed,
        'channel_id': task.channel_id,
        'channel_name': task.channel.name if task.channel else None,
        'assignee': task.assignee.username if task.assignee else None,
        'created_by': task.created_by.username if task.created_by else None,
        'labels': [{'id': l.id, 'name': l.name, 'color': l.color} for l in task.labels.all()],
    }


def serialize_chat_message(msg, user=None):
    sender_name = msg.sender.username if msg.sender else 'Unknown'
    initial = (sender_name[:1] or '?').upper()
    reactions = {}
    for r in msg.reactions.all():
        reactions.setdefault(r.emoji, {'emoji': r.emoji, 'count': 0, 'mine': False, 'users': []})
        reactions[r.emoji]['count'] += 1
        reactions[r.emoji]['users'].append(r.user.username if r.user else '?')
        if user and r.user_id == user.id:
            reactions[r.emoji]['mine'] = True

    bookmarked = False
    if user is not None:
        bookmarked = msg.bookmarks.filter(user=user).exists()

    reply_to = None
    if msg.parent_id and msg.parent:
        p = msg.parent
        p_sender = p.sender.username if p.sender else 'Unknown'
        p_text = p.content if p.content else (p.task.title if p.task else 'Message')
        if len(p_text) > 80:
            p_text = p_text[:80] + '…'
        reply_to = {
            'id': p.id,
            'sender': p_sender,
            'content': p_text,
            'is_task': p.message_type == 'task',
            'task_key': p.task.task_key if p.task else None,
        }

    data = {
        'id': msg.id,
        'content': msg.content,
        'message_type': msg.message_type,
        'sender': sender_name,
        'sender_id': msg.sender_id,
        'sender_initial': initial,
        'channel_id': msg.channel_id,
        'created_at': timezone.localtime(msg.created_at).strftime('%I:%M %p').lstrip('0'),
        'created_full': timezone.localtime(msg.created_at).strftime('%d-%b-%Y %H:%M'),
        'timestamp_iso': msg.created_at.isoformat(),
        'is_pinned': msg.is_pinned,
        'edited_at': timezone.localtime(msg.edited_at).strftime('%I:%M %p').lstrip('0') if msg.edited_at else None,
        'bookmarked': bookmarked,
        'reactions': list(reactions.values()),
        'has_image': bool(msg.image_data),
        'image_name': msg.image_name or '',
        'image_size': msg.image_size or 0,
        'image_url': f'/api/chat/message/{msg.id}/image/' if msg.image_data else None,
        'task': serialize_chat_task(msg.task) if msg.task_id else None,
        'reply_to': reply_to,
    }
    return data


def get_or_create_dm(user_a, user_b):
    if user_a.id == user_b.id:
        return None
    ids = sorted([user_a.id, user_b.id])
    slug = f'dm-{ids[0]}-{ids[1]}'
    channel, created = ChatChannel.objects.get_or_create(
        slug=slug,
        defaults={
            'name': f'{user_a.username} · {user_b.username}',
            'is_dm': True,
            'is_private': True,
            'description': 'Direct message',
        },
    )
    if created or channel.members.count() < 2:
        channel.members.add(user_a, user_b)
        # Prefer other person's name for each viewer handled in API
    return channel


def channel_unread_count(channel, user):
    read = ChatChannelRead.objects.filter(channel=channel, user=user).first()
    last_id = read.last_read_id if read else 0
    return ChatMessage.objects.filter(channel=channel, id__gt=last_id).exclude(sender=user).count()


@login_required
def chat_view(request):
    ensure_default_chat_channels()
    user = request.user
    public_channels = list(ChatChannel.objects.filter(is_dm=False).order_by('name'))
    public_channels = [c for c in public_channels if user_can_access_channel(user, c)]

    # Ensure DMs list of other active users
    other_users = list(
        User.objects.filter(is_active=True).exclude(pk=user.pk).order_by('username')[:40]
    )

    dm_channels = list(
        ChatChannel.objects.filter(is_dm=True, members=user).prefetch_related('members')
    )
    for ch in dm_channels:
        other = next((m for m in ch.members.all() if m.id != user.id), None)
        if other:
            ch.name = other.username

    active_slug = request.GET.get('channel') or 'general'
    active = None
    for c in public_channels + dm_channels:
        if c.slug == active_slug:
            active = c
            break
    if not active and public_channels:
        active = public_channels[0]

    # Task counts
    my_tasks = ChatTask.objects.filter(
        Q(created_by=user) | Q(assignee=user)
    )
    today = timezone.localdate()
    task_counts = {
        'today': my_tasks.filter(completed=False, due_date=today).count(),
        'upcoming': my_tasks.filter(completed=False, due_date__gt=today).count(),
        'inbox': my_tasks.filter(completed=False).count(),
    }

    mention_users = [
        {'id': u.id, 'username': u.username, 'initial': (u.username[:1] or '?').upper()}
        for u in User.objects.filter(is_active=True).order_by('username')[:100]
    ]

    return render(request, 'chat.html', {
        'channels': public_channels,
        'dm_users': other_users,
        'dm_channels': dm_channels,
        'active_channel': active,
        'task_counts': task_counts,
        'workspace_name': 'FabricTrack',
        'mention_users_json': json.dumps(mention_users),
    })


@login_required
@require_http_methods(['GET'])
def chat_bootstrap_api(request):
    """Sidebar data: channels, DMs, task counts."""
    ensure_default_chat_channels()
    user = request.user
    channels = []
    for c in ChatChannel.objects.filter(is_dm=False).order_by('name'):
        if not user_can_access_channel(user, c):
            continue
        channels.append({
            'id': c.id,
            'name': c.name,
            'slug': c.slug,
            'description': c.description,
            'is_private': c.is_private,
            'is_dm': False,
            'unread': channel_unread_count(c, user),
        })

    dms = []
    for c in ChatChannel.objects.filter(is_dm=True, members=user).prefetch_related('members'):
        other = next((m for m in c.members.all() if m.id != user.id), None)
        dms.append({
            'id': c.id,
            'name': other.username if other else c.name,
            'slug': c.slug,
            'user_id': other.id if other else None,
            'is_dm': True,
            'unread': channel_unread_count(c, user),
            'initial': (other.username[:1] if other else '?').upper(),
        })

    # Users without existing DM (for sidebar stubs)
    existing_ids = {d['user_id'] for d in dms if d.get('user_id')}
    people = []
    # All active users for @mention autocomplete
    mention_users = []
    for u in User.objects.filter(is_active=True).order_by('username')[:100]:
        mention_users.append({
            'id': u.id,
            'username': u.username,
            'initial': (u.username[:1] or '?').upper(),
        })
        if u.id != user.id and u.id not in existing_ids:
            people.append({
                'id': u.id,
                'username': u.username,
                'initial': (u.username[:1] or '?').upper(),
            })

    today = timezone.localdate()
    my_tasks = ChatTask.objects.filter(Q(created_by=user) | Q(assignee=user), completed=False)
    return JsonResponse({
        'channels': channels,
        'dms': dms,
        'people': people,
        'users': mention_users,
        'task_counts': {
            'today': my_tasks.filter(due_date=today).count(),
            'upcoming': my_tasks.filter(due_date__gt=today).count(),
            'inbox': my_tasks.count(),
        },
        'me': {
            'id': user.id,
            'username': user.username,
            'initial': user.username[:1].upper(),
        },
    })


@login_required
@require_http_methods(['GET'])
def chat_messages_api(request, channel_id):
    channel = get_object_or_404(ChatChannel, pk=channel_id)
    if not user_can_access_channel(request.user, channel):
        return JsonResponse({'error': 'Forbidden'}, status=403)

    after_id = request.GET.get('after_id')
    qs = (
        ChatMessage.objects.filter(channel=channel)
        .select_related('sender', 'task', 'task__channel', 'task__assignee', 'task__created_by', 'parent', 'parent__sender', 'parent__task')
        .prefetch_related('reactions__user', 'bookmarks')
    )

    wait = request.GET.get('wait') == '1'
    if wait and after_id is not None and str(after_id).isdigit():
        after_id_int = int(after_id)
        deadline = time.time() + 18
        while time.time() < deadline:
            messages_list = list(qs.filter(id__gt=after_id_int).order_by('id')[:80])
            if messages_list:
                return JsonResponse({
                    'messages': [serialize_chat_message(m, request.user) for m in messages_list],
                    'channel_id': channel.id,
                })
            time.sleep(1.1)
        return JsonResponse({'messages': [], 'channel_id': channel.id})

    if after_id is not None and str(after_id).isdigit():
        messages_list = list(qs.filter(id__gt=int(after_id)).order_by('id')[:80])
    else:
        recent = list(qs.order_by('-id')[:100])
        messages_list = list(reversed(recent))
        # mark read
        if messages_list:
            ChatChannelRead.objects.update_or_create(
                channel=channel, user=request.user,
                defaults={'last_read_id': messages_list[-1].id},
            )

    return JsonResponse({
        'messages': [serialize_chat_message(m, request.user) for m in messages_list],
        'channel_id': channel.id,
        'channel_name': channel.name,
        'is_dm': channel.is_dm,
    })


@login_required
@require_POST
def chat_send_api(request, channel_id):
    channel = get_object_or_404(ChatChannel, pk=channel_id)
    if not user_can_access_channel(request.user, channel):
        return JsonResponse({'error': 'Forbidden'}, status=403)

    content_type = request.content_type or ''
    image_file = None
    if 'multipart/form-data' in content_type:
        content = (request.POST.get('content') or '').strip()
        image_file = request.FILES.get('image')
        as_task = request.POST.get('as_task') == '1'
        priority = request.POST.get('priority') or 'p3'
        reply_to_id = request.POST.get('reply_to_id')
    else:
        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except json.JSONDecodeError:
            payload = {}
        content = (payload.get('content') or '').strip()
        as_task = bool(payload.get('as_task'))
        priority = payload.get('priority') or 'p3'
        reply_to_id = payload.get('reply_to_id')

    parent_msg = None
    if reply_to_id and str(reply_to_id).isdigit():
        parent_msg = ChatMessage.objects.filter(pk=int(reply_to_id), channel=channel).first()

    # /task command
    task_cmd = re.match(r'^/task\s+(.+)$', content, re.I | re.S)
    if task_cmd or as_task:
        raw = task_cmd.group(1) if task_cmd else content
        title, due, pr, assignee_name = parse_task_natural_language(raw)
        if priority in ('p1', 'p2', 'p3', 'p4'):
            pr = priority
        if not title:
            return JsonResponse({'error': 'Task title required.'}, status=400)
        task = ChatTask.objects.create(
            title=title,
            created_by=request.user,
            assignee=resolve_task_assignee(assignee_name, request.user),
            channel=channel,
            due_date=due,
            priority=pr,
        )
        msg = ChatMessage.objects.create(
            channel=channel,
            sender=request.user,
            content=title,
            message_type=ChatMessage.TYPE_TASK,
            task=task,
            parent=parent_msg,
        )
        return JsonResponse({'message': serialize_chat_message(msg, request.user)}, status=201)

    if not content and not image_file:
        return JsonResponse({'error': 'Message cannot be empty.'}, status=400)
    if len(content) > 4000:
        return JsonResponse({'error': 'Message too long.'}, status=400)

    msg = ChatMessage(
        channel=channel,
        sender=request.user,
        content=content,
        message_type=ChatMessage.TYPE_TEXT,
        parent=parent_msg,
    )
    if image_file:
        data = image_file.read()
        if len(data) > 5 * 1024 * 1024:
            return JsonResponse({'error': 'Image max 5MB.'}, status=400)
        msg.image_data = data
        msg.image_name = image_file.name[:255]
        msg.image_content_type = getattr(image_file, 'content_type', None) or 'image/jpeg'
        msg.image_size = len(data)
    msg.save()
    return JsonResponse({'message': serialize_chat_message(msg, request.user)}, status=201)


@login_required
@require_POST
def chat_message_to_task_api(request, message_id):
    msg = get_object_or_404(ChatMessage.objects.select_related('channel'), pk=message_id)
    if not user_can_access_channel(request.user, msg.channel):
        return JsonResponse({'error': 'Forbidden'}, status=403)
    if msg.task_id:
        return JsonResponse({'message': serialize_chat_message(msg, request.user), 'task': serialize_chat_task(msg.task)})

    title, due, pr, assignee_name = parse_task_natural_language(msg.content or 'Task from message')
    task = ChatTask.objects.create(
        title=title or 'Task from chat',
        created_by=request.user,
        assignee=resolve_task_assignee(assignee_name, request.user),
        channel=msg.channel,
        due_date=due,
        priority=pr,
    )
    card = ChatMessage.objects.create(
        channel=msg.channel,
        sender=request.user,
        content=task.title,
        message_type=ChatMessage.TYPE_TASK,
        task=task,
    )
    return JsonResponse({
        'task': serialize_chat_task(task),
        'message': serialize_chat_message(card, request.user),
    }, status=201)


@login_required
@require_POST
def chat_react_api(request, message_id):
    msg = get_object_or_404(ChatMessage, pk=message_id)
    if not user_can_access_channel(request.user, msg.channel):
        return JsonResponse({'error': 'Forbidden'}, status=403)
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        payload = {}
    emoji = (payload.get('emoji') or '👍').strip()[:16]
    existing = ChatReaction.objects.filter(message=msg, user=request.user, emoji=emoji).first()
    if existing:
        existing.delete()
        toggled = False
    else:
        ChatReaction.objects.create(message=msg, user=request.user, emoji=emoji)
        toggled = True
    msg = ChatMessage.objects.prefetch_related('reactions__user', 'bookmarks').select_related(
        'sender', 'task', 'task__channel'
    ).get(pk=msg.pk)
    return JsonResponse({'toggled': toggled, 'message': serialize_chat_message(msg, request.user)})


@login_required
@require_POST
def chat_bookmark_api(request, message_id):
    msg = get_object_or_404(ChatMessage, pk=message_id)
    if not user_can_access_channel(request.user, msg.channel):
        return JsonResponse({'error': 'Forbidden'}, status=403)
    bm, created = ChatBookmark.objects.get_or_create(user=request.user, message=msg)
    if not created:
        bm.delete()
        bookmarked = False
    else:
        bookmarked = True
    return JsonResponse({'bookmarked': bookmarked})


@login_required
@require_POST
def chat_pin_api(request, message_id):
    msg = get_object_or_404(ChatMessage, pk=message_id)
    if not user_can_access_channel(request.user, msg.channel):
        return JsonResponse({'error': 'Forbidden'}, status=403)
    msg.is_pinned = not msg.is_pinned
    msg.save(update_fields=['is_pinned'])
    return JsonResponse({'is_pinned': msg.is_pinned})


@login_required
@require_POST
def chat_edit_api(request, message_id):
    msg = get_object_or_404(ChatMessage, pk=message_id)
    if not user_can_access_channel(request.user, msg.channel):
        return JsonResponse({'error': 'Forbidden'}, status=403)
    if msg.sender_id != request.user.id and not request.user.is_superuser:
        return JsonResponse({'error': 'Cannot edit others messages'}, status=403)
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        payload = {}
    content = (payload.get('content') or '').strip()
    if not content:
        return JsonResponse({'error': 'Content required'}, status=400)
    msg.content = content
    msg.edited_at = timezone.now()
    msg.save(update_fields=['content', 'edited_at'])
    return JsonResponse({'id': msg.id, 'content': msg.content, 'edited_at': msg.edited_at.isoformat()})


@login_required
@require_POST
def chat_message_delete_api(request, message_id):
    msg = get_object_or_404(ChatMessage, pk=message_id)
    if not user_can_access_channel(request.user, msg.channel):
        return JsonResponse({'error': 'Forbidden'}, status=403)
    if msg.sender_id != request.user.id and not request.user.is_superuser:
        return JsonResponse({'error': 'Cannot delete message created by another user'}, status=403)

    task_id = msg.task_id
    msg.delete()
    if task_id and not ChatMessage.objects.filter(task_id=task_id).exists():
        ChatTask.objects.filter(pk=task_id).delete()

    return JsonResponse({'success': True, 'message_id': message_id})


@login_required
@require_http_methods(['GET'])
def chat_message_image_api(request, message_id):
    msg = get_object_or_404(ChatMessage, pk=message_id)
    if not user_can_access_channel(request.user, msg.channel):
        raise Http404()
    if not msg.image_data:
        raise Http404()
    return HttpResponse(bytes(msg.image_data), content_type=msg.image_content_type or 'image/jpeg')


@login_required
@require_http_methods(['GET', 'POST'])
def chat_tasks_api(request):
    user = request.user
    if request.method == 'POST':
        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except json.JSONDecodeError:
            payload = {}
        raw = (payload.get('title') or '').strip()
        if not raw:
            return JsonResponse({'error': 'Title required'}, status=400)
        title, due, pr, assignee_name = parse_task_natural_language(raw)
        channel_id = payload.get('channel_id')
        channel = None
        if channel_id:
            channel = ChatChannel.objects.filter(pk=channel_id).first()
        task = ChatTask.objects.create(
            title=title,
            created_by=user,
            assignee=resolve_task_assignee(assignee_name, user),
            channel=channel,
            due_date=due or None,
            priority=payload.get('priority') or pr,
        )
        label_ids = payload.get('label_ids', [])
        if label_ids:
            task.labels.set(ChatTaskLabel.objects.filter(id__in=label_ids))
        if channel and payload.get('post_to_channel'):
            ChatMessage.objects.create(
                channel=channel,
                sender=user,
                content=task.title,
                message_type=ChatMessage.TYPE_TASK,
                task=task,
            )
        return JsonResponse({'task': serialize_chat_task(task)}, status=201)

    scope = request.GET.get('scope', 'inbox')
    channel_id = request.GET.get('channel_id')
    today = timezone.localdate()
    qs = ChatTask.objects.select_related('channel', 'assignee', 'created_by').prefetch_related('labels')
    if channel_id and str(channel_id).isdigit():
        qs = qs.filter(channel_id=int(channel_id))
    else:
        qs = qs.filter(Q(created_by=user) | Q(assignee=user))

    if scope == 'today':
        qs = qs.filter(completed=False, due_date=today)
    elif scope == 'upcoming':
        qs = qs.filter(completed=False, due_date__gt=today)
    elif scope == 'done':
        qs = qs.filter(completed=True)
    elif scope == 'channel':
        qs = qs.filter(completed=False)
    else:
        qs = qs.filter(completed=False)

    tasks = [serialize_chat_task(t) for t in qs.order_by('due_date', '-created_at')[:100]]
    return JsonResponse({'tasks': tasks})


@login_required
@require_POST
def chat_task_toggle_api(request, task_id):
    task = get_object_or_404(ChatTask, pk=task_id)
    task.completed = not task.completed
    task.completed_at = timezone.now() if task.completed else None
    task.save(update_fields=['completed', 'completed_at'])
    return JsonResponse({'task': serialize_chat_task(task)})


@login_required
@require_POST
def chat_task_delete_api(request, task_id):
    task = get_object_or_404(ChatTask, pk=task_id)
    if task.created_by_id != request.user.id and task.assignee_id != request.user.id and not request.user.is_superuser:
        return JsonResponse({'error': 'Cannot delete task created by another user'}, status=403)

    ChatMessage.objects.filter(task=task).delete()
    task.delete()
    return JsonResponse({'success': True, 'task_id': task_id})



@login_required
@require_http_methods(['GET', 'POST'])
def chat_labels_api(request):
    user = request.user
    if request.method == 'POST':
        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except json.JSONDecodeError:
            payload = {}
        name = (payload.get('name') or '').strip()
        color = payload.get('color', '#a855f7')
        if not name:
            return JsonResponse({'error': 'Name required'}, status=400)
        label, created = ChatTaskLabel.objects.get_or_create(
            name__iexact=name,
            defaults={'name': name, 'color': color, 'created_by': user}
        )
        return JsonResponse({'id': label.id, 'name': label.name, 'color': label.color}, status=201)

    labels = ChatTaskLabel.objects.all().order_by('name')
    return JsonResponse({'labels': [{'id': l.id, 'name': l.name, 'color': l.color} for l in labels]})


@login_required
@require_http_methods(['GET', 'POST', 'DELETE'])
def chat_task_labels_api(request, task_id):
    task = get_object_or_404(ChatTask, pk=task_id)
    if request.method == 'GET':
        labels = [{'id': l.id, 'name': l.name, 'color': l.color} for l in task.labels.all()]
        return JsonResponse({'labels': labels})

    if request.method == 'POST':
        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except json.JSONDecodeError:
            payload = {}
        label_ids = payload.get('label_ids', [])
        task.labels.set(ChatTaskLabel.objects.filter(id__in=label_ids))
        return JsonResponse({'labels': [{'id': l.id, 'name': l.name, 'color': l.color} for l in task.labels.all()]})

    if request.method == 'DELETE':
        label_id = request.GET.get('label_id')
        if label_id:
            task.labels.remove(label_id)
        return JsonResponse({'labels': [{'id': l.id, 'name': l.name, 'color': l.color} for l in task.labels.all()]})


@login_required
@require_POST
def chat_open_dm_api(request):
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        payload = {}
    uid = payload.get('user_id')
    other = get_object_or_404(User, pk=uid, is_active=True)
    channel = get_or_create_dm(request.user, other)
    if not channel:
        return JsonResponse({'error': 'Invalid DM'}, status=400)
    return JsonResponse({
        'channel': {
            'id': channel.id,
            'name': other.username,
            'slug': channel.slug,
            'is_dm': True,
        }
    })


@login_required
@require_http_methods(['GET'])
def chat_saved_api(request):
    kind = request.GET.get('kind', 'bookmarks')
    user = request.user
    if kind == 'pinned':
        msgs = (
            ChatMessage.objects.filter(is_pinned=True)
            .select_related('sender', 'channel', 'task')
            .prefetch_related('reactions__user', 'bookmarks')
            .order_by('-created_at')[:50]
        )
        msgs = [m for m in msgs if user_can_access_channel(user, m.channel)]
    else:
        bms = (
            ChatBookmark.objects.filter(user=user)
            .select_related('message__sender', 'message__channel', 'message__task')
            .prefetch_related('message__reactions__user', 'message__bookmarks')
            .order_by('-created_at')[:50]
        )
        msgs = [b.message for b in bms]
    return JsonResponse({
        'messages': [serialize_chat_message(m, user) for m in msgs]
    })

